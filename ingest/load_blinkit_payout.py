"""Load Blinkit monthly payout sheet into the orders table.

Processes one payout folder (e.g. blinkit_reports/payout sheets/payout_sheet_2026-03/).
Reads 'Forward & Return Cancelled Orders.xlsx':
  - Forward Orders tab   (header row 5, data row 7+): upserts FULFILLED orders
  - Cancelled/Returned tab (header row 5, data row 7+): marks matching orders SALE_RETURN

Forward Orders columns (row 5):
  0  S.No.          10  Customer City     20  Order Status      30  CGST Value
  1  Invoice ID     11  Customer State    21  HSN Code          31  SGST Value
  2  Order ID       12  IRN               22  Quantity          32  CESS Value
  3  Order Type     13  Item ID           23  MRP (Rs)          33  Total Tax
  4  Order Date     14  Product Name      24  Selling Price     34  Total Gross Bill Amount
  5  Customer Name  15  Variant Desc      25  IGST %            ...
  6  GST Name       16  Business Cat      26  CGST %
  7  GST Number     17  L0 Category       27  SGST %
  8  Supply State   18  L1 Category       28  CESS %
  9  State GST      19  L2 Category       29  IGST Value

Cancelled/Returned columns (row 5):
  0  S.No.               5  Return Order Date    10  Supply State
  1  Forward Invoice ID  6  Customer Name        11  State GST
  2  Forward Inv Date    7  GST Number           12  Customer City
  3  Return Invoice ID   8  GST Name             13  Customer State
  4  Return Order ID     9  Supply City          15  Item ID

Usage:
  python ingest/load_blinkit_payout.py --folder <payout_sheet_YYYY-MM path> [--env dev|prod] [--dry-run]
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from ingest.utils import resolve_blinkit_sku, get_sku_cogs_at_date

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

BLK_CHANNEL_ID = 4
_PAYOUT_FILE = "Forward & Return Cancelled Orders.xlsx"
_UPSERT_BATCH = 100

# Forward Orders column indices (0-based, header row 5, data from row 7)
_COL_INVOICE_ID   = 1
_COL_ORDER_ID     = 2
_COL_ORDER_DATE   = 4
_COL_SUPPLY_STATE = 8   # "Supply State" — used for state-level FIFO lot consumption
_COL_CUST_CITY    = 10
_COL_CUST_STATE   = 11
_COL_ITEM_ID      = 13
_COL_ORDER_STATUS = 20
_COL_QTY          = 22
_COL_MRP          = 23
_COL_SP           = 24
_COL_GROSS_BILL   = 34


def _parse_date(val) -> date:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d %B %Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Cannot parse date: {val!r}")


def _load_forward_orders(
    ws, source_file: str
) -> tuple[list[dict], list[dict], dict[str, dict[str, str]], set[tuple[str, str]], dict[str, str]]:
    """Parse Forward Orders sheet.

    Returns:
        delivered_rows    — FULFILLED order dicts (upserted with ignore_duplicates=True)
        cancelled_rows    — CANCELLED order dicts (force-upserted to fix any prior mis-tagging)
        inv_map           — {invoice_id: {str(item_id): order_id}} for linking returns
        cancelled_inv_ids — {(invoice_id, str(item_id))} of cancelled line items
        supply_state_map  — {order_id: supply_state} for FIFO lot consumption
    """
    delivered_rows: list[dict] = []
    cancelled_rows: list[dict] = []
    inv_map: dict[str, dict[str, str]] = {}
    cancelled_inv_ids: set[tuple[str, str]] = set()
    supply_state_map: dict[str, str] = {}

    for raw in ws.iter_rows(min_row=7, values_only=True):
        if raw[0] is None:
            continue

        invoice_id      = str(raw[_COL_INVOICE_ID])   if raw[_COL_INVOICE_ID]   else None
        order_id_raw    = raw[_COL_ORDER_ID]
        order_date_raw  = raw[_COL_ORDER_DATE]
        supply_state    = str(raw[_COL_SUPPLY_STATE]).strip() if raw[_COL_SUPPLY_STATE] else None
        blinkit_item_id = raw[_COL_ITEM_ID]
        order_status    = str(raw[_COL_ORDER_STATUS]).strip().upper() if raw[_COL_ORDER_STATUS] else "FULFILLED"

        if not order_id_raw or not blinkit_item_id:
            continue

        try:
            order_date = _parse_date(order_date_raw)
        except ValueError:
            logger.warning("Unparseable date %r in Forward Orders — skipping", order_date_raw)
            continue

        sku_id = resolve_blinkit_sku(int(blinkit_item_id), order_date)
        if sku_id is None:
            continue

        order_id = f"BLK-{order_id_raw}-{sku_id}"

        if invoice_id:
            inv_map.setdefault(invoice_id, {})[str(blinkit_item_id)] = order_id

        mrp = float(raw[_COL_MRP]) if raw[_COL_MRP] is not None else None
        sp  = float(raw[_COL_SP])  if raw[_COL_SP]  is not None else None
        qty = int(raw[_COL_QTY])   if raw[_COL_QTY]  is not None else 1

        base = {
            "order_id":          order_id,
            "channel_id":        BLK_CHANNEL_ID,
            "order_date":        order_date.isoformat(),
            "sku_id":            sku_id,
            "quantity":          qty,
            "mrp":               mrp,
            "city":              raw[_COL_CUST_CITY],
            "state":             raw[_COL_CUST_STATE],
            "fulfillment_type":  "SOR",
            "platform_order_id": str(order_id_raw),
            "source_file":       source_file,
        }

        if order_status == "CANCELLED":
            if invoice_id:
                cancelled_inv_ids.add((invoice_id, str(blinkit_item_id)))
            cancelled_rows.append({**base, "status": "CANCELLED",
                                   "selling_price": None, "gross_value": None,
                                   "discount_pct": None, "cogs": None})
        else:
            gv = float(raw[_COL_GROSS_BILL]) if raw[_COL_GROSS_BILL] is not None else None
            discount_pct = round((mrp - sp) / mrp * 100, 2) if mrp and sp and mrp > 0 else None
            delivered_rows.append({**base, "status": "FULFILLED",
                                   "selling_price": sp, "gross_value": gv,
                                   "discount_pct": discount_pct,
                                   "cogs": get_sku_cogs_at_date(sku_id, order_date)})
            if supply_state:
                supply_state_map[order_id] = supply_state

    return delivered_rows, cancelled_rows, inv_map, cancelled_inv_ids, supply_state_map


def _load_returns(
    ws, inv_map: dict[str, dict[str, str]], cancelled_inv_ids: set[tuple[str, str]]
) -> list[dict]:
    """Parse Cancelled/Returned Orders sheet.

    Skips rows whose (forward_invoice_id, item_id) belongs to a CANCELLED forward order.
    Uses item_id (col 15) to resolve the exact line item for multi-SKU orders.
    Returns list of {order_id, return_date} update dicts for genuine SALE_RETURN orders.
    """
    updates: list[dict] = []
    unmatched = 0

    for raw in ws.iter_rows(min_row=7, values_only=True):
        if raw[0] is None:
            continue

        fwd_invoice_id = str(raw[1]) if raw[1] else None
        return_date_raw = raw[5]
        return_item_id = str(raw[15]) if len(raw) > 15 and raw[15] else None

        if not fwd_invoice_id:
            continue

        if (fwd_invoice_id, return_item_id) in cancelled_inv_ids:
            continue  # cancellation confirmation, not a customer return

        order_id = inv_map.get(fwd_invoice_id, {}).get(return_item_id)
        if order_id is None:
            logger.warning(
                "Forward Invoice ID %s / item %s not found in Forward Orders — skipping return",
                fwd_invoice_id, return_item_id,
            )
            unmatched += 1
            continue

        try:
            return_date = _parse_date(return_date_raw)
        except ValueError:
            return_date = None

        updates.append({"order_id": order_id, "return_date": return_date})

    if unmatched:
        logger.warning("%d return rows had no matching Forward Order", unmatched)

    return updates


def _apply_lot_cogs(db, delivered_rows: list[dict],
                    supply_state_map: dict[str, str]) -> tuple[int, int]:
    """
    For each FULFILLED order not yet lot-finalized: consume sku_cogs_lots FIFO
    and update orders.cogs + lot_cogs_finalized.

    Returns (finalized_count, skipped_already_done).
    """
    from tcb.inventory import consume_sor_sale

    # Fetch which order_ids are already finalized to avoid double-consumption.
    all_ids = [r["order_id"] for r in delivered_rows]
    already_done: set[str] = set()
    for i in range(0, len(all_ids), 500):
        rows = (db.table("orders")
                  .select("order_id")
                  .in_("order_id", all_ids[i : i + 500])
                  .eq("lot_cogs_finalized", True)
                  .execute().data)
        already_done.update(r["order_id"] for r in rows)

    finalized = 0
    skipped   = 0
    for row in delivered_rows:
        oid = row["order_id"]
        if oid in already_done:
            skipped += 1
            continue

        supply_state = supply_state_map.get(oid)
        result = consume_sor_sale(
            sku_id=row["sku_id"],
            qty=row["quantity"],
            channel_id=BLK_CHANNEL_ID,
            supply_state=supply_state,
        )

        update: dict = {"lot_cogs_finalized": True}
        if result is not None:
            unit_cogs, lot_id = result
            update["cogs"] = round(unit_cogs * row["quantity"], 2)
            if lot_id is not None:
                update["lot_id"] = lot_id

        db.table("orders").update(update).eq("order_id", oid).eq("channel_id", BLK_CHANNEL_ID).execute()
        finalized += 1

    return finalized, skipped


def _flag_daily_not_in_payout(db, delivered_rows: list[dict]) -> int:
    """
    Warn about orders loaded from daily sales files (lot_cogs_finalized=False)
    whose order_date falls within the payout period but are absent from the payout.
    Returns count of flagged orders.
    """
    if not delivered_rows:
        return 0

    payout_ids = {r["order_id"] for r in delivered_rows}
    dates = [r["order_date"] for r in delivered_rows]
    min_date, max_date = min(dates), max(dates)

    db_rows = (db.table("orders")
                 .select("order_id")
                 .eq("channel_id", BLK_CHANNEL_ID)
                 .eq("status", "FULFILLED")
                 .eq("lot_cogs_finalized", False)
                 .gte("order_date", min_date)
                 .lte("order_date", max_date)
                 .execute().data)

    missing = [r["order_id"] for r in db_rows if r["order_id"] not in payout_ids]
    if missing:
        logger.warning(
            "%d order(s) in DB for %s–%s are FULFILLED but absent from this payout — "
            "review manually: %s",
            len(missing), min_date, max_date, missing[:10],
        )
    return len(missing)


def load_payout_folder(folder: Path, db, dry_run: bool = False) -> tuple[int, int, int]:
    """Load one payout folder. Returns (new_orders, skipped_orders, returns_marked)."""
    payout_file = folder / _PAYOUT_FILE
    if not payout_file.exists():
        raise FileNotFoundError(f"{_PAYOUT_FILE} not found in {folder}")

    wb = openpyxl.load_workbook(str(payout_file), read_only=True, data_only=True)
    ws_fwd = wb["Forward Orders"]
    ws_ret = wb["Cancelled or Returned Orders"]

    delivered_rows, cancelled_rows, inv_map, cancelled_inv_ids, supply_state_map = \
        _load_forward_orders(ws_fwd, payout_file.name)
    return_updates = _load_returns(ws_ret, inv_map, cancelled_inv_ids)

    wb.close()

    if dry_run:
        print(
            f"[DRY RUN] Forward Orders: {len(delivered_rows)} delivered, "
            f"{len(cancelled_rows)} cancelled | Returns: {len(return_updates)} matched | "
            f"Supply states captured: {len(supply_state_map)}"
        )
        return len(delivered_rows) + len(cancelled_rows), 0, len(return_updates)

    # Upsert delivered orders — existing rows not overwritten (idempotent)
    new_count = 0
    for i in range(0, len(delivered_rows), _UPSERT_BATCH):
        result = (
            db.table("orders")
            .upsert(delivered_rows[i : i + _UPSERT_BATCH],
                    on_conflict="order_id,channel_id",
                    ignore_duplicates=True)
            .execute()
        )
        new_count += len(result.data) if result.data else 0

    skipped = len(delivered_rows) - new_count

    # Force-upsert cancelled orders
    for i in range(0, len(cancelled_rows), _UPSERT_BATCH):
        result = (
            db.table("orders")
            .upsert(cancelled_rows[i : i + _UPSERT_BATCH],
                    on_conflict="order_id,channel_id",
                    ignore_duplicates=False)
            .execute()
        )
        new_count += len(result.data) if result.data else 0

    # Consume FIFO lots and finalize COGS for all delivered orders
    finalized, already_done = _apply_lot_cogs(db, delivered_rows, supply_state_map)
    logger.info("Lot COGS finalized: %d | already done: %d", finalized, already_done)

    # Flag daily-loaded orders absent from this payout
    _flag_daily_not_in_payout(db, delivered_rows)

    # Mark genuine customer returns
    returns_marked = 0
    for upd in return_updates:
        result = (
            db.table("orders")
            .update({"status": "SALE_RETURN",
                     "return_date": upd["return_date"].isoformat() if upd["return_date"] else None})
            .eq("order_id", upd["order_id"])
            .eq("channel_id", BLK_CHANNEL_ID)
            .execute()
        )
        if result.data:
            returns_marked += 1

    return new_count, skipped, returns_marked


def main() -> None:
    parser = argparse.ArgumentParser(description="Load Blinkit payout sheet → orders table")
    parser.add_argument("--folder", required=True, help="Path to payout_sheet_YYYY-MM folder")
    parser.add_argument("--env",     choices=["dev", "prod"], default="prod")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    os.environ.setdefault("TCB_ENV", args.env)

    from tcb.db import get_client
    db = get_client()

    folder = Path(args.folder)
    new, skipped, returns = load_payout_folder(folder, db, dry_run=args.dry_run)

    tag = " [DRY RUN]" if args.dry_run else ""
    print(f"{folder.name}{tag}: {new} new orders | {skipped} already existed | {returns} returns marked")


if __name__ == "__main__":
    main()
