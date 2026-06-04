"""Monthly FC order reconciliation, return tagging, and reason enrichment.

Two inputs -- place both files in --folder before running:

  1. FC Excel reports (Delivered + Shipped .xlsx files downloaded from FC portal)
     -> Gap detection only: flag any order in the file that is NOT found in DB.
        These are App misses -- do NOT auto-insert; review and punch manually.
     -> No updates, no inserts from the Excel files.

  2. Google Sheet 'First Cry' tab (auto-fetched as xlsx)
     Sheet: https://docs.google.com/spreadsheets/d/1_rwqWJJt0aOuLwJnxNyKlCYy21C8RO0rNXI4U1MYKYc
     -> Tags FULFILLED orders as RTO or SALE_RETURN (from 'SR / RTO' column)
     -> Sets return_date (from 'SR/RTO Date' column)
     -> Enriches return_reason, return_responsible, return_customer_verbatim
     -> Enriches city / state on DB orders where those fields are currently NULL

All rows in the First Cry tab are treated as confirmed returns (no 'Solved' filter).

Cross-checks:
  A. Excel orders with Is SR=Yes that have no entry in the Google Sheet
     -> return reported in file, not yet added to your tracking sheet
  B. DB FC orders as RTO/SALE_RETURN (3-month window) not in Google Sheet
     -> return tagged in DB but missing from your tracking sheet
  C. Google Sheet rows whose Order No is not found in DB at all
     -> order in sheet but missing from DB entirely -- likely an App miss

Usage:
  python ingest/load_fc_sales.py --folder data/firstcry/manual [--env dev|prod] [--dry-run]
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import requests

sys.path.insert(0, str(Path(__file__).parent.parent))

from ingest.utils import resolve_fc_sku

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

FC_CHANNEL_ID = 6
SHEET_ID      = "1_rwqWJJt0aOuLwJnxNyKlCYy21C8RO0rNXI4U1MYKYc"
SHEET_URL     = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
FC_SHEET_NAME = "First Cry"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _rolling_window() -> tuple[date, date]:
    """3 full calendar months before current month."""
    today = date.today()
    first_current = today.replace(day=1)
    last_prev = first_current - timedelta(days=1)
    m = first_current
    for _ in range(3):
        m = (m - timedelta(days=1)).replace(day=1)
    return m, last_prev


def _clean(val) -> str:
    if val is None:
        return "Unknown"
    s = str(val).strip()
    return s if s and s.lower() not in ("na", "n/a", "nan") else "Unknown"


def _parse_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("na", "n/a", ""):
        return None
    for fmt in ("%d-%b %Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _shipping_ref_str(val) -> str | None:
    """Normalise Shipping Ref (can be float like 11166963297.0) to string."""
    if val is None:
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


# ── Part 1: Excel report loading ───────────────────────────────────────────────

def _load_excel_orders(folder: Path) -> list[dict]:
    """Load all .xlsx files in folder. Returns list of order dicts."""
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        logger.warning("No .xlsx files found in %s", folder)
        return []

    all_rows: list[dict] = []
    for f in files:
        logger.info("Loading Excel file: %s", f.name)
        df = pd.read_excel(f)
        for _, r in df.iterrows():
            row = r.to_dict()
            product_id = str(row.get("Product ID", "")).strip()
            sku_id     = resolve_fc_sku(product_id)
            order_id   = str(row.get("OrderID", "")).strip()
            if not order_id or not sku_id:
                continue
            shipping_ref = _shipping_ref_str(row.get("Shipping Ref. No."))
            is_sr        = str(row.get("Is SR", "No")).strip().lower() == "yes"
            order_date   = _parse_date(row.get("Order Date"))
            all_rows.append({
                "order_id":     order_id,
                "shipping_ref": shipping_ref,
                "sku_id":       sku_id,
                "is_sr":        is_sr,
                "order_date":   order_date,
                "source_file":  f.name,
            })
        logger.info("  -> %d rows parsed from %s", len(df), f.name)

    logger.info("Excel total: %d orders across %d file(s)", len(all_rows), len(files))
    return all_rows


# ── Part 2: Google Sheet fetching ──────────────────────────────────────────────

def _fetch_sheet_entries() -> list[dict]:
    """Download spreadsheet xlsx and parse the 'First Cry' tab."""
    logger.info("Fetching Google Sheet...")
    resp = requests.get(SHEET_URL, allow_redirects=True, timeout=30)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    if FC_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{FC_SHEET_NAME}' not found in workbook. "
            f"Available: {wb.sheetnames}"
        )
    ws   = wb[FC_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col(row, name):
        try:
            return row[headers.index(name)]
        except ValueError:
            return None

    entries: list[dict] = []
    for raw in rows[1:]:
        if all(v is None for v in raw):
            continue
        order_no     = str(col(raw, "Order No") or "").strip()
        sku          = str(col(raw, "SKU")      or "").strip()
        sr_rto       = str(col(raw, "SR / RTO") or "").strip().upper()
        shipping_ref = _shipping_ref_str(col(raw, "Shipping Ref"))
        if not order_no or not sku or sr_rto not in ("SR", "RTO"):
            continue
        entries.append({
            "order_no":    order_no,
            "shipping_ref": shipping_ref,
            "sku":         sku,
            "sr_rto":      sr_rto,           # "SR" or "RTO"
            "return_date": _parse_date(col(raw, "SR/RTO Date")),
            "verbatim":    _clean(col(raw, "Customer Verbatim")),
            "reason":      _clean(col(raw, "Reason")),
            "responsible": _clean(col(raw, "Responsible")),
            "city":        str(col(raw, "City")  or "").strip() or None,
            "state":       str(col(raw, "State") or "").strip() or None,
        })

    logger.info("Google Sheet 'First Cry': %d confirmed return entries", len(entries))
    return entries


# ── Part 3: DB helpers ─────────────────────────────────────────────────────────

def _load_db_fc_orders(db) -> dict[str, dict]:
    """Return {platform_order_id: db_row} for all FC orders."""
    rows = (db.table("orders")
              .select("order_id, platform_order_id, sku_id, status, city, state, channel_id")
              .eq("channel_id", FC_CHANNEL_ID)
              .execute().data) or []
    return {r["platform_order_id"]: r for r in rows if r.get("platform_order_id")}


def _find_db_order(order_id: str, shipping_ref: str | None,
                   db_by_id: dict[str, dict]) -> dict | None:
    """Look up a DB order by OrderID first, then Shipping Ref fallback."""
    row = db_by_id.get(order_id)
    if row is None and shipping_ref:
        row = db_by_id.get(shipping_ref)
    return row


# ── Main run ───────────────────────────────────────────────────────────────────

def run(folder: Path, dry_run: bool = False) -> None:
    from tcb.db import get_client
    db = get_client()

    window_start, window_end = _rolling_window()
    logger.info("3-month window for cross-checks: %s to %s", window_start, window_end)

    # Load all inputs
    excel_orders  = _load_excel_orders(folder)
    sheet_entries = _fetch_sheet_entries()
    db_by_id      = _load_db_fc_orders(db)
    logger.info("DB: %d FC orders total", len(db_by_id))

    # Sheet order_nos for quick lookup in cross-checks
    sheet_order_nos: set[str] = {e["order_no"] for e in sheet_entries}
    sheet_refs: set[str]      = {e["shipping_ref"] for e in sheet_entries if e["shipping_ref"]}

    # ── Part 1: Excel gap detection ───────────────────────────────────────────
    not_in_db:   list[dict] = []
    excel_returns_not_in_sheet: list[dict] = []

    for row in excel_orders:
        db_row = _find_db_order(row["order_id"], row["shipping_ref"], db_by_id)

        if db_row is None:
            not_in_db.append(row)
            continue

        # Order is in DB — check if it's a reported return but not yet in sheet
        if row["is_sr"]:
            in_sheet = (row["order_id"] in sheet_order_nos
                        or (row["shipping_ref"] and row["shipping_ref"] in sheet_refs))
            if not in_sheet:
                excel_returns_not_in_sheet.append(row)

    # ── Part 2: Google Sheet — status update + enrichment ─────────────────────
    tagged_rto = tagged_sr = already_tagged = enriched = city_state_enriched = 0

    for entry in sheet_entries:
        db_row = _find_db_order(entry["order_no"], entry["shipping_ref"], db_by_id)
        if db_row is None:
            continue  # cross-check C will catch this

        new_status = "SALE_RETURN" if entry["sr_rto"] == "SR" else "RTO"

        # Build enrichment payload
        payload: dict = {
            "return_reason":           entry["reason"],
            "return_responsible":      entry["responsible"],
            "return_customer_verbatim": entry["verbatim"],
        }
        if entry["return_date"]:
            payload["return_date"] = entry["return_date"].isoformat()

        # Enrich city/state only if currently NULL in DB
        if entry["city"] and not db_row.get("city"):
            payload["city"]  = entry["city"]
            city_state_enriched += 1
        if entry["state"] and not db_row.get("state"):
            payload["state"] = entry["state"]

        current_status = db_row["status"]

        if current_status == "FULFILLED":
            payload["status"] = new_status
            if dry_run:
                logger.info("[DRY-RUN] %s -> %s | reason=%s responsible=%s",
                            entry["order_no"], new_status, entry["reason"], entry["responsible"])
            else:
                db.table("orders").update(payload).eq(
                    "order_id", db_row["order_id"]
                ).eq("channel_id", FC_CHANNEL_ID).execute()
                logger.info("Tagged %s -> %s | reason=%s responsible=%s",
                            entry["order_no"], new_status, entry["reason"], entry["responsible"])
            if new_status == "RTO":
                tagged_rto += 1
            else:
                tagged_sr += 1

        elif current_status in ("RTO", "SALE_RETURN"):
            # Already tagged — just refresh enrichment fields
            if dry_run:
                logger.info("[DRY-RUN] Enrich %s (%s) | reason=%s",
                            entry["order_no"], current_status, entry["reason"])
            else:
                db.table("orders").update(payload).eq(
                    "order_id", db_row["order_id"]
                ).eq("channel_id", FC_CHANNEL_ID).execute()
                logger.info("Refreshed %s (%s) | reason=%s",
                            entry["order_no"], current_status, entry["reason"])
            already_tagged += 1

        else:
            logger.warning("Unexpected DB status %s for %s — skipped",
                           current_status, entry["order_no"])
            continue

        enriched += 1

    # ── Cross-checks ──────────────────────────────────────────────────────────

    # [A] already computed above: excel_returns_not_in_sheet

    # [B] DB FC RTO/SALE_RETURN in window not in sheet
    db_returns_in_window = (
        db.table("orders")
          .select("platform_order_id, sku_id, status")
          .eq("channel_id", FC_CHANNEL_ID)
          .in_("status", ["RTO", "SALE_RETURN"])
          .gte("order_date", str(window_start))
          .lte("order_date", str(window_end))
          .execute().data
    ) or []

    not_in_sheet_b = [
        r for r in db_returns_in_window
        if r["platform_order_id"] not in sheet_order_nos
    ]

    # [C] Sheet rows not found in DB at all
    not_in_db_c = [
        e for e in sheet_entries
        if _find_db_order(e["order_no"], e["shipping_ref"], db_by_id) is None
    ]

    # ── Print results ─────────────────────────────────────────────────────────

    if not_in_db:
        logger.warning("\n[GAP] %d order(s) in Excel report NOT found in DB"
                       " -- App miss, review and punch manually:", len(not_in_db))
        for r in not_in_db:
            logger.warning("    OrderID=%-25s  SKU=%-10s  Date=%s  SR=%s  file=%s",
                           r["order_id"], r["sku_id"],
                           r["order_date"] or "?", "YES" if r["is_sr"] else "no",
                           r["source_file"])

    if excel_returns_not_in_sheet:
        logger.warning("\n[A] %d order(s) have Is SR=Yes in Excel but NO entry in Google Sheet"
                       " -- add to tracking sheet:", len(excel_returns_not_in_sheet))
        for r in excel_returns_not_in_sheet:
            logger.warning("    OrderID=%-25s  SKU=%-10s  Date=%s",
                           r["order_id"], r["sku_id"], r["order_date"] or "?")

    if not_in_sheet_b:
        logger.warning("\n[B] %d FC order(s) are RTO/SALE_RETURN in DB (%s to %s)"
                       " but NOT in Google Sheet:", len(not_in_sheet_b), window_start, window_end)
        for r in not_in_sheet_b:
            logger.warning("    %-28s  (%s)", r["platform_order_id"], r["status"])

    if not_in_db_c:
        logger.warning("\n[C] %d order(s) in Google Sheet not found in DB at all"
                       " -- likely App miss:", len(not_in_db_c))
        for e in not_in_db_c:
            logger.warning("    OrderNo=%-25s  SKU=%-10s  %s  return_date=%s",
                           e["order_no"], e["sku"], e["sr_rto"], e["return_date"] or "?")

    # ── Summary ───────────────────────────────────────────────────────────────
    tag = "  [DRY-RUN -- nothing written]" if dry_run else ""
    print(
        f"\nFC Reconciliation{tag}\n"
        f"  Excel orders checked   : {len(excel_orders)}\n"
        f"  [GAP] Not in DB        : {len(not_in_db)}  (App miss -- review manually)\n"
        f"\n"
        f"  Tagged RTO             : {tagged_rto}\n"
        f"  Tagged SALE_RETURN     : {tagged_sr}\n"
        f"  Enriched (already tagged): {already_tagged}\n"
        f"  City/state enriched    : {city_state_enriched}\n"
        f"\n"
        f"  [A] SR in file, not in sheet : {len(excel_returns_not_in_sheet)}\n"
        f"  [B] RTO/SR in DB, not in sheet: {len(not_in_sheet_b)}\n"
        f"  [C] In sheet, not in DB      : {len(not_in_db_c)}\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="FC monthly reconciliation: gap detection + return tagging"
    )
    parser.add_argument("--folder", required=True,
                        help="Folder containing FC Excel files (Delivered + Shipped)")
    parser.add_argument("--env",     choices=["dev", "prod"], default="prod")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    os.environ.setdefault("TCB_ENV", args.env)
    run(Path(args.folder), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
