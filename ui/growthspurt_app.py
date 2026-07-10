"""
The Cradle Box — Sales MIS Dashboard (Phase C)
Read-only analytics. Auth via Streamlit Cloud viewer allowlist (no code needed).
"""
import sys, os
from pathlib import Path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

# Bridge Streamlit Cloud secrets → env vars so that tcb/db.py can read them.
# On Streamlit Cloud there is no .env file; credentials live in app secrets.
# Locally, .env is loaded by db.py directly — this block is a no-op locally.
try:
    import streamlit as _st
    for _k in ("SUPABASE_URL", "SUPABASE_KEY"):
        if _k in _st.secrets and not os.environ.get(_k):
            os.environ[_k] = _st.secrets[_k]
except Exception:
    pass

import calendar
from datetime import date, timedelta

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from tcb.db import get_orders_raw, get_blinkit_city_ds, get_skus, get_client, get_replen_plan
from tcb.forecasting import (
    get_forecast_display as _fc_get_display,
    get_forecast_channel_breakdown as _fc_get_channels,
    lock_sku_month as _fc_lock,
    reset_sku_month as _fc_reset,
    get_last_run_date as _fc_last_run,
    FORECAST_MONTHS as _FC_MONTHS,
    _next_n_month_starts as _fc_months,
)

st.set_page_config(
    page_title="TCB Sales MIS",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

GROSS_STATUSES  = {"FULFILLED", "PENDING", "RTO", "SALE_RETURN", "REPLACEMENT"}
NET_STATUSES    = {"FULFILLED", "PENDING"}
RETURN_STATUSES = {"RTO", "SALE_RETURN", "REPLACEMENT"}
ALL_STATUSES    = {"FULFILLED", "PENDING", "CANCELLED", "RTO", "SALE_RETURN", "REPLACEMENT"}
STATUS_ORDER    = ["FULFILLED", "REPLACEMENT", "PENDING", "RTO", "SALE_RETURN", "CANCELLED"]

# Canonical display names; DB has alternate spellings mapped below
BLINKIT_CITIES      = ["Bengaluru", "Chennai", "Ghaziabad", "Gurgaon", "Hyderabad", "New Delhi"]
BLINKIT_CITY_DB_MAP = {c: [c] for c in BLINKIT_CITIES}
BLINKIT_CITY_DB_MAP["Gurgaon"]   = ["Gurgaon", "Gurugram"]   # same city, two spellings
BLINKIT_CITY_DB_MAP["New Delhi"]  = ["New Delhi", "Delhi"]    # same city, two spellings
# All DB city values that belong to a Blinkit city (used for "Others" exclusion)
_ALL_BLINKIT_DB_CITIES = {v for vals in BLINKIT_CITY_DB_MAP.values() for v in vals}

CHANNEL_COLOR_MAP = {
    "Amazon":         "#FF9900",
    "Blinkit":        "#FFED29",
    "First Cry":      "#A12134",
    "Ferns & Petals": "#7D8035",
    "Peeko":          "#CBF2B8",
    "Ozi":            "#F1DFEC",
    "Kiddo":          "#F46060",
    "D2C":            "#222222",
}

# Channels billed by purchase order (bulk invoice) — current-month data must NOT be
# linearly projected because one billing early in the month ≠ daily run-rate.
OUTRIGHT_CHANNELS = {"Peeko", "Kiddo"}


# ── Data loading ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_data() -> pd.DataFrame:
    rows = get_orders_raw()
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["order_date"] = pd.to_datetime(df["order_date"])
    df["month"]      = df["order_date"].dt.to_period("M")
    df["month_dt"]   = df["month"].dt.to_timestamp()
    df["year"]       = df["order_date"].dt.year
    df["quarter"]    = df["order_date"].dt.quarter
    for col in ("gross_value", "selling_price", "mrp", "quantity"):
        df[col] = pd.to_numeric(df.get(col, 0), errors="coerce").fillna(0)
    df["status"] = df["status"].fillna("FULFILLED")
    df["channel_name"] = df["channel_name"].replace({"Amazon FBM": "Amazon", "Own Website": "D2C"})
    return df


# ── Helpers ────────────────────────────────────────────────────────────────────
def fmt_inr(val: float) -> str:
    if val >= 1_00_000:
        return f"₹{val/1_00_000:.1f}L"
    if val >= 1_000:
        return f"₹{val/1_000:.1f}K"
    return f"₹{val:.0f}"


def pct(num, denom) -> float:
    return round(100 * num / denom, 1) if denom else 0.0


def return_rate_pct(grp: pd.DataFrame) -> float:
    base = grp[grp["status"].isin({"FULFILLED", "RTO", "SALE_RETURN", "REPLACEMENT"})]
    return pct(
        base[base["status"].isin(RETURN_STATUSES)]["quantity"].sum(),
        base["quantity"].sum(),
    )


def fmt_delta(val) -> str:
    if val is None:
        return "—"
    return f"{'+'if val>0 else ''}{val:.1f}%"


def colour_delta(val) -> str:
    if not isinstance(val, str) or val == "—":
        return ""
    try:
        v = float(val.replace("%", "").replace("+", ""))
        if v > 0:
            return "color: green"
        if v < 0:
            return "color: red"
    except ValueError:
        pass
    return ""


def blue_gradient(col: pd.Series) -> pd.Series:
    """Column-wise blue gradient — each month column normalised independently."""
    max_val = col.max() if col.max() else 1
    return col.map(
        lambda v: f"background-color: rgba(30, 100, 255, {min(v / max_val, 1) * 0.6:.2f})"
    )


def red_gradient(col: pd.Series) -> pd.Series:
    """Column-wise red gradient — each column normalised independently."""
    max_val = col.max() if col.max() else 1
    return col.map(
        lambda v: f"background-color: rgba(255, 80, 0, {min(v / max_val, 1) * 0.6:.2f})"
    )


def colour_rr(val) -> str:
    if not isinstance(val, str):
        return ""
    try:
        v = float(val.replace("%", ""))
        if v < 5:
            return "color: green"
        if v < 15:
            return "color: orange"
        return "color: red"
    except ValueError:
        return ""


def _cur_mult() -> float:
    """Multiplier to project current month-to-date to full month-end.
    Data is available through yesterday, so use today.day - 1 as elapsed days."""
    today = date.today()
    data_days = max(today.day - 1, 1)
    return calendar.monthrange(today.year, today.month)[1] / data_days


def _project_col(df: pd.DataFrame, value_col: str, mult: float) -> pd.DataFrame:
    """Scale current-month rows by mult, rounded to nearest whole number.
    Outright channels (Peeko, Kiddo) are billed by PO and are never projected."""
    today  = date.today()
    cur_ts = pd.Timestamp(date(today.year, today.month, 1))
    out    = df.copy()
    mask   = out["month_dt"] == cur_ts
    if "channel_name" in out.columns:
        mask = mask & ~out["channel_name"].isin(OUTRIGHT_CHANNELS)
    out.loc[mask, value_col] = (out.loc[mask, value_col] * mult).round(0)
    return out


# ── Sidebar ────────────────────────────────────────────────────────────────────
def sidebar(df: pd.DataFrame) -> dict:
    st.sidebar.markdown("## Filters")

    all_channels = sorted(df["channel_name"].dropna().unique().tolist())
    sel_ch = st.sidebar.multiselect("Channel", ["All Channels"] + all_channels, default=["All Channels"])
    if "All Channels" in sel_ch or not sel_ch:
        sel_ch = all_channels

    all_skus = sorted(df["sku_id"].dropna().unique().tolist())
    sel_sku = st.sidebar.multiselect("SKU", ["All SKUs"] + all_skus, default=["All SKUs"])
    if "All SKUs" in sel_sku or not sel_sku:
        sel_sku = all_skus

    months_avail = sorted(df["month_dt"].dt.to_period("M").unique().astype(str).tolist())
    if months_avail:
        sel_range = st.sidebar.select_slider(
            "Month range", options=months_avail,
            value=(months_avail[0], months_avail[-1]),
        )
    else:
        sel_range = (None, None)

    city_options = BLINKIT_CITIES + ["Others"]
    sel_city = st.sidebar.multiselect("City", ["All Cities"] + city_options, default=["All Cities"])
    if "All Cities" in sel_city or not sel_city:
        sel_city = ["__all__"]

    mode = st.sidebar.radio(
        "Revenue mode",
        ["Gross (excl. Cancelled)", "Net (Fulfilled only)"],
        index=0,
    )
    st.sidebar.markdown("---")
    st.sidebar.caption("Gross = all statuses except Cancelled  \nNet = Fulfilled + Pending only")

    return {"channels": sel_ch, "skus": sel_sku, "range": sel_range, "net_mode": "Net" in mode, "cities": sel_city}


def _apply_city_filter(df: pd.DataFrame, cities: list) -> pd.DataFrame:
    if "__all__" in cities:
        return df
    specific  = [c for c in cities if c != "Others"]
    db_values = {v for c in specific for v in BLINKIT_CITY_DB_MAP.get(c, [c])}
    if "Others" in cities:
        return df[df["city"].isin(db_values) | ~df["city"].isin(_ALL_BLINKIT_DB_CITIES)]
    return df[df["city"].isin(db_values)]


def apply_filters(df: pd.DataFrame, f: dict) -> pd.DataFrame:
    out = df[df["channel_name"].isin(f["channels"]) & df["sku_id"].isin(f["skus"])]
    if f["range"][0]:
        out = out[out["month_dt"].dt.to_period("M").astype(str) >= f["range"][0]]
    if f["range"][1]:
        out = out[out["month_dt"].dt.to_period("M").astype(str) <= f["range"][1]]
    out = _apply_city_filter(out, f.get("cities", ["__all__"]))
    return out[out["status"] != "CANCELLED"]


def active_df(df: pd.DataFrame, net_mode: bool) -> pd.DataFrame:
    return df[df["status"].isin(NET_STATUSES if net_mode else GROSS_STATUSES)]


# ── Velocity snapshot ──────────────────────────────────────────────────────────
def velocity_snapshot(raw_df: pd.DataFrame) -> None:
    """SKU velocity: one column per day (7 days, most recent first), curr/last month avg."""
    today      = date.today()
    yesterday  = today - timedelta(days=1)
    days       = [yesterday - timedelta(days=i) for i in range(7)]

    cur_period      = pd.Period(today, "M")
    last_period     = cur_period - 1
    last_month_days = calendar.monthrange(last_period.year, last_period.month)[1]

    gross = raw_df[raw_df["status"].isin(GROSS_STATUSES)].copy()
    gross["order_day"] = gross["order_date"].dt.date

    l7_data = gross[(gross["order_day"] >= days[-1]) & (gross["order_day"] <= yesterday)]
    daily_u = l7_data.groupby(["sku_id", "order_day"])["quantity"].sum()

    cm_u = gross[gross["month"] == cur_period].groupby("sku_id")["quantity"].sum()
    lm_u = gross[gross["month"] == last_period].groupby("sku_id")["quantity"].sum()

    sku_name_map = (
        raw_df.dropna(subset=["sku_id"])
        .drop_duplicates("sku_id")
        .set_index("sku_id")["sku_name"]
        .to_dict()
    )
    all_sku_ids = sorted(raw_df["sku_id"].dropna().unique())

    day_cols   = [f"{d.day} {d.strftime('%b')}" for d in days]
    avg_col_cm = f"{today.strftime('%b')} Avg"
    avg_col_lm = f"{last_period.strftime('%b')} Avg"

    vs_lm_col = "vs. LM"

    def _vs_lm_pct(cm_avg, lm_avg):
        return (cm_avg - lm_avg) / lm_avg * 100 if lm_avg else float("nan")

    def _fmt_vs_lm(v):
        return "—" if (v != v) else f"{'+'if v > 0 else ''}{v:.1f}%"

    def _color_vs_lm(v):
        if v != v:
            return ""
        return "color: green" if v > 5 else ("color: red" if v < -5 else "")

    rows = []
    for sku_id in all_sku_ids:
        row = {"SKU Code": sku_id, "SKU Name": sku_name_map.get(sku_id, sku_id)}
        for d, col in zip(days, day_cols):
            row[col] = int(daily_u.get((sku_id, d), 0))
        row[avg_col_cm] = round(cm_u.get(sku_id, 0) / max(today.day - 1, 1), 1)
        row[avg_col_lm] = round(lm_u.get(sku_id, 0) / last_month_days, 1)
        row[vs_lm_col]  = _vs_lm_pct(row[avg_col_cm], row[avg_col_lm])
        rows.append(row)

    vel_df = pd.DataFrame(rows).sort_values("SKU Code").reset_index(drop=True)

    # Drop SKUs with zero units across all columns
    numeric_cols = day_cols + [avg_col_cm, avg_col_lm]
    vel_df = vel_df[vel_df[numeric_cols].sum(axis=1) > 0].reset_index(drop=True)

    total_row = {"SKU Code": "TOTAL", "SKU Name": ""}
    for col in numeric_cols:
        total_row[col] = round(vel_df[col].sum(), 1)
    total_row[vs_lm_col] = _vs_lm_pct(total_row[avg_col_cm], total_row[avg_col_lm])
    total_df = pd.DataFrame([total_row])

    def _style_vel(df):
        return (
            df.style
            .format({avg_col_cm: "{:.1f}", avg_col_lm: "{:.1f}", vs_lm_col: _fmt_vs_lm})
            .map(_color_vs_lm, subset=[vs_lm_col])
        )

    st.subheader("Gross Units Sold")
    st.dataframe(_style_vel(total_df), use_container_width=True, hide_index=True)
    st.dataframe(_style_vel(vel_df),   use_container_width=True, hide_index=True)


# ── Tab 1 — Overview ───────────────────────────────────────────────────────────
def tab_overview(raw_df: pd.DataFrame, fdf: pd.DataFrame, net_mode: bool, filters: dict = None):
    st.caption("Cancelled orders are excluded from all metrics on this tab.")
    adf = active_df(fdf, net_mode)

    # Velocity snapshot + MTD projection use raw_df (no date filter) but should respect
    # channel / SKU / city filters from the sidebar
    filtered_raw = raw_df.copy()
    if filters:
        filtered_raw = filtered_raw[
            filtered_raw["channel_name"].isin(filters["channels"]) &
            filtered_raw["sku_id"].isin(filters["skus"])
        ]
        filtered_raw = _apply_city_filter(filtered_raw, filters.get("cities", ["__all__"]))

    total_orders = adf[["order_id", "channel_id"]].drop_duplicates().shape[0]
    total_units  = int(adf["quantity"].sum())
    gross_rev    = fdf[fdf["status"].isin(GROSS_STATUSES)]["gross_value"].sum()
    net_rev      = fdf[fdf["status"].isin(NET_STATUSES)]["gross_value"].sum()

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Orders", f"{total_orders:,}")
    c2.metric("Units", f"{total_units:,}")
    c3.metric("Gross Revenue", fmt_inr(gross_rev))
    c4.metric("Net Revenue", fmt_inr(net_rev))
    c5.metric("Return Rate", f"{return_rate_pct(fdf):.1f}%")

    today = date.today()
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    cur = filtered_raw[
        (filtered_raw["order_date"].dt.year  == today.year) &
        (filtered_raw["order_date"].dt.month == today.month) &
        (filtered_raw["status"].isin(NET_STATUSES))
    ]
    mtd_units = int(cur["quantity"].sum())
    mtd_rev   = cur["gross_value"].sum()
    data_days = max(today.day - 1, 1)
    mult = days_in_month / data_days
    cur_out  = cur[cur["channel_name"].isin(OUTRIGHT_CHANNELS)]
    cur_reg  = cur[~cur["channel_name"].isin(OUTRIGHT_CHANNELS)]
    proj_units = round(cur_reg["quantity"].sum() * mult) + int(cur_out["quantity"].sum())
    proj_rev   = cur_reg["gross_value"].sum() * mult + cur_out["gross_value"].sum()
    st.info(
        f"**{today.strftime('%b %Y')} projection:** "
        f"{mtd_units:,} units so far ({data_days} of {days_in_month} days) "
        f"→ **~{proj_units:,} units** projected | "
        f"{fmt_inr(mtd_rev)} revenue so far → **~{fmt_inr(proj_rev)}** projected"
    )

    velocity_snapshot(filtered_raw)

    st.markdown("---")

    all_channels = sorted(fdf["channel_name"].dropna().unique().tolist())
    monthly = (
        fdf[fdf["status"].isin(GROSS_STATUSES)]
        .groupby(["month_dt", "channel_name"], as_index=False)["gross_value"].sum()
    )
    monthly["Month"] = monthly["month_dt"].dt.strftime("%b %Y")

    col_l, col_r = st.columns([3, 2])
    with col_l:
        st.subheader("Revenue by Channel")
        if not monthly.empty:
            monthly = monthly.sort_values("month_dt")
            monthly["gross_value_L"] = monthly["gross_value"] / 100_000
            fig = px.bar(
                monthly,
                x="Month", y="gross_value_L", color="channel_name",
                labels={"gross_value_L": "Gross Revenue (₹ Lacs)", "channel_name": "Channel"},
                color_discrete_map=CHANNEL_COLOR_MAP,
                category_orders={
                    "Month":        monthly["Month"].unique().tolist(),
                    "channel_name": all_channels,
                },
            )
            # Total labels on top of each bar
            totals = (
                monthly.groupby(["Month", "month_dt"], sort=False)[["gross_value", "gross_value_L"]]
                .sum().reset_index().sort_values("month_dt")
            )
            for _, row in totals.iterrows():
                fig.add_annotation(
                    x=row["Month"], y=row["gross_value_L"],
                    text=fmt_inr(row["gross_value"]),
                    showarrow=False, yshift=10,
                    font=dict(size=10, color="#333"),
                )
            fig.update_traces(
                hovertemplate="<b>%{fullData.name}</b><br>Month: %{x}<br>Revenue: %{y:.1f}L<extra></extra>"
            )
            fig.update_layout(
                legend_title="Channel", xaxis_title=None, margin=dict(t=35),
                yaxis=dict(ticksuffix="L", tickformat=".0f"),
            )
            st.plotly_chart(fig, use_container_width=True)

    with col_r:
        st.subheader("Revenue Share")
        ch_share = (
            fdf[fdf["status"].isin(GROSS_STATUSES)]
            .groupby("channel_name", as_index=False)["gross_value"].sum()
            .sort_values("channel_name")
        )
        if not ch_share.empty:
            fig2 = px.pie(
                ch_share, values="gross_value", names="channel_name",
                hole=0.4, color="channel_name",
                color_discrete_map=CHANNEL_COLOR_MAP,
            )
            fig2.update_traces(texttemplate="%{percent:.1%}")
            fig2.update_layout(legend_title="Channel", margin=dict(t=20))
            st.plotly_chart(fig2, use_container_width=True)

    st.subheader("Units by Channel")
    pivot_statuses = NET_STATUSES if net_mode else GROSS_STATUSES
    pivot_df = (
        fdf[fdf["status"].isin(pivot_statuses)]
        .groupby(["channel_name", "month_dt"])["quantity"].sum()
        .unstack("month_dt").fillna(0).astype(int)
    )
    if not pivot_df.empty:
        pivot_df.columns = [pd.Timestamp(c).strftime("%b %Y") for c in pivot_df.columns]
        pivot_df["Total"] = pivot_df.sum(axis=1)
        total_row = pivot_df.sum(axis=0).rename("TOTAL")
        pivot_df = pd.concat([pivot_df, total_row.to_frame().T])
        pivot_df.index.name = "Channels"  # set after concat so reset_index names it correctly
        pivot_df = pivot_df.reset_index()

        num_cols      = [c for c in pivot_df.columns if c != "Channels"]
        is_total_mask = [ch == "TOTAL" for ch in pivot_df["Channels"]]

        def _gradient_excl_total(col):
            # normalise against channel rows only — TOTAL would always dominate otherwise
            max_val = max((v for flag, v in zip(is_total_mask, col) if not flag), default=1) or 1
            return pd.Series(
                [
                    "" if flag
                    else f"background-color: rgba(30, 100, 255, {min(v / max_val, 1) * 0.6:.2f})"
                    for flag, v in zip(is_total_mask, col)
                ],
                index=col.index,
            )

        st.dataframe(
            pivot_df.style
            .apply(_gradient_excl_total, axis=0, subset=num_cols)
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if is_total_mask[row.name] else [""] * len(row),
                axis=1,
            ),
            use_container_width=True,
            hide_index=True,
        )


# ── Tab 2 — Trends ─────────────────────────────────────────────────────────────
def tab_trends(fdf: pd.DataFrame, net_mode: bool):
    st.caption("Cancelled orders are excluded from all metrics on this tab.")
    adf = active_df(fdf, net_mode)

    today        = date.today()
    mult         = _cur_mult()
    all_channels = sorted(fdf["channel_name"].dropna().unique().tolist())

    # ── Monthly units trend ───────────────────────────────────────────────────
    st.subheader("Units by Channel")
    monthly_ch = adf.groupby(["month_dt", "channel_name"])["quantity"].sum().reset_index()
    monthly_ch = _project_col(monthly_ch, "quantity", mult)
    monthly_ch["Month"] = monthly_ch["month_dt"].dt.strftime("%b %Y")

    if not monthly_ch.empty:
        monthly_ch = monthly_ch.sort_values("month_dt")
        fig = px.line(
            monthly_ch,
            x="Month", y="quantity", color="channel_name", markers=True,
            labels={"quantity": "Units Sold", "channel_name": "Channel"},
            color_discrete_map=CHANNEL_COLOR_MAP,
            category_orders={
                "Month":        monthly_ch["Month"].unique().tolist(),
                "channel_name": all_channels,
            },
        )
        fig.update_layout(xaxis_title=None, margin=dict(t=20))
        st.plotly_chart(fig, use_container_width=True)
    st.caption(
        f"⚠️ {today.strftime('%b %Y')} shows projected month-end "
        f"({max(today.day - 1, 1)} days elapsed, {mult:.1f}× multiplier)."
    )

    # ── MoM comparison — 3 tables, channels as rows ───────────────────────────
    st.subheader("Month-on-Month Comparison")

    cur_period = pd.Period(today, "M")
    periods    = [cur_period - i for i in range(4)]  # M, M-1, M-2, M-3
    col_m      = f"{cur_period} (proj)"
    col_m1     = str(cur_period - 1)
    col_m2     = str(cur_period - 2)
    col_avg    = "L3M"
    vs_m1  = f"vs. {col_m1}"   # M(proj) vs M-1
    vs_m2  = f"vs. {col_m2}"   # M-1 vs M-2
    vs_avg = "vs. L3M"          # M(proj) vs 3-month avg
    all_vs = [vs_m1, vs_m2, vs_avg]

    def _grp(ch, period):
        sub = fdf[(fdf["channel_name"] == ch) & (fdf["month"] == period)]
        return sub[sub["status"].isin(GROSS_STATUSES)]

    def _grp_all(period):
        sub = fdf[fdf["month"] == period]
        return sub[sub["status"].isin(GROSS_STATUSES)]

    def _ch_val(ch, period, metric):
        g = _grp(ch, period)
        if metric == "units":   return float(g["quantity"].sum())
        if metric == "orders":  return float(g[["order_id", "channel_id"]].drop_duplicates().shape[0])
        if metric == "revenue": return float(g["gross_value"].sum())
        if metric == "asp":
            u = float(g["quantity"].sum())
            return float(g["gross_value"].sum()) / u if u else float("nan")
        if metric == "cart_size":
            o = float(g[["order_id", "channel_id"]].drop_duplicates().shape[0])
            return float(g["gross_value"].sum()) / o if o else float("nan")
        return 0.0

    def _total_val(period, metric):
        g = _grp_all(period)
        if metric == "units":   return float(g["quantity"].sum())
        if metric == "orders":  return float(g[["order_id", "channel_id"]].drop_duplicates().shape[0])
        if metric == "revenue": return float(g["gross_value"].sum())
        if metric == "asp":
            u = float(g["quantity"].sum())
            return float(g["gross_value"].sum()) / u if u else float("nan")
        if metric == "cart_size":
            o = float(g[["order_id", "channel_id"]].drop_duplicates().shape[0])
            return float(g["gross_value"].sum()) / o if o else float("nan")
        return 0.0

    def _nanavg(values):
        valid = [v for v in values if v == v]
        return sum(valid) / len(valid) if valid else float("nan")

    def _vs_pct(a, b):
        if a != a or b != b: return float("nan")
        return (a - b) / b * 100 if b else float("nan")

    def _fmt_vs(v):
        return "—" if (v != v) else f"{'+'if v > 0 else ''}{v:.1f}%"

    def _color_vs(v):
        if v != v:
            return ""
        return "color: red" if v < -5 else ("color: green" if v > 5 else "")

    def _mom_table(metric, fmt_fn, project=True):
        channels = sorted(fdf["channel_name"].dropna().unique())
        rows = []
        for ch in channels:
            vals   = [_ch_val(ch, p, metric) for p in periods]
            avg    = _nanavg(vals[1:])
            m_proj = vals[0] if (not project or ch in OUTRIGHT_CHANNELS) else vals[0] * mult
            rows.append({
                "Channel": ch,
                col_m:   fmt_fn(m_proj),
                col_m1:  fmt_fn(vals[1]),
                vs_m1:   _vs_pct(m_proj, vals[1]),
                col_m2:  fmt_fn(vals[2]),
                vs_m2:   _vs_pct(vals[1], vals[2]),
                col_avg: fmt_fn(avg),
                vs_avg:  _vs_pct(m_proj, avg),
            })
        totals = [_total_val(p, metric) for p in periods]
        t_avg  = _nanavg(totals[1:])
        if project and metric not in ("asp", "cart_size"):
            out_total = sum(_ch_val(ch, periods[0], metric) for ch in channels if ch in OUTRIGHT_CHANNELS)
            t_proj    = (totals[0] - out_total) * mult + out_total
        else:
            t_proj = totals[0]
        rows.append({
            "Channel": "TOTAL",
            col_m:   fmt_fn(t_proj),
            col_m1:  fmt_fn(totals[1]),
            vs_m1:   _vs_pct(t_proj, totals[1]),
            col_m2:  fmt_fn(totals[2]),
            vs_m2:   _vs_pct(totals[1], totals[2]),
            col_avg: fmt_fn(t_avg),
            vs_avg:  _vs_pct(t_proj, t_avg),
        })
        df = pd.DataFrame(rows).set_index("Channel")
        return (
            df.style
            .format({c: _fmt_vs for c in all_vs})
            .map(_color_vs, subset=all_vs)
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if row.name == "TOTAL" else [""] * len(row),
                axis=1,
            )
        )

    _fmt_inr_safe = lambda v: "—" if (v != v) else f"₹{v:.0f}"

    st.markdown("**MoM Units**")
    st.dataframe(_mom_table("units",  lambda v: round(v)), use_container_width=True)

    st.markdown("**MoM Orders**")
    st.dataframe(_mom_table("orders", lambda v: round(v)), use_container_width=True)

    st.markdown("**MoM Revenue**")
    st.dataframe(_mom_table("revenue", lambda v: f"₹{round(v):,}"), use_container_width=True)

    st.markdown("**MoM ASP**")
    st.dataframe(_mom_table("asp", _fmt_inr_safe, project=False), use_container_width=True)

    st.markdown("**MoM AOV**")
    st.dataframe(_mom_table("cart_size", _fmt_inr_safe, project=False), use_container_width=True)

    st.caption(
        f"⚠️ {today.strftime('%b %Y')} (proj) = month-to-date × {mult:.1f} "
        f"({max(today.day - 1, 1)} of {calendar.monthrange(today.year, today.month)[1]} days elapsed)."
    )


# ── Tab 3 — By Channel ─────────────────────────────────────────────────────────
def tab_channel(fdf: pd.DataFrame, net_mode: bool):
    st.caption("Cancelled orders are excluded from all metrics on this tab.")

    channels_avail = sorted(fdf["channel_name"].dropna().unique().tolist())
    ch_sel = st.radio("View", ["All Channels"] + channels_avail, horizontal=True)
    cdf    = fdf if ch_sel == "All Channels" else fdf[fdf["channel_name"] == ch_sel]

    # ── Performance summary ───────────────────────────────────────────────────
    st.subheader("Channel Performance")
    today      = date.today()
    cur_period = pd.Period(today, "M")

    def _period_metrics(df_sub: pd.DataFrame, p: pd.Period) -> dict:
        sub    = df_sub[df_sub["month"] == p]
        g      = sub[sub["status"].isin(GROSS_STATUSES)]
        orders = g[["order_id", "channel_id"]].drop_duplicates().shape[0]
        units  = int(g["quantity"].sum())
        rev    = g["gross_value"].sum()
        return {
            "orders":    orders,
            "units":     units,
            "gross_rev": rev,
            "asp":       rev / units  if units  else 0.0,
            "aov":       rev / orders if orders else 0.0,
        }

    mult = _cur_mult()
    pm   = [_period_metrics(cdf, cur_period - i) for i in range(4)]
    M_raw, M1, M2, M3 = pm
    # Outright channels are billed by PO — exclude them from projection, keep actuals
    cdf_reg = cdf[~cdf["channel_name"].isin(OUTRIGHT_CHANNELS)]
    cdf_out = cdf[cdf["channel_name"].isin(OUTRIGHT_CHANNELS)]
    M_reg   = _period_metrics(cdf_reg, cur_period)
    M_out   = _period_metrics(cdf_out, cur_period)
    M = {}
    for k in M_raw:
        if k in ("asp", "aov"):
            M[k] = M_raw[k]
        else:
            M[k] = M_reg[k] * mult + M_out[k]
    avg = {k: (M1[k] + M2[k] + M3[k]) / 3 for k in M}

    def _fmt(key, val) -> str:
        if key in ("gross_rev", "asp", "aov"):
            return f"₹{round(val):,}"
        return f"{round(val):,}"

    def _vs_pct_ch(m_val, prev_val):
        return (m_val - prev_val) / prev_val * 100 if prev_val else float("nan")

    def _fmt_vs_ch(v):
        return "—" if (v != v) else f"{'+'if v > 0 else ''}{v:.1f}%"

    def _color_vs_ch(v):
        if v != v:
            return ""
        return "color: green" if v > 5 else ("color: red" if v < -5 else "")

    vs_cols_ch    = ["vs. M-1", "vs. M-2", "vs. L3M"]
    metric_labels = [
        ("orders",    "Orders"),
        ("units",     "Units"),
        ("gross_rev", "Gross Revenue"),
        ("asp",       "ASP"),
        ("aov",       "AOV"),
    ]
    all_labels    = [label for _, label in metric_labels]

    summary_rows = []
    for key, label in metric_labels:
        vs_vals = {
            "vs. M-1": _vs_pct_ch(M[key], M1[key]),
            "vs. M-2": _vs_pct_ch(M[key], M2[key]),
            "vs. L3M": _vs_pct_ch(M[key], avg[key]),
        }
        summary_rows.append({
            "Metric":                label,
            f"{cur_period} (Proj)":  _fmt(key, M[key]),
            f"{cur_period-1} (M-1)": _fmt(key, M1[key]),
            "vs. M-1":               vs_vals["vs. M-1"],
            f"{cur_period-2} (M-2)": _fmt(key, M2[key]),
            "vs. M-2":               vs_vals["vs. M-2"],
            "L3M":                   _fmt(key, avg[key]),
            "vs. L3M":               vs_vals["vs. L3M"],
        })

    summary_df = pd.DataFrame(summary_rows).set_index("Metric")
    st.dataframe(
        summary_df.style
        .format({c: _fmt_vs_ch for c in vs_cols_ch}, subset=pd.IndexSlice[all_labels, vs_cols_ch])
        .map(_color_vs_ch, subset=pd.IndexSlice[all_labels, vs_cols_ch]),
        use_container_width=True,
    )

    # ── Trend — Units or Revenue, stacked, alphabetical, projected ────────────
    st.subheader("Monthly Trend")
    metric_sel = st.radio("Metric", ["Units", "Revenue"], horizontal=True, key="ch_metric")
    is_revenue = metric_sel == "Revenue"
    value_col  = "gross_value" if is_revenue else "quantity"
    y_label    = "Gross Revenue (₹ Lacs)" if is_revenue else "Units Sold"

    mult     = _cur_mult()
    cdf_plot = cdf[cdf["status"].isin(NET_STATUSES if net_mode else GROSS_STATUSES)]
    trend    = cdf_plot.groupby(["month_dt", "channel_name"])[value_col].sum().reset_index()
    trend    = _project_col(trend, value_col, mult)
    trend["Month"] = trend["month_dt"].dt.strftime("%b %Y")

    # Scale revenue to Lacs for the chart
    if is_revenue:
        trend["plot_val"] = trend[value_col] / 100_000
    else:
        trend["plot_val"] = trend[value_col]

    all_channels_in_view = sorted(cdf_plot["channel_name"].dropna().unique())

    if not trend.empty:
        trend = trend.sort_values("month_dt")
        fig = px.bar(
            trend, x="Month", y="plot_val", color="channel_name",
            barmode="stack",
            labels={"plot_val": y_label, "channel_name": "Channel"},
            color_discrete_map=CHANNEL_COLOR_MAP,
            category_orders={
                "Month":        trend["Month"].unique().tolist(),
                "channel_name": all_channels_in_view,
            },
        )
        # Totals on top of each bar
        totals = trend.groupby(["Month", "month_dt"], sort=False)["plot_val"].sum().reset_index()
        totals = totals.sort_values("month_dt")
        for _, row in totals.iterrows():
            val = row["plot_val"]
            label_text = f"₹{val:.1f}L" if is_revenue else f"{int(val):,}"
            fig.add_annotation(
                x=row["Month"], y=val,
                text=label_text,
                showarrow=False, yshift=10,
                font=dict(size=10, color="#333"),
            )
        layout_kwargs = dict(xaxis_title=None, margin=dict(t=35))
        if is_revenue:
            layout_kwargs["yaxis"] = dict(ticksuffix="L", tickformat=".0f")
        fig.update_layout(**layout_kwargs)
        if is_revenue:
            fig.update_traces(
                hovertemplate="<b>%{fullData.name}</b><br>Month: %{x}<br>Revenue: %{y:.1f}L<extra></extra>"
            )
        st.plotly_chart(fig, use_container_width=True)
    st.caption(
        f"⚠️ {today.strftime('%b %Y')} shows projected month-end "
        f"({max(today.day - 1, 1)} days elapsed, {mult:.1f}× multiplier)."
    )

    # ── Fulfillment type split ────────────────────────────────────────────────
    st.subheader("Fulfillment Type Split")
    if "fulfillment_type" in cdf_plot.columns:
        ft = cdf_plot.groupby("fulfillment_type")[["order_id"]].count().reset_index()
        ft.columns = ["Fulfillment Type", "Orders"]
        n_slices = len(ft)
        blues = [f"hsl(220, 70%, {int(30 + i * 40 / max(n_slices - 1, 1))}%)" for i in range(n_slices)]
        fig_ft = px.pie(
            ft, values="Orders", names="Fulfillment Type",
            hole=0.4, color_discrete_sequence=blues,
        )
        fig_ft.update_traces(texttemplate="%{percent:.1%}")
        fig_ft.update_layout(margin=dict(t=20))
        st.plotly_chart(fig_ft, use_container_width=True)
    else:
        st.info("Fulfillment type data not available.")


# ── Tab 4 — By SKU ─────────────────────────────────────────────────────────────
def tab_sku(fdf: pd.DataFrame, net_mode: bool):
    st.caption("Cancelled orders are excluded from all metrics on this tab.")
    adf = active_df(fdf, net_mode)

    st.subheader("SKU Performance")
    g_all = fdf[fdf["status"].isin(GROSS_STATUSES)]
    n_all = fdf[fdf["status"].isin(NET_STATUSES)]
    total_units     = int(g_all["quantity"].sum()) or 1
    total_gross_rev = g_all["gross_value"].sum() or 1
    sku_rows = []
    for sku_id, grp in fdf.groupby("sku_id"):
        g   = grp[grp["status"].isin(GROSS_STATUSES)]
        n   = grp[grp["status"].isin(NET_STATUSES)]
        asp = (n["gross_value"].sum() / n["quantity"].sum()) if n["quantity"].sum() else 0
        sku_rows.append({
            "SKU":           sku_id,
            "SKU Name":      grp["sku_name"].iloc[0],
            "Orders":        g[["order_id", "channel_id"]].drop_duplicates().shape[0],
            "Units":         int(g["quantity"].sum()),
            "Gross Revenue": f"₹{round(g['gross_value'].sum()):,}",
            "Net Revenue":   f"₹{round(n['gross_value'].sum()):,}",
            "ASP (₹)":       f"₹{asp:.0f}",
            "Return Rate %": f"{return_rate_pct(grp):.1f}%",
            "% of Units":    f"{pct(g['quantity'].sum(), total_units):.1f}%",
            "% of Revenue":  f"{pct(g['gross_value'].sum(), total_gross_rev):.1f}%",
        })

    if sku_rows:
        sku_df = pd.DataFrame(sku_rows).sort_values("Units", ascending=False).reset_index(drop=True)

        # TOTAL row computed from raw data to avoid double-counting orders
        t_asp = total_gross_rev / total_units if total_units else 0
        total_row_sku = {
            "SKU":           "TOTAL",
            "SKU Name":      "",
            "Orders":        g_all[["order_id", "channel_id"]].drop_duplicates().shape[0],
            "Units":         total_units,
            "Gross Revenue": f"₹{round(total_gross_rev):,}",
            "Net Revenue":   f"₹{round(n_all['gross_value'].sum()):,}",
            "ASP (₹)":       f"₹{t_asp:.0f}",
            "Return Rate %": f"{return_rate_pct(fdf):.1f}%",
            "% of Units":    "100.0%",
            "% of Revenue":  "100.0%",
        }
        sku_df = pd.concat([sku_df, pd.DataFrame([total_row_sku])], ignore_index=True)
        is_total = [False] * (len(sku_df) - 1) + [True]

        def _units_heatmap(col):
            max_val = max((v for flag, v in zip(is_total, col) if not flag), default=1) or 1
            return pd.Series(
                [
                    "" if flag
                    else f"background-color: rgba(30, 100, 255, {min(v / max_val, 1) * 0.6:.2f})"
                    for flag, v in zip(is_total, col)
                ],
                index=col.index,
            )

        st.dataframe(
            sku_df.style
            .apply(_units_heatmap, subset=["Units"])
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if is_total[row.name] else [""] * len(row),
                axis=1,
            )
            .map(colour_rr, subset=["Return Rate %"]),
            use_container_width=True, hide_index=True,
        )

    st.subheader("SKU Comparison")
    today = date.today()
    mult  = _cur_mult()
    sku_monthly = adf.groupby(["month_dt", "sku_id", "sku_name"])["quantity"].sum().reset_index()
    if not sku_monthly.empty:
        all_sku_ids  = sorted(adf["sku_id"].dropna().unique().tolist())
        defaults     = [s for s in ["TCB004", "TCB005", "TCB006"] if s in all_sku_ids]
        # Fall back to first 3 available if defaults aren't in the data
        if len(defaults) < 3:
            defaults = all_sku_ids[:3]
        sel_skus = st.multiselect(
            "Select up to 3 SKUs to compare",
            options=all_sku_ids,
            default=defaults,
            max_selections=3,
            key="sku_compare_sel",
        )
        if sel_skus:
            sku_id_name  = dict(zip(sku_monthly["sku_id"], sku_monthly["sku_name"]))
            trend_data   = sku_monthly[sku_monthly["sku_id"].isin(sel_skus)].copy()
            trend_data   = _project_col(trend_data, "quantity", mult)
            trend_data["Series"] = trend_data["sku_id"].map(
                lambda x: f"{x} — {sku_id_name.get(x, x)}"
            )
            trend_data["Month"] = trend_data["month_dt"].dt.strftime("%b %Y")
            trend_data = trend_data.sort_values("month_dt")
            fig = px.line(
                trend_data, x="Month", y="quantity",
                color="Series", markers=True,
                labels={"quantity": "Units", "Series": "SKU"},
                category_orders={"Month": trend_data["Month"].unique().tolist()},
            )
            fig.update_layout(xaxis_title=None, margin=dict(t=20), legend_title="SKU")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Select at least one SKU above to plot the comparison.")
    st.caption(
        f"⚠️ {today.strftime('%b %Y')} shows projected month-end "
        f"({max(today.day - 1, 1)} days elapsed, {mult:.1f}× multiplier)."
    )

    # ── Theme Comparison ──────────────────────────────────────────────────────
    st.subheader("Theme Comparison")
    THEMES = {
        "Milestone Gifts":   ["TCB005", "TCB006", "TCB010"],
        "Parent Gifts":      ["TCB007", "TCB009", "TCB009_1", "TCB009_2"],
        "Premium Gifts":     ["TCB003", "TCB004", "TCB001", "TCB002"],
        "Entry Price Gifts": ["TCB008", "TCB011", "TCB012"],
    }
    theme_metric = st.radio("Metric", ["Units", "Revenue"], horizontal=True, key="theme_metric")
    theme_val_col = "quantity" if theme_metric == "Units" else "gross_value"
    theme_y_label = "Units Sold" if theme_metric == "Units" else "Gross Revenue (₹)"

    # Use full gross-status data for revenue; adf (net) for units
    theme_base = fdf[fdf["status"].isin(GROSS_STATUSES)].copy() if theme_metric == "Revenue" else adf.copy()

    theme_rows = []
    for theme_name, sku_ids in THEMES.items():
        grp = (
            theme_base[theme_base["sku_id"].isin(sku_ids)]
            .groupby("month_dt")[theme_val_col].sum()
            .reset_index()
        )
        grp["Theme"] = theme_name
        theme_rows.append(grp)

    if theme_rows:
        theme_df = pd.concat(theme_rows, ignore_index=True)
        theme_df = _project_col(theme_df, theme_val_col, mult)
        theme_df["Month"] = theme_df["month_dt"].dt.strftime("%b %Y")
        theme_df = theme_df.sort_values("month_dt")
        fig_th = px.line(
            theme_df, x="Month", y=theme_val_col,
            color="Theme", markers=True,
            labels={theme_val_col: theme_y_label, "Theme": "Theme"},
            category_orders={"Month": theme_df["Month"].unique().tolist()},
        )
        fig_th.update_layout(xaxis_title=None, margin=dict(t=20), legend_title="Theme")
        st.plotly_chart(fig_th, use_container_width=True)
    st.caption(
        f"⚠️ {today.strftime('%b %Y')} shows projected month-end "
        f"({max(today.day - 1, 1)} days elapsed, {mult:.1f}× multiplier)."
    )

    st.subheader("SKU × Channel Heatmap (Units)")
    heat = (
        adf.groupby(["sku_name", "channel_name"])["quantity"]
        .sum().unstack("channel_name").fillna(0).astype(int)
    )
    if not heat.empty:
        # Global max across all SKU/channel cells (excluding Total) for consistent shading
        global_max = int(heat.values.max()) or 1

        total_heat = heat.sum(axis=0).rename("TOTAL")
        heat = pd.concat([heat, total_heat.to_frame().T])
        heat.index.name = "SKUs"  # set after concat so name is preserved

        is_total_heat = [False] * (len(heat) - 1) + [True]

        def _global_red_excl_total(col):
            return pd.Series(
                [
                    "" if flag
                    else f"background-color: rgba(255, 100, 0, {min(v / global_max, 1) * 0.6:.2f})"
                    for flag, v in zip(is_total_heat, col)
                ],
                index=col.index,
            )

        st.dataframe(
            heat.style
            .apply(_global_red_excl_total, axis=0)
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if row.name == "TOTAL" else [""] * len(row),
                axis=1,
            ),
            use_container_width=True,
        )


# ── Tab 5 — Returns ────────────────────────────────────────────────────────────
def tab_returns(raw_df: pd.DataFrame, fdf: pd.DataFrame, filters: dict = None):
    st.warning("**This tab shows ALL order statuses including Cancelled. All other tabs exclude Cancelled orders.**")

    if not raw_df.empty and not fdf.empty:
        _min_month = fdf["month_dt"].min()
        _max_month = fdf["month_dt"].max()
        all_fdf = raw_df[
            raw_df["channel_name"].isin(fdf["channel_name"].unique()) &
            raw_df["sku_id"].isin(fdf["sku_id"].unique()) &
            (raw_df["month_dt"] >= _min_month) &
            (raw_df["month_dt"] <= _max_month)
        ]
        all_fdf = _apply_city_filter(all_fdf, (filters or {}).get("cities", ["__all__"]))
    else:
        all_fdf = fdf

    # ── Status breakdown — pie + table ────────────────────────────────────────
    st.subheader("Order Status Breakdown")
    # For CANCELLED orders Amazon often sends qty=0; treat each such row as 1 unit
    _counts_df = all_fdf.copy()
    _cancelled_zero = (_counts_df["status"] == "CANCELLED") & (_counts_df["quantity"].fillna(0) <= 0)
    _counts_df.loc[_cancelled_zero, "quantity"] = 1

    status_counts = _counts_df.groupby("status").agg(
        Orders=("order_id", "count"),
        Units=("quantity", "sum"),
    ).reindex(STATUS_ORDER).dropna(how="all").fillna(0).astype(int)
    status_counts["% of Orders"] = (status_counts["Orders"] / status_counts["Orders"].sum() * 100).round(1)

    col_pie, col_tbl = st.columns([1, 1])
    with col_pie:
        fig_pie = px.pie(
            status_counts.reset_index(),
            values="Orders", names="status",
            color_discrete_sequence=px.colors.qualitative.Set2,
            hole=0.35,
        )
        fig_pie.update_traces(
            texttemplate="%{label}<br>%{percent:.1%}",
            hovertemplate="<b>%{label}</b><br>Orders: %{value:,}<br>Share: %{percent:.1%}<extra></extra>",
            textposition="inside",
            insidetextorientation="horizontal",
            pull=[0.03] * len(status_counts),
        )
        fig_pie.update_layout(
            showlegend=True,
            uniformtext_minsize=11,
            uniformtext_mode="hide",   # hide labels that are too small to fit inside — legend covers them
            margin=dict(t=20, b=20, l=20, r=20),
        )
        st.plotly_chart(fig_pie, use_container_width=True)
    with col_tbl:
        total_sc = pd.DataFrame(
            [{"Orders": status_counts["Orders"].sum(),
              "Units":  status_counts["Units"].sum(),
              "% of Orders": 100.0}],
            index=["TOTAL"],
        )
        sc_with_total = pd.concat([status_counts, total_sc])
        st.dataframe(
            sc_with_total.style
            .format({"% of Orders": "{:.1f}%"})
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if row.name == "TOTAL" else [""] * len(row),
                axis=1,
            ),
            use_container_width=True,
        )

    def _bold_total_row(df):
        last = len(df) - 1
        return df.style.apply(
            lambda row: ["font-weight: bold"] * len(row) if row.name == last else [""] * len(row),
            axis=1,
        )

    # ── Cancellation by channel ───────────────────────────────────────────────
    st.subheader("Cancellation by Channel")
    canc_rows = []
    for ch, g in all_fdf.groupby("channel_name"):
        total     = len(g)
        cancelled = int((g["status"] == "CANCELLED").sum())
        canc_rows.append({
            "Channel":       ch,
            "Total Orders":  total,
            "Cancelled":     cancelled,
            "_sort":         pct(cancelled, total),
        })
    canc = pd.DataFrame(canc_rows).sort_values("_sort", ascending=False)
    canc["Cancel Rate %"] = canc["_sort"]  # keep numeric for heatmap
    canc = canc.drop(columns=["_sort"])
    t_orders = canc["Total Orders"].sum()
    t_canc   = canc["Cancelled"].sum()
    canc = pd.concat([canc, pd.DataFrame([{
        "Channel": "TOTAL", "Total Orders": t_orders,
        "Cancelled": t_canc, "Cancel Rate %": pct(t_canc, t_orders),
    }])], ignore_index=True)
    _is_total_canc = [False] * (len(canc) - 1) + [True]

    def _cancel_rate_bg(col):
        max_val = max((v for flag, v in zip(_is_total_canc, col) if not flag), default=1) or 1
        return pd.Series(
            ["" if flag else f"background-color: rgba(165, 55, 55, {min(v / max_val, 1) * 0.55:.2f})"
             for flag, v in zip(_is_total_canc, col)],
            index=col.index,
        )

    st.dataframe(
        _bold_total_row(canc)
        .apply(_cancel_rate_bg, subset=["Cancel Rate %"])
        .format({"Cancel Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    # ── Shared return base + filters ─────────────────────────────────────────
    ret_base = all_fdf[all_fdf["status"].isin(RETURN_STATUSES)].copy()
    for _col in ("return_reason", "return_responsible", "return_customer_verbatim"):
        if _col not in ret_base.columns:
            ret_base[_col] = "Unknown"
        ret_base[_col] = ret_base[_col].fillna("Unknown")

    st.subheader("Return Rate by Channel")
    f1, f2 = st.columns(2)
    with f1:
        sel_reason = st.multiselect(
            "Filter by Reason",
            sorted(ret_base["return_reason"].unique()),
            default=[],
            key="ret_reason_filter",
        )
    with f2:
        sel_resp = st.multiselect(
            "Filter by Responsible",
            sorted(ret_base["return_responsible"].unique()),
            default=[],
            key="ret_resp_filter",
        )

    ret_filtered = ret_base.copy()
    if sel_reason:
        ret_filtered = ret_filtered[ret_filtered["return_reason"].isin(sel_reason)]
    if sel_resp:
        ret_filtered = ret_filtered[ret_filtered["return_responsible"].isin(sel_resp)]

    ret_tcb = ret_base[ret_base["return_responsible"] == "TCB"]

    if sel_reason or sel_resp:
        st.caption(
            f"Filters active — showing {len(ret_filtered):,} of {len(ret_base):,} return orders. "
            "Return Rate % uses all dispatched orders as denominator."
        )

    def _filt_rate(ch_filt, ch_ret_base, grp):
        denom = int((grp["status"] == "FULFILLED").sum()) + len(ch_ret_base)
        return pct(len(ch_filt), denom)

    def _tcb_rate(ch_tcb, ch_ret_base, grp):
        denom = int((grp["status"] == "FULFILLED").sum()) + len(ch_ret_base)
        return pct(len(ch_tcb), denom)

    # ── Return rate by channel ────────────────────────────────────────────────
    ret_rows = []
    for ch, grp in all_fdf.groupby("channel_name"):
        ch_ret  = ret_base[ret_base["channel_name"] == ch]
        ch_filt = ret_filtered[ret_filtered["channel_name"] == ch]
        ch_tcb  = ret_tcb[ret_tcb["channel_name"] == ch]
        ret_rows.append({
            "Channel":           ch,
            "Fulfilled":         int((grp["status"] == "FULFILLED").sum()),
            "RTO":               int((ch_filt["status"] == "RTO").sum()),
            "Sale Return":       int((ch_filt["status"] == "SALE_RETURN").sum()),
            "Replacement":       int((ch_filt["status"] == "REPLACEMENT").sum()),
            "Return Rate %":     _filt_rate(ch_filt, ch_ret, grp),
            "TCB Responsible Return Rate %": _tcb_rate(ch_tcb, ch_ret, grp),
        })
    ret_df = pd.DataFrame(ret_rows)
    ret_df = pd.concat([ret_df, pd.DataFrame([{
        "Channel":           "TOTAL",
        "Fulfilled":         int(ret_df["Fulfilled"].sum()),
        "RTO":               int(ret_df["RTO"].sum()),
        "Sale Return":       int(ret_df["Sale Return"].sum()),
        "Replacement":       int(ret_df["Replacement"].sum()),
        "Return Rate %":     _filt_rate(ret_filtered, ret_base, all_fdf),
        "TCB Responsible Return Rate %": _tcb_rate(ret_tcb, ret_base, all_fdf),
    }])], ignore_index=True)
    _is_total_ret = [False] * (len(ret_df) - 1) + [True]

    def _rr_bg(col, is_total_flags):
        max_val = max((v for flag, v in zip(is_total_flags, col) if not flag), default=1) or 1
        return pd.Series(
            ["" if flag else f"background-color: rgba(165, 55, 55, {min(v / max_val, 1) * 0.55:.2f})"
             for flag, v in zip(is_total_flags, col)],
            index=col.index,
        )

    st.dataframe(
        _bold_total_row(ret_df)
        .apply(_rr_bg, is_total_flags=_is_total_ret, subset=["Return Rate %", "TCB Responsible Return Rate %"])
        .format({"Return Rate %": "{:.1f}%", "TCB Responsible Return Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    # ── Return rate by SKU ────────────────────────────────────────────────────
    st.subheader("Return Rate by SKU")
    ret_sku_rows = []
    for sku, grp in all_fdf.groupby("sku_id"):
        sku_ret  = ret_base[ret_base["sku_id"] == sku]
        sku_filt = ret_filtered[ret_filtered["sku_id"] == sku]
        sku_tcb  = ret_tcb[ret_tcb["sku_id"] == sku]
        ret_sku_rows.append({
            "SKU":               sku,
            "SKU Name":          grp["sku_name"].iloc[0],
            "Fulfilled":         int((grp["status"] == "FULFILLED").sum()),
            "RTO":               int((sku_filt["status"] == "RTO").sum()),
            "Sale Return":       int((sku_filt["status"] == "SALE_RETURN").sum()),
            "Replacement":       int((sku_filt["status"] == "REPLACEMENT").sum()),
            "Return Rate %":     _filt_rate(sku_filt, sku_ret, grp),
            "TCB Responsible Return Rate %": _tcb_rate(sku_tcb, sku_ret, grp),
        })
    ret_sku_df = pd.DataFrame(ret_sku_rows)
    ret_sku_df = pd.concat([ret_sku_df, pd.DataFrame([{
        "SKU":               "TOTAL",
        "SKU Name":          "",
        "Fulfilled":         int(ret_sku_df["Fulfilled"].sum()),
        "RTO":               int(ret_sku_df["RTO"].sum()),
        "Sale Return":       int(ret_sku_df["Sale Return"].sum()),
        "Replacement":       int(ret_sku_df["Replacement"].sum()),
        "Return Rate %":     _filt_rate(ret_filtered, ret_base, all_fdf),
        "TCB Responsible Return Rate %": _tcb_rate(ret_tcb, ret_base, all_fdf),
    }])], ignore_index=True)
    _is_total_ret_sku = [False] * (len(ret_sku_df) - 1) + [True]

    st.dataframe(
        _bold_total_row(ret_sku_df)
        .apply(_rr_bg, is_total_flags=_is_total_ret_sku, subset=["Return Rate %", "TCB Responsible Return Rate %"])
        .format({"Return Rate %": "{:.1f}%", "TCB Responsible Return Rate %": "{:.1f}%"}),
        use_container_width=True, hide_index=True,
    )

    # ── Return reasons ────────────────────────────────────────────────────────
    st.subheader("Return Reasons")
    rdf = ret_base.copy()
    reasons_grp = (
        rdf.groupby(["return_reason", "return_responsible"])
        .agg(Orders=("order_id", "count"), Units=("quantity", "sum"))
        .reset_index()
        .sort_values("Orders", ascending=False)
    )
    if not reasons_grp.empty:
        reasons_grp["% of Returns"] = (reasons_grp["Orders"] / reasons_grp["Orders"].sum() * 100).round(1)
        reasons_grp.columns = ["Reason", "Responsible", "Orders", "Units", "% of Returns"]
        total_reason = pd.DataFrame([{
            "Reason":         "TOTAL",
            "Responsible":    "",
            "Orders":         int(reasons_grp["Orders"].sum()),
            "Units":          int(reasons_grp["Units"].sum()),
            "% of Returns":   100.0,
        }])
        reasons_display = pd.concat([reasons_grp, total_reason], ignore_index=True)
        st.dataframe(
            _bold_total_row(reasons_display)
            .format({"% of Returns": "{:.1f}%"}),
            use_container_width=True, hide_index=True,
        )

        # Raw data download
        pid_col = "platform_order_id" if "platform_order_id" in rdf.columns else None
        raw_dl = rdf[[
            "channel_name",
            *(["platform_order_id"] if pid_col else []),
            "order_date",
            "return_reason",
            "return_responsible",
            "return_customer_verbatim",
        ]].copy()
        raw_dl.columns = [
            "Channel Name",
            *(["Order Number"] if pid_col else []),
            "Order Date",
            "Return Reason",
            "Responsible",
            "Customer Verbatim",
        ]
        st.download_button(
            "⬇ Download raw return data",
            raw_dl.to_csv(index=False),
            file_name="returns_raw.csv",
            mime="text/csv",
        )
    else:
        st.info("No return data in selected filters.")

    # ── Monthly return rate trend ─────────────────────────────────────────────
    st.subheader("Monthly Return Rate Trend")
    monthly_status = all_fdf.groupby(["month_dt", "status"])["quantity"].sum().unstack("status").fillna(0)
    for s in ALL_STATUSES:
        if s not in monthly_status.columns:
            monthly_status[s] = 0

    denom = (
        monthly_status["FULFILLED"] + monthly_status["RTO"]
        + monthly_status["SALE_RETURN"] + monthly_status["REPLACEMENT"]
    ).replace(0, pd.NA)

    monthly_status["return_rate"] = (
        (monthly_status["RTO"] + monthly_status["SALE_RETURN"] + monthly_status["REPLACEMENT"])
        / denom * 100
    ).fillna(0).round(1)
    monthly_status["rto_rate"] = (monthly_status["RTO"] / denom * 100).fillna(0).round(1)
    monthly_status["sr_rate"]  = (monthly_status["SALE_RETURN"] / denom * 100).fillna(0).round(1)

    trend_plot = monthly_status[["return_rate", "rto_rate", "sr_rate"]].reset_index()
    trend_plot = trend_plot.sort_values("month_dt")
    trend_plot["Month"] = trend_plot["month_dt"].dt.strftime("%b %Y")
    fig_rt = go.Figure()
    fig_rt.add_trace(go.Scatter(x=trend_plot["Month"], y=trend_plot["return_rate"],
                                name="Overall Return Rate %", mode="lines+markers"))
    fig_rt.add_trace(go.Scatter(x=trend_plot["Month"], y=trend_plot["rto_rate"],
                                name="RTO Rate %", mode="lines+markers", line=dict(dash="dash")))
    fig_rt.add_trace(go.Scatter(x=trend_plot["Month"], y=trend_plot["sr_rate"],
                                name="Sale Return Rate %", mode="lines+markers", line=dict(dash="dot")))
    fig_rt.update_layout(yaxis_title="Rate %", xaxis_title=None, margin=dict(t=20))
    st.plotly_chart(fig_rt, use_container_width=True)


# ── Geography tab ──────────────────────────────────────────────────────────────
def tab_geography(fdf: pd.DataFrame, net_mode: bool) -> None:
    """6th tab: city/state breakdown across channels."""

    # All non-cancelled filtered orders (apply_filters already drops CANCELLED)
    gdf = active_df(fdf, net_mode).copy()

    # In-memory normalisation (mirrors ingest/utils.py — catches any stragglers)
    _CITY_NORM  = {
        "Delhi": "New Delhi", "Bangalore": "Bengaluru",
        "Gurugram": "Gurgaon", "Bengalore": "Bengaluru", "Bangaluru": "Bengaluru",
        "Benagluru": "Bengaluru", "Bangalure": "Bengaluru", "Bengalure": "Bengaluru",
        "Vishakhapatnam": "Visakhapatnam", "Vishakhapatanam": "Visakhapatnam",
    }
    _STATE_NORM = {
        "Tamilnadu": "Tamil Nadu", "Asom": "Assam", "Asom (Assam)": "Assam",
        "Orissa": "Odisha", "Pondicherry": "Puducherry", "Uttaranchal": "Uttarakhand",
    }
    # Groupings for Top Cities chart + table only — raw city data is never changed
    _CITY_GROUPS = {
        "Bengaluru":         "Bengaluru",
        # NCR
        "New Delhi":         "NCR",  "Gurgaon":           "NCR",
        "Ghaziabad":         "NCR",  "Faridabad":         "NCR",
        "Noida":             "NCR",  "Greater Noida":     "NCR",
        "East Delhi":        "NCR",  "Central Delhi":     "NCR",
        "North Delhi":       "NCR",  "South Delhi":       "NCR",
        "West Delhi":        "NCR",  "North East Delhi":  "NCR",
        "North West Delhi":  "NCR",  "South West Delhi":  "NCR",
        "Shahdara":          "NCR",
        "Gautam Buddha Nagar": "NCR", "Manesar":          "NCR",
        "Loni":              "NCR",  "Sohna":             "NCR",
        "Bhiwadi":           "NCR",
        # Greater Mumbai
        "Mumbai":            "Greater Mumbai",  "Navi Mumbai":       "Greater Mumbai",
        "Thane":             "Greater Mumbai",  "Thane West":        "Greater Mumbai",
        "Kalyan":            "Greater Mumbai",  "Badlapur":          "Greater Mumbai",
        "Mira-Bhayandar":    "Greater Mumbai",  "Mira Bhayandar":    "Greater Mumbai",
        "Mira Bhayander":    "Greater Mumbai",
        "Vasai-Virar":       "Greater Mumbai",  "Vasai Virar":       "Greater Mumbai",
        "Vasai":             "Greater Mumbai",  "Virar":             "Greater Mumbai",
        "Panvel":            "Greater Mumbai",  "Bhiwandi-Nizampur": "Greater Mumbai",
        "Bhiwandi":          "Greater Mumbai",  "Mumbra":            "Greater Mumbai",
        # Greater Pune
        "Pune":              "Greater Pune",    "Pimpri Chinchwad":  "Greater Pune",
        "Pimpri-Chinchwad":  "Greater Pune",
    }
    gdf["city"]  = gdf["city"].str.strip().str.title().replace(_CITY_NORM)
    gdf["state"] = gdf["state"].str.strip().str.title().replace(_STATE_NORM)

    geo = gdf[gdf["city"].notna() & gdf["state"].notna()].copy()
    geo["display_city"] = geo["city"].map(_CITY_GROUPS).fillna(geo["city"])

    # ── Coverage callout ──────────────────────────────────────────────────────
    total_rows = len(gdf)
    geo_rows   = len(geo)
    coverage   = pct(geo_rows, total_rows)
    st.info(
        f"All orders that have geo data (currently {coverage:.0f}%) are included in the charts below. "
        f"Cancelled orders are excluded from this tab."
    )

    # ── KPI strip ─────────────────────────────────────────────────────────────
    n_cities  = geo["city"].nunique()
    n_states  = geo["state"].nunique()
    top_city  = geo.groupby("city")["quantity"].sum().idxmax()  if n_cities else "—"
    top_state = geo.groupby("state")["quantity"].sum().idxmax() if n_states else "—"

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Cities Reached", n_cities)
    c2.metric("States Reached", n_states)
    c3.metric("Top City",       top_city)
    c4.metric("Top State",      top_state)

    if geo.empty:
        st.info("No orders with city/state data for the selected filters.")
        return

    st.markdown("---")

    # ── Metric selector ───────────────────────────────────────────────────────
    metric = st.radio(
        "Metric", ["Units", "Orders", "Revenue"], horizontal=True, key="geo_metric"
    )
    if metric == "Units":
        value_col = "quantity"
        y_label   = "Units"
    elif metric == "Orders":
        geo = geo.copy()
        geo["_orders"] = 1
        value_col = "_orders"
        y_label   = "Orders"
    else:
        value_col = "gross_value"
        y_label   = "Revenue (₹)"

    def _bar_totals(fig, grouped_df, x_col):
        """Add totals as a text-mode scatter trace above each stacked bar."""
        totals = grouped_df.groupby(x_col)[value_col].sum().reset_index()
        totals.columns = [x_col, "_tot"]
        text_vals = [
            fmt_inr(v) if metric == "Revenue" else f"{int(v):,}"
            for v in totals["_tot"]
        ]
        fig.add_trace(go.Scatter(
            x=totals[x_col], y=totals["_tot"],
            mode="text", text=text_vals,
            textposition="top center", showlegend=False,
            textfont=dict(size=9, color="#333"),
        ))
        max_v = totals["_tot"].max() or 1
        fig.update_layout(yaxis_range=[0, max_v * 1.18])

    # ── Top Cities chart ──────────────────────────────────────────────────────
    city_ch = (
        geo.groupby(["display_city", "channel_name"])[value_col].sum()
        .reset_index()
    )
    top15_cities = city_ch.groupby("display_city")[value_col].sum().nlargest(15).index
    city_ch = city_ch[city_ch["display_city"].isin(top15_cities)]
    city_order = (
        city_ch.groupby("display_city")[value_col].sum()
        .sort_values(ascending=False).index.tolist()
    )

    st.subheader("Top Cities")
    if not city_ch.empty:
        fig_city = px.bar(
            city_ch,
            x="display_city", y=value_col, color="channel_name",
            barmode="stack",
            labels={value_col: y_label, "channel_name": "Channel", "display_city": "City"},
            color_discrete_map=CHANNEL_COLOR_MAP,
            category_orders={"display_city": city_order},
        )
        fig_city.update_layout(
            xaxis_title=None, margin=dict(t=30),
            xaxis=dict(tickangle=-30),
        )
        if metric == "Revenue":
            fig_city.update_layout(yaxis=dict(tickformat=",.0f", tickprefix="₹"))
        _bar_totals(fig_city, city_ch, "display_city")
        st.plotly_chart(fig_city, use_container_width=True)

    # ── City trend table ──────────────────────────────────────────────────────
    today = date.today()
    cur_p = pd.Period(today, "M")
    m1_p  = cur_p - 1
    m2_p  = cur_p - 2
    m3_p  = cur_p - 3
    mult  = _cur_mult()

    city_monthly = (
        geo.groupby(["display_city", "month"])["quantity"].sum()
        .reset_index()
        .rename(columns={"quantity": "units", "display_city": "city"})
    )

    def _period_units(period):
        sub = city_monthly[city_monthly["month"] == period]
        return sub.set_index("city")["units"]

    cm_raw = _period_units(cur_p)
    m1_u   = _period_units(m1_p)
    m2_u   = _period_units(m2_p)
    m3_u   = _period_units(m3_p)
    cm_proj = (cm_raw * mult).round(0)

    def _vs_pct(cur, ref):
        if ref and ref > 0:
            return round((cur - ref) / ref * 100, 1)
        return None

    def _fmt_vs(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "—"
        return f"+{v:.1f}%" if v >= 0 else f"{v:.1f}%"

    def _color_vs(val):
        if not isinstance(val, str) or val == "—":
            return ""
        try:
            v = float(val.replace("%", "").replace("+", ""))
            if v > 5:
                return "color: green; font-weight: bold"
            if v < -5:
                return "color: red; font-weight: bold"
        except ValueError:
            pass
        return ""

    cm_col = f"CM ({today.strftime('%b')})"
    tbl_rows = []
    for city in sorted(geo["display_city"].unique()):
        cm_v  = float(cm_proj.get(city, 0) or 0)
        m1_v  = float(m1_u.get(city, 0) or 0)
        m2_v  = float(m2_u.get(city, 0) or 0)
        m3_v  = float(m3_u.get(city, 0) or 0)
        l3m_v = (m1_v + m2_v + m3_v) / 3
        tbl_rows.append({
            "City":      city,
            cm_col:      int(cm_v),
            "M-1":       int(m1_v),
            "vs. M-1":   _fmt_vs(_vs_pct(cm_v, m1_v)),
            "M-2":       int(m2_v),
            "vs. M-2":   _fmt_vs(_vs_pct(cm_v, m2_v)),
            "L3M Avg":   round(l3m_v, 1),
            "vs. L3M":   _fmt_vs(_vs_pct(cm_v, l3m_v)),
        })

    city_tbl = pd.DataFrame(tbl_rows).sort_values(cm_col, ascending=False)
    vs_cols  = ["vs. M-1", "vs. M-2", "vs. L3M"]
    st.caption(f"Units only. {cm_col} projected to full month.")
    st.dataframe(
        city_tbl.style.map(_color_vs, subset=vs_cols),
        use_container_width=True,
        height=350,
        hide_index=True,
    )

    # ── Top States chart ───────────────────────────────────────────────────────
    state_ch = (
        geo.groupby(["state", "channel_name"])[value_col].sum()
        .reset_index()
    )
    state_order = (
        state_ch.groupby("state")[value_col].sum()
        .sort_values(ascending=False).index.tolist()
    )

    st.subheader("Top States")
    if not state_ch.empty:
        fig_state = px.bar(
            state_ch,
            x="state", y=value_col, color="channel_name",
            barmode="stack",
            labels={value_col: y_label, "channel_name": "Channel", "state": "State"},
            color_discrete_map=CHANNEL_COLOR_MAP,
            category_orders={"state": state_order},
        )
        fig_state.update_layout(
            xaxis_title=None, margin=dict(t=30),
            xaxis=dict(tickangle=-40),
            height=450,
        )
        if metric == "Revenue":
            fig_state.update_layout(yaxis=dict(tickformat=",.0f", tickprefix="₹"))
        _bar_totals(fig_state, state_ch, "state")
        st.plotly_chart(fig_state, use_container_width=True)

    # ── City-wise Trend ────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("City-wise Trend")

    _TREND_GROUPS = {
        "Bengaluru":          "Bengaluru",
        "New Delhi":          "NCR",           "Gurgaon":           "NCR",
        "Ghaziabad":          "NCR",           "Faridabad":         "NCR",
        "Noida":              "NCR",           "Greater Noida":     "NCR",
        "East Delhi":         "NCR",           "Central Delhi":     "NCR",
        "North Delhi":        "NCR",           "South Delhi":       "NCR",
        "West Delhi":         "NCR",           "North East Delhi":  "NCR",
        "North West Delhi":   "NCR",           "South West Delhi":  "NCR",
        "Shahdara":           "NCR",
        "Gautam Buddha Nagar": "NCR",          "Manesar":           "NCR",
        "Loni":               "NCR",           "Sohna":             "NCR",
        "Bhiwadi":            "NCR",
        "Hyderabad":          "Hyderabad",     "Secunderabad":      "Hyderabad",
        "Mumbai":             "Greater Mumbai", "Navi Mumbai":      "Greater Mumbai",
        "Thane":              "Greater Mumbai", "Thane West":       "Greater Mumbai",
        "Kalyan":             "Greater Mumbai", "Badlapur":         "Greater Mumbai",
        "Mira-Bhayandar":     "Greater Mumbai", "Mira Bhayandar":  "Greater Mumbai",
        "Mira Bhayander":     "Greater Mumbai",
        "Vasai-Virar":        "Greater Mumbai", "Vasai Virar":      "Greater Mumbai",
        "Vasai":              "Greater Mumbai", "Virar":            "Greater Mumbai",
        "Panvel":             "Greater Mumbai", "Bhiwandi-Nizampur":"Greater Mumbai",
        "Bhiwandi":           "Greater Mumbai", "Mumbra":           "Greater Mumbai",
        "Chennai":            "Chennai",
        "Pune":               "Greater Pune",  "Pimpri Chinchwad":  "Greater Pune",
        "Pimpri-Chinchwad":   "Greater Pune",
        "Ahmedabad":          "Ahmedabad",
        "Kolkata":            "Kolkata",
    }
    _GROUP_ORDER = [
        "Bengaluru", "NCR", "Hyderabad", "Greater Mumbai",
        "Chennai", "Greater Pune", "Ahmedabad", "Kolkata", "All Others",
    ]

    _trend_metric = st.radio(
        "Metric", ["Units", "Revenue"], horizontal=True, key="geo_city_trend_metric"
    )
    _trend_vcol = "quantity" if _trend_metric == "Units" else "gross_value"

    _geo_tr = geo.copy()
    _geo_tr["city_group"] = _geo_tr["city"].map(_TREND_GROUPS).fillna("All Others")

    _trend_agg = (
        _geo_tr.groupby(["month_dt", "city_group"])[_trend_vcol]
        .sum()
        .reset_index()
    )
    _trend_agg["Month"] = _trend_agg["month_dt"].dt.strftime("%b %Y")
    _trend_agg = _project_col(_trend_agg, _trend_vcol, _cur_mult())
    _month_order = _trend_agg.sort_values("month_dt")["Month"].unique().tolist()

    if not _trend_agg.empty:
        _fig_ct = px.line(
            _trend_agg,
            x="Month", y=_trend_vcol, color="city_group",
            markers=True,
            labels={_trend_vcol: _trend_metric, "city_group": "City"},
            category_orders={"Month": _month_order, "city_group": _GROUP_ORDER},
        )
        _fig_ct.update_layout(
            xaxis_title=None,
            yaxis_title=_trend_metric,
            margin=dict(t=10),
            height=420,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        if _trend_metric == "Revenue":
            _fig_ct.update_layout(yaxis=dict(tickformat=",.0f", tickprefix="₹"))
        st.caption("Current month projected to full month")
        st.plotly_chart(_fig_ct, use_container_width=True)


# ── Auth gate ──────────────────────────────────────────────────────────────────
def _require_auth() -> None:
    """Password gate for public deployment. No-op if APP_PASSWORD secret is not set."""
    try:
        pwd_secret = st.secrets.get("APP_PASSWORD", "")
    except Exception:
        return  # secrets not configured — local dev, skip gate

    if not pwd_secret:
        return  # secret present but empty — treat as disabled

    if st.session_state.get("_auth_ok"):
        return

    st.markdown("## Growth Spurt Dashboard")
    pwd = st.text_input("Password", type="password", key="_pwd")
    if st.button("Enter"):
        if pwd == pwd_secret:
            st.session_state["_auth_ok"] = True
            st.rerun()
        else:
            st.error("Incorrect password")
    st.stop()


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    _require_auth()

    import base64
    logo_path = os.path.join(os.path.dirname(__file__), "..", "assets", "logo.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as _f:
            logo_b64 = base64.b64encode(_f.read()).decode()
        st.markdown(
            f"""<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
                <img src="data:image/png;base64,{logo_b64}" width="70"/>
                <h2 style="margin:0;padding:0;">Growth Spurt Dashboard</h2>
            </div>""",
            unsafe_allow_html=True,
        )
    else:
        st.markdown("## Growth Spurt Dashboard")

    raw_df = load_data()
    if raw_df.empty:
        st.warning("No orders data found. Run an ingest script first.")
        return

    filters = sidebar(raw_df)
    fdf     = apply_filters(raw_df, filters)

    if fdf.empty:
        st.info("No data for the selected filters.")
        return

    t1, t2, t3, t4, t5, t6, t7, t8 = st.tabs([
        "📊 Overview", "📈 Trends", "🏪 By Channel", "📦 By SKU",
        "🔄 Returns", "🗺️ Geo", "🟡 Blinkit Deepdive", "🔮 Forecast",
    ])
    with t1:
        tab_overview(raw_df, fdf, filters["net_mode"], filters)
    with t2:
        tab_trends(fdf, filters["net_mode"])
    with t3:
        tab_channel(fdf, filters["net_mode"])
    with t4:
        tab_sku(fdf, filters["net_mode"])
    with t5:
        tab_returns(raw_df, fdf, filters)
    with t6:
        tab_geography(fdf, filters["net_mode"])
    with t7:
        tab_blinkit_deepdive(fdf, filters["net_mode"])
    with t8:
        tab_forecast()


# ── Blinkit Deepdive tab ───────────────────────────────────────────────────────

_STATUS_META = [
    ("active",                  "Active",              "#22C55E"),
    ("launch_awaited",          "Not launched",        "#F97316"),
    ("sku_moved_out_low_sales",  "Inactive (low sale)", "#EAB308"),
    ("ds_choked",               "Choked",              "#06B6D4"),
    ("sku_city_exited",         "City exit",           "#EF4444"),
    ("sku_recalled",            "Recalled",            "#7C3AED"),
    ("darkstore_closed",        "DS closed",           "#94A3B8"),
    ("no_data",                 "Not deployed",        "#CBD5E1"),
]
_STATUS_LABEL = {k: lbl for k, lbl, _ in _STATUS_META}
_STATUS_COLOR = {k: col for k, _, col in _STATUS_META}


@st.cache_data(ttl=300)
def _load_city_ds(sku_id: str) -> pd.DataFrame:
    rows = get_blinkit_city_ds(sku_id)
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


_REPLEN_PARQUET = Path("data/blinkit/auto/replenishment/replenishment_plan_latest.parquet")


@st.cache_data(show_spinner=False)
def _load_replen_plan(_mtime: float) -> "pd.DataFrame":
    """Read pre-computed replen plan from parquet. Cache busts only when the file changes."""
    if _REPLEN_PARQUET.exists():
        return pd.read_parquet(_REPLEN_PARQUET)
    return pd.DataFrame()


@st.cache_data(ttl=300, show_spinner=False)
def _load_replen_plan_from_db() -> "pd.DataFrame":
    """Fallback: fetch latest replen plan from Supabase (used on Streamlit Cloud)."""
    rows = get_replen_plan()
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def _get_replen_plan() -> "pd.DataFrame":
    if _REPLEN_PARQUET.exists():
        mtime = _REPLEN_PARQUET.stat().st_mtime
        return _load_replen_plan(mtime)
    return _load_replen_plan_from_db()


def _load_wh_summary(sku_id: str) -> list:
    plan_df = _get_replen_plan()
    if plan_df.empty:
        return []
    sku_rows = plan_df[plan_df["sku_id"] == sku_id]
    result = []
    for _, r in sku_rows.iterrows():
        result.append({
            "wh_name":         r["wh_name"],
            "active_ds":       int(r["active_ds_count"]),
            "choked_ds_count": int(r.get("ds_choked_count", 0)),
            "wh_stock":        int(r["units_wh"]) + int(r["units_incoming"]),
            "ds_stock":        int(r["units_ds"]) + int(r["units_transit"]),
            "total_inventory": int(r["effective_stock"]),
            "target_stock":    int(r["target_stock"]),
            "oos_ds_count":    int(r["ds_with_oos"]),
        })
    return sorted(result, key=lambda x: x["wh_name"])


@st.cache_data(ttl=300)
def _load_wh_list() -> list:
    db = get_client()
    rows = (db.table("partner_locations")
              .select("location_id,name,code")
              .eq("channel_id", 4)
              .eq("location_type", "WH")
              .eq("is_active", True)
              .execute().data)
    return sorted(rows, key=lambda r: r["name"])


@st.cache_data(ttl=300)
def _load_all_wh_names() -> dict:
    """location_id -> name for ALL Blinkit WH rows, including deactivated ones —
    used for display labels (e.g. Launch Status' Serving WH column), where a DS
    can still point at a since-deactivated WH and should show its real name."""
    db = get_client()
    rows = (db.table("partner_locations")
              .select("location_id,name")
              .eq("channel_id", 4)
              .eq("location_type", "WH")
              .execute().data)
    return {r["location_id"]: r["name"] for r in rows}


@st.cache_data(ttl=300)
def _load_all_ds_raw() -> list:
    db = get_client()
    return (db.table("partner_locations")
              .select("location_id,name,city,parent_location_id,is_active")
              .eq("channel_id", 4)
              .eq("location_type", "DARKSTORE")
              .eq("is_active", True)
              .limit(2000)
              .execute().data)


@st.cache_data(ttl=300)
def _load_all_elig() -> list:
    db = get_client()
    rows, offset, page_size = [], 0, 1000
    while True:
        batch = (db.table("blinkit_ds_sku_eligibility")
                   .select("location_id,sku_id,status")
                   .range(offset, offset + page_size - 1)
                   .execute().data)
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


@st.cache_data(ttl=300, show_spinner=False)
def _load_latest_stock_flags(sku_id: str, location_ids: tuple, days_back: int = 20) -> dict:
    """
    For each given DS (location_id), return whether it currently has stock —
    the inventory_available flag from its most recent blinkit_performance_detail
    row for this SKU within the last `days_back` days. A DS with no row at all
    in that window (nothing recent to go on) is treated as no-stock — the
    conservative default, since we can't confirm availability either way.
    location_ids must be a tuple (not list/set) so st.cache_data can hash it.
    """
    if not location_ids:
        return {}
    db = get_client()
    cutoff = (date.today() - timedelta(days=days_back)).isoformat()
    rows, offset, page_size = [], 0, 1000
    while True:
        batch = (
            db.table("blinkit_performance_detail")
            .select("location_id,data_date,inventory_available")
            .eq("sku_id", sku_id)
            .in_("location_id", list(location_ids))
            .gte("data_date", cutoff)
            .order("data_date", desc=True)
            .range(offset, offset + page_size - 1)
            .execute()
            .data
        )
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size

    latest: dict = {}
    for r in rows:
        loc_id = r["location_id"]
        if loc_id not in latest:  # first occurrence = latest, since sorted desc
            latest[loc_id] = bool(r["inventory_available"])
    return latest


@st.cache_data(ttl=300, show_spinner=False)
def _load_blinkit_fc_month1(month1_str: str) -> "pd.DataFrame":
    db = get_client()
    rows = (
        db.table('demand_forecasts')
        .select('sku_id,forecast_units')
        .eq('channel_id', 4)
        .eq('forecast_month', month1_str)
        .eq('model', 'VELOCITY_BASE')
        .execute().data
    )
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=['sku_id', 'forecast_units'])
    df['forecast_units'] = pd.to_numeric(df['forecast_units'], errors='coerce').fillna(0).astype(int)
    return df.rename(columns={'forecast_units': 'fc_m1'})


def tab_blinkit_deepdive(fdf: pd.DataFrame, net_mode: bool) -> None:
    """Blinkit Deepdive — stacked horizontal bar of dark store status per city, per SKU."""
    st.markdown("#### Darkstore Launch Status")

    # SKU selector
    sku_rows  = get_skus(active_only=True)
    sku_names = {r["sku_id"]: f"{r['sku_id']} — {r['name']}" for r in sku_rows}
    sku_ids   = list(sku_names.keys())
    if not sku_ids:
        st.warning("No SKUs found.")
        return

    selected_sku = st.selectbox(
        "Select SKU", sku_ids,
        format_func=lambda x: sku_names[x],
        key="blk_deepdive_sku",
    )

    df = _load_city_ds(selected_sku)
    if df.empty:
        st.info("No Blinkit dark store data found. Run the performance loader first.")
        return

    # ── Top 15 cities by total DS count ───────────────────────────────────────
    city_totals   = df.groupby("city").size().sort_values(ascending=False)
    top15_cities  = city_totals.head(15).index.tolist()
    df10          = df[df["city"].isin(top15_cities)].copy()

    # Pivot: rows = city (sorted by total, bottom-to-top for horizontal bar), cols = status counts
    pivot = (
        df10.groupby(["city", "status"])
        .size()
        .unstack(fill_value=0)
    )
    # Ensure all status columns exist
    for key, _, _ in _STATUS_META:
        if key not in pivot.columns:
            pivot[key] = 0
    # Sort cities: largest total at top (plotly horizontal bars are bottom-up)
    pivot = pivot.loc[top15_cities[::-1]]

    city_labels = [f"{c}  ({city_totals[c]})" for c in pivot.index]

    # ── Summary KPIs above the chart ──────────────────────────────────────────
    total      = len(df10)
    active_us  = int((df10["status"] == "active").sum())
    not_launch = int((df10["status"] == "launch_awaited").sum())
    low_sales  = int((df10["status"] == "sku_moved_out_low_sales").sum())
    choked     = int((df10["status"] == "ds_choked").sum())
    exited     = int((df10["status"] == "sku_city_exited").sum())
    recalled   = int((df10["status"] == "sku_recalled").sum())
    closed     = int((df10["status"] == "darkstore_closed").sum())

    c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
    c1.metric("Total DS",            total)
    c2.metric("DS closed",           closed)
    c3.metric("City exit",           exited)
    c4.metric("Recalled",            recalled)
    c5.metric("Inactive (low sale)", low_sales)
    c6.metric("Not launched",        not_launch)
    c7.metric("Choked",              choked)
    c8.metric("Active",              active_us)
    st.caption(
        "These counts and the bar chart below cover only the top 15 cities by dark store count. "
        "\"Launch Status\" further down covers **all** cities, so its totals will be larger."
    )

    fig = go.Figure()
    for key, label, color in _STATUS_META:
        counts = pivot[key].tolist()
        if sum(counts) == 0:
            continue
        fig.add_trace(go.Bar(
            name=label,
            x=counts,
            y=city_labels,
            orientation="h",
            marker_color=color,
            hovertemplate=f"<b>%{{y}}</b><br>{label}: %{{x}}<extra></extra>",
            text=[str(v) if v > 0 else "" for v in counts],
            textposition="inside",
            insidetextanchor="middle",
            textfont=dict(color="white", size=11),
        ))

    fig.update_layout(
        barmode="stack",
        height=480,
        margin=dict(l=10, r=20, t=10, b=10),
        xaxis=dict(title="Dark store count", tickfont=dict(size=11)),
        yaxis=dict(tickfont=dict(size=12)),
        legend=dict(
            orientation="h", yanchor="bottom", y=-0.30,
            xanchor="left", x=0, font=dict(size=11),
        ),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
    )
    st.plotly_chart(fig, use_container_width=True)

    # ── Launch Status ──────────────────────────────────────────────────────────
    st.markdown("#### Launch Status")
    st.caption(
        "Per-city launch state for this SKU, for expansion decisions — covers **all** cities "
        "(unlike the top-15-only counts and chart above)."
    )

    df_city = df[df["city"].notna() & (df["city"] != "")]
    active_rows          = df_city[df_city["status"] == "active"]
    launch_awaited_rows  = df_city[df_city["status"] == "launch_awaited"]

    active_ds_ids = tuple(sorted(active_rows["location_id"].unique().tolist()))
    stock_flags   = _load_latest_stock_flags(selected_sku, active_ds_ids)
    wh_names      = _load_all_wh_names()

    def _serving_wh(grp: pd.DataFrame) -> str:
        """Comma-separated, sorted list of distinct WH names serving this
        city's DS group — a city can legitimately sit under 2+ WHs."""
        whs = sorted({
            wh_names.get(pid, f"location_id={pid}")
            for pid in grp["parent_location_id"].dropna().unique()
        })
        return ", ".join(whs) if whs else "—"

    with_stock, no_stock = [], []
    for city, grp in active_rows.groupby("city"):
        loc_ids   = grp["location_id"].tolist()
        cnt       = len(loc_ids)
        has_stock = any(stock_flags.get(lid, False) for lid in loc_ids)
        row = (city, cnt, _serving_wh(grp))
        (with_stock if has_stock else no_stock).append(row)
    with_stock.sort(key=lambda x: -x[1])
    no_stock.sort(key=lambda x: -x[1])

    # Not launched: city has launch_awaited DS AND zero active DS for this SKU.
    # A city with ANY active DS counts as launched — even if some other DS
    # there are still pending launch — so "launched" always takes priority
    # over "not launched" for the same city.
    launched_cities   = set(active_rows["city"].unique().tolist())
    not_launched_pool = launch_awaited_rows[~launch_awaited_rows["city"].isin(launched_cities)]
    not_launched = sorted(
        ((city, len(grp), _serving_wh(grp)) for city, grp in not_launched_pool.groupby("city")),
        key=lambda x: -x[1],
    )

    def _city_count_df(rows: list[tuple]) -> pd.DataFrame:
        if not rows:
            return pd.DataFrame({"City": ["— none —"], "Serving WH": [""]})
        return pd.DataFrame({
            "City":       [f"{city} ({cnt})" for city, cnt, _ in rows],
            "Serving WH": [wh for _, _, wh in rows],
        })

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.markdown(f"**Cities Launched with stock** ({len(with_stock)})")
        st.dataframe(_city_count_df(with_stock), use_container_width=True, hide_index=True)
    with col_b:
        st.markdown(f"**Cities Launched but no stock** ({len(no_stock)})")
        st.dataframe(_city_count_df(no_stock), use_container_width=True, hide_index=True)
    with col_c:
        st.markdown(f"**Cities not Launched** ({len(not_launched)})")
        st.dataframe(_city_count_df(not_launched), use_container_width=True, hide_index=True)

    # ── Inventory Health table ─────────────────────────────────────────────────
    st.markdown("#### Inventory Health")
    wh_rows = _load_wh_summary(selected_sku)
    if wh_rows:
        wh_df = pd.DataFrame(wh_rows)

        def _health_pct(total_inv, target):
            if target > 0:
                return int(round(total_inv / target * 100))
            return None

        def _health_str(total_inv, target):
            v = _health_pct(total_inv, target)
            return f"{v}%" if v is not None else "N/A"

        display_rows = []
        for r in wh_rows:
            display_rows.append({
                "Warehouse":              r["wh_name"],
                "Active DS":              r["active_ds"],
                "Choked DS":              r["choked_ds_count"],
                "WH Stock (SOH+Planned)": r["wh_stock"],
                "DS Stock (SOH+In-transit)": r["ds_stock"],
                "Total Inventory":        r["total_inventory"],
                "Target Stock":           r["target_stock"],
                "Health":                 _health_str(r["total_inventory"], r["target_stock"]),
                "Starving DS":            r["oos_ds_count"],
            })

        total_inv = wh_df["total_inventory"].sum()
        total_tgt = wh_df["target_stock"].sum()
        display_rows.append({
            "Warehouse":              "TOTAL",
            "Active DS":              int(wh_df["active_ds"].sum()),
            "Choked DS":              int(wh_df["choked_ds_count"].sum()),
            "WH Stock (SOH+Planned)": int(wh_df["wh_stock"].sum()),
            "DS Stock (SOH+In-transit)": int(wh_df["ds_stock"].sum()),
            "Total Inventory":        int(total_inv),
            "Target Stock":           int(total_tgt),
            "Health":                 _health_str(total_inv, total_tgt),
            "Starving DS":            int(wh_df["oos_ds_count"].sum()),
        })

        inv_df   = pd.DataFrame(display_rows)
        is_total = [False] * (len(inv_df) - 1) + [True]

        def _health_color(val):
            if not isinstance(val, str) or val == "N/A":
                return ""
            try:
                v = float(val.replace("%", ""))
                if v < 70:
                    return "color: red; font-weight: bold"
                if v < 85:
                    return "color: darkorange; font-weight: bold"
                return "color: green; font-weight: bold"
            except ValueError:
                return ""

        st.dataframe(
            inv_df.style
            .map(_health_color, subset=["Health"])
            .apply(
                lambda row: ["font-weight: bold"] * len(row) if is_total[row.name] else [""] * len(row),
                axis=1,
            ),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("No warehouse inventory data. Run the inventory and performance loaders first.")

    # ── Data Quality Alerts ────────────────────────────────────────────────────
    st.markdown("#### Data Quality Alerts")
    st.caption("These alerts fire only when anomalies are detected — all clear means no output below.")

    try:
        db = get_client()
        plan_df = _get_replen_plan()

        # Trigger 1: DB active count > plan active count for this SKU
        # Signals DS that are marked active in DB but absent from current performance data
        db_active_rows = (
            db.table("blinkit_ds_sku_eligibility")
            .select("location_id")
            .eq("sku_id", selected_sku)
            .eq("status", "active")
            .execute()
            .data
        )
        db_active_count = len(db_active_rows)
        plan_active_count = (
            int(plan_df[plan_df["sku_id"] == selected_sku]["active_ds_count"].sum())
            if not plan_df.empty else 0
        )
        if not plan_df.empty and db_active_count > plan_active_count:
            discrepancy = db_active_count - plan_active_count
            st.warning(
                f"**Trigger 1 — Vanished-Y DS:** {db_active_count} DS marked `active` in DB "
                f"but only {plan_active_count} appear in current performance data. "
                f"**{discrepancy} DS may have been dropped from assessment while still active** "
                f"(went Y → absent without going through N). Raise with Blinkit."
            )

        # Trigger 2: Column Q data integrity guard
        # inventory_available=False but total_orders>0 should never happen — if it does,
        # Blinkit's Column Q has a data quality issue.
        t2_rows = (
            db.table("blinkit_performance_detail")
            .select("data_date, location_id")
            .eq("sku_id", selected_sku)
            .eq("inventory_available", False)
            .gt("total_orders", 0)
            .limit(50)
            .execute()
            .data
        )
        if t2_rows:
            st.warning(
                f"**Trigger 2 — Column Q integrity:** {len(t2_rows)} row(s) found where "
                f"`inventory_available=False` but `total_orders>0`. Orders were placed despite "
                f"Column Q flagging inventory as unavailable — possible Column Q data quality issue. "
                f"Raise with Blinkit."
            )

        # Trigger 3: Orphaned WH stock (risk flag, not a certainty — Blinkit gives
        # no DS-level inventory quantity anywhere; SOH is WH-aggregate only, and
        # performance detail only has a per-DS Y/N flag). Checked across ALL SKUs,
        # not just the one selected above, since a WH can be affected regardless
        # of which SKU tab is currently open.
        snap_date_rows = (
            db.table("blinkit_inventory_snapshots")
            .select("snapshot_date")
            .order("snapshot_date", desc=True)
            .limit(1)
            .execute()
            .data
        )
        if snap_date_rows:
            latest_snap = snap_date_rows[0]["snapshot_date"]
            soh_rows = (
                db.table("blinkit_inventory_snapshots")
                .select("location_id, total_sellable")
                .eq("snapshot_date", latest_snap)
                .execute()
                .data
            )
            wh_stock: dict = {}
            for r in soh_rows:
                wh_stock[r["location_id"]] = wh_stock.get(r["location_id"], 0) + (r.get("total_sellable") or 0)
            stocked_wh_ids = [wid for wid, qty in wh_stock.items() if qty > 0]

            if stocked_wh_ids:
                wh_names = {
                    r["location_id"]: r["name"]
                    for r in db.table("partner_locations").select("location_id, name")
                                .in_("location_id", stocked_wh_ids).execute().data
                }
                # Join via the DS network (parent_location_id), NOT partner_locations.name
                # vs. blinkit_performance_detail.serving_wh — those two are known to
                # disagree for several WHs (e.g. "Bengaluru B3" vs. "Bengaluru B3 - Feeder";
                # migration 024 renamed some but not all WH rows), which would otherwise
                # make this trigger fire false positives for perfectly healthy WHs.
                all_ds = (
                    db.table("partner_locations")
                    .select("location_id, parent_location_id")
                    .eq("location_type", "DARKSTORE")
                    .in_("parent_location_id", stocked_wh_ids)
                    .execute()
                    .data
                )
                ds_by_wh: dict = {}
                for d in all_ds:
                    ds_by_wh.setdefault(d["parent_location_id"], []).append(d["location_id"])

                cutoff = (date.today() - timedelta(days=15)).isoformat()
                orphaned = []
                for wid in stocked_wh_ids:
                    wh_name = wh_names.get(wid, f"location_id={wid}")
                    ds_ids = ds_by_wh.get(wid, [])
                    if not ds_ids:
                        orphaned.append((wh_name, wh_stock[wid]))
                        continue
                    recent = (
                        db.table("blinkit_performance_detail")
                        .select("data_date")
                        .in_("location_id", ds_ids)
                        .gte("data_date", cutoff)
                        .limit(1)
                        .execute()
                        .data
                    )
                    if not recent:
                        orphaned.append((wh_name, wh_stock[wid]))

                if orphaned:
                    orphaned_str = ", ".join(f"{name} ({int(qty)} units)" for name, qty in sorted(orphaned))
                    st.warning(
                        f"**Trigger 3 — Orphaned WH stock (risk flag):** {len(orphaned)} warehouse(s) "
                        f"show nonzero sellable stock in the latest SOH snapshot ({latest_snap}) but "
                        f"zero dark stores with current performance data in the last 15 days: "
                        f"{orphaned_str}. Blinkit gives no DS-level inventory quantity anywhere, so "
                        f"this can't be proven as stranded — but it's worth confirming directly with Blinkit."
                    )

    except Exception as exc:
        st.error(f"Data quality check failed: {exc}")

    # ── Warehouse Status ───────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Warehouse Status")

    try:
        wh_list  = _load_wh_list()
        all_ds   = _load_all_ds_raw()
        all_elig = _load_all_elig()
    except Exception as exc:
        st.error(f"Failed to load warehouse data: {exc}")
        return

    if not wh_list:
        st.info("No warehouse data found.")
        return

    sku_name_map = {r["sku_id"]: r["name"] for r in get_skus(active_only=False)}

    col_wh, col_city = st.columns([1, 1])
    with col_wh:
        sel_wh_name = st.selectbox("Warehouse", [r["name"] for r in wh_list], key="wh_status_wh")
    sel_wh_id = next(r["location_id"] for r in wh_list if r["name"] == sel_wh_name)

    wh_all_ds   = [d for d in all_ds if d["parent_location_id"] == sel_wh_id]
    cities_here = sorted({d["city"] for d in wh_all_ds if d.get("city")})
    with col_city:
        sel_city = st.selectbox("City", ["All"] + cities_here, key="wh_status_city")

    if sel_city == "All":
        city_ds_ids = {d["location_id"] for d in wh_all_ds}
    else:
        city_ds_ids = {d["location_id"] for d in wh_all_ds if d.get("city") == sel_city}

    # Count statuses per SKU for the selected DS set
    from collections import defaultdict
    sku_statuses: dict = defaultdict(lambda: defaultdict(int))
    for e in all_elig:
        if e["location_id"] in city_ds_ids:
            sku_statuses[e["sku_id"]][e["status"]] += 1

    # Replen plan for this WH (WH-level columns)
    try:
        plan_df = _get_replen_plan()
        wh_plan = plan_df[plan_df["wh_location_id"] == sel_wh_id] if not plan_df.empty else pd.DataFrame()
    except Exception:
        wh_plan = pd.DataFrame()

    wh_status_rows = []
    for sku_id in sorted(sku_statuses):
        s = sku_statuses[sku_id]
        if s.get("active", 0) == 0:
            continue
        pr = wh_plan[wh_plan["sku_id"] == sku_id] if not wh_plan.empty else pd.DataFrame()
        wh_status_rows.append({
            "SKU Code":        sku_id,
            "SKU Name":        sku_name_map.get(sku_id, sku_id),
            "Total DS":        sum(s.values()),
            "Closed":          s.get("darkstore_closed", 0) + s.get("sku_moved_out_low_sales", 0),
            "Exited/Recalled": s.get("sku_city_exited", 0) + s.get("sku_recalled", 0),
            "Choked":          s.get("ds_choked", 0),
            "Not Launched":    s.get("launch_awaited", 0),
            "Active":          s.get("active", 0),
            "Total ADS":       round(float(pr["total_ads"].iloc[0]), 2) if not pr.empty else 0.0,
            "Total Inventory": int(pr["effective_stock"].iloc[0]) if not pr.empty else 0,
            "Target Stock":    int(pr["target_stock"].iloc[0]) if not pr.empty else 0,
            "To Ship":         int(pr["units_to_ship"].iloc[0]) if not pr.empty else 0,
            "Invoice Value (₹)": int(pr["invoice_value"].iloc[0]) if not pr.empty else 0,
        })

    if wh_status_rows:
        _wh_cols = ["SKU Code", "SKU Name", "Active", "Not Launched",
                    "Total ADS", "Total Inventory", "Target Stock", "To Ship", "Invoice Value (₹)"]
        _wh_df = pd.DataFrame(wh_status_rows)[_wh_cols].copy()

        # Totals row
        _totals = {
            "SKU Code":          "",
            "SKU Name":          "Total",
            "Active":            _wh_df["Active"].sum(),
            "Not Launched":      _wh_df["Not Launched"].sum(),
            "Total ADS":         round(_wh_df["Total ADS"].sum(), 2),
            "Total Inventory":   _wh_df["Total Inventory"].sum(),
            "Target Stock":      _wh_df["Target Stock"].sum(),
            "To Ship":           _wh_df["To Ship"].sum(),
            "Invoice Value (₹)": _wh_df["Invoice Value (₹)"].sum(),
        }
        _wh_df = pd.concat([_wh_df, pd.DataFrame([_totals])], ignore_index=True)

        _last = len(_wh_df) - 1
        _styled = _wh_df.style.apply(
            lambda row: ["font-weight: bold" if row.name == _last else "" for _ in row],
            axis=1,
        )
        st.dataframe(
            _styled,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Total ADS": st.column_config.NumberColumn("Total ADS", format="%.2f"),
                "Invoice Value (₹)": st.column_config.NumberColumn(
                    "Invoice Value (₹)", format="%,.0f"
                ),
            },
        )
    else:
        st.info("No SKUs found for this WH / City selection.")

    # ── Replen plan refresh + download ────────────────────────────────────────
    import glob as _glob, subprocess as _sp, datetime as _dt

    _IST = _dt.timezone(_dt.timedelta(hours=5, minutes=30))
    _plan_age = ""
    if _REPLEN_PARQUET.exists():
        _mtime = _dt.datetime.fromtimestamp(
            _REPLEN_PARQUET.stat().st_mtime, tz=_dt.timezone.utc
        ).astimezone(_IST)
        _plan_age = f"Plan last generated: {_mtime.strftime('%d-%b-%Y %H:%M')} IST"

    _col_age, _col_btn = st.columns([3, 1])
    with _col_age:
        if _plan_age:
            st.caption(_plan_age)
        else:
            st.caption("No plan generated yet — click Refresh to generate.")
    with _col_btn:
        _refresh = st.button("🔄 Refresh & Download Replen Plan", key="replen_refresh_btn")

    if _refresh:
        with st.spinner("Generating replenishment plan (~30s)…"):
            _result = _sp.run(
                [sys.executable, "-m", "tcb.replenishment"],
                capture_output=True, text=True, encoding="utf-8",
                cwd=str(Path(".").resolve()),
            )
        if _result.returncode == 0:
            _replen_files = sorted(_glob.glob(
                "data/blinkit/auto/replenishment/replenishment_plan_*.xlsx"), reverse=True)
            if _replen_files:
                with open(_replen_files[0], "rb") as _fh:
                    st.session_state["_replen_dl_bytes"] = _fh.read()
                    st.session_state["_replen_dl_fname"] = os.path.basename(_replen_files[0])
            st.success("Plan refreshed!")
        else:
            st.error(f"Plan generation failed:\n{_result.stderr[-600:]}")

    # Show download button: use freshly generated bytes if available, else latest file on disk
    _dl_bytes = st.session_state.get("_replen_dl_bytes")
    _dl_fname = st.session_state.get("_replen_dl_fname")
    if not _dl_bytes:
        _replen_files = sorted(_glob.glob(
            "data/blinkit/auto/replenishment/replenishment_plan_*.xlsx"), reverse=True)
        if _replen_files:
            with open(_replen_files[0], "rb") as _fh:
                _dl_bytes = _fh.read()
            _dl_fname = os.path.basename(_replen_files[0])
    if _dl_bytes:
        st.download_button(
            "⬇ Download Replenishment Plan",
            _dl_bytes,
            file_name=_dl_fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_replen_excel",
        )

    # ── Warehouse — City Mapping ───────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Warehouse — City Mapping")

    ds_city_lkp = {d["location_id"]: (d.get("city") or "") for d in all_ds}

    html_rows = []
    for wh in wh_list:
        wid   = wh["location_id"]
        w_ds  = [d for d in all_ds if d["parent_location_id"] == wid]
        w_ids = {d["location_id"] for d in w_ds}

        # Closed DS: has darkstore_closed for ANY SKU AND no active row for ANY SKU.
        # Same invariant as propagate_darkstore_closed in blinkit_performance_loader.py:
        # if any SKU is active on a DS, the store has reopened → not closed.
        active_ds_ids_wh = {
            e["location_id"] for e in all_elig
            if e["location_id"] in w_ids and e["status"] == "active"
        }
        has_dark_closed_ids = {
            e["location_id"] for e in all_elig
            if e["location_id"] in w_ids and e["status"] == "darkstore_closed"
        }
        closed_ds_ids_wh   = has_dark_closed_ids - active_ds_ids_wh
        ds_closed_count    = len(closed_ds_ids_wh)

        # Skip WHs with no live DS presence at all (every DS closed, or none
        # exist) — same principle as Farukhnagar, which never appears here
        # because it's flagged is_active=False on the WH row itself. These
        # WHs (e.g. Chennai C5, fully wound down per Blinkit) don't have that
        # DB flag set yet, but showing them clutters the table with nothing
        # actionable — Total DS == DS Closed means zero open dark stores.
        if len(w_ds) == ds_closed_count:
            continue

        # Not-closed DS: every DS that is NOT in closed_ds_ids_wh.
        # Includes active, launch_awaited, sku_moved_out, etc. — any status
        # except a confirmed-closed store.  City (N) shows how these distribute.
        not_closed_ds_ids_wh = w_ids - closed_ds_ids_wh

        # All cities served by this WH (from DS records)
        all_wh_cities = sorted({
            ds_city_lkp.get(d["location_id"], "")
            for d in w_ds if ds_city_lkp.get(d["location_id"], "")
        })

        # Active SKU count per city (drives color) + WH-level SKUs launched
        city_active_skus: dict = defaultdict(set)
        wh_active_skus: set = set()
        for e in all_elig:
            if e["location_id"] in w_ids and e["status"] == "active":
                city = ds_city_lkp.get(e["location_id"], "")
                if city:
                    city_active_skus[city].add(e["sku_id"])
                wh_active_skus.add(e["sku_id"])

        # Not-closed DS count per city → shown as (N) in city list
        city_open_ds_count: dict = defaultdict(int)
        for d in w_ds:
            city = ds_city_lkp.get(d["location_id"], "")
            if city and d["location_id"] in not_closed_ds_ids_wh:
                city_open_ds_count[city] += 1

        city_parts = []
        for city in all_wh_cities:
            active_skus = len(city_active_skus.get(city, set()))
            open_ds     = city_open_ds_count.get(city, 0)
            if active_skus == 0:
                color = "#374151"  # black/dark — no SKUs launched in this city
            elif active_skus >= 10:
                color = "#22C55E"
            elif active_skus >= 5:
                color = "#F97316"
            else:
                color = "#EF4444"
            city_parts.append(
                f'<span style="color:{color};font-weight:500">{city}&nbsp;({open_ds})</span>'
            )

        html_rows.append(
            f"<tr style='border-bottom:1px solid #e5e7eb'>"
            f"<td style='padding:6px 14px;white-space:nowrap'>{wh['name']}</td>"
            f"<td style='padding:6px 14px;text-align:center'>{len(w_ds)}</td>"
            f"<td style='padding:6px 14px;text-align:center'>{ds_closed_count}</td>"
            f"<td style='padding:6px 14px;text-align:center'>{len(wh_active_skus)}</td>"
            f"<td style='padding:6px 14px;line-height:1.8'>{',&nbsp; '.join(city_parts) or '—'}</td>"
            f"</tr>"
        )

    mapping_html = (
        "<table style='border-collapse:collapse;width:100%;font-size:13px'>"
        "<thead><tr style='border-bottom:2px solid #9ca3af;background:rgba(0,0,0,0.03)'>"
        "<th style='padding:7px 14px;text-align:left'>Warehouse</th>"
        "<th style='padding:7px 14px;text-align:center'>Total DS</th>"
        "<th style='padding:7px 14px;text-align:center'>DS Closed</th>"
        "<th style='padding:7px 14px;text-align:center'>SKUs Launched</th>"
        "<th style='padding:7px 14px;text-align:left'>Cities Served (Open DS count)</th>"
        "</tr></thead>"
        f"<tbody>{''.join(html_rows)}</tbody></table>"
    )
    st.markdown(mapping_html, unsafe_allow_html=True)

    # ── City Performance ───────────────────────────────────────────────────────
    _CITY_NORM_BK = {
        "Delhi": "New Delhi", "Bangalore": "Bengaluru",
        "Gurugram": "Gurgaon", "Bengalore": "Bengaluru", "Bangaluru": "Bengaluru",
    }
    _gdf_bk = active_df(fdf, net_mode).copy()
    _gdf_bk["city"] = _gdf_bk["city"].str.strip().str.title().replace(_CITY_NORM_BK)
    _geo_bk = _gdf_bk[_gdf_bk["city"].notna() & _gdf_bk["state"].notna()]
    blinkit_geo = _geo_bk[_geo_bk["channel_name"] == "Blinkit"].copy()

    if not blinkit_geo.empty:
        st.markdown("---")
        st.markdown("#### City Performance")

        col_left, col_right = st.columns([1, 2])

        with col_left:
            bk_city = (
                blinkit_geo.groupby("city")
                .agg(
                    Orders=("order_id", "count"),
                    Units=("quantity", "sum"),
                    Revenue=("gross_value", "sum"),
                )
                .reset_index()
                .sort_values("Units", ascending=False)
            )
            bk_city["Revenue"] = bk_city["Revenue"].apply(fmt_inr)
            total_row = pd.DataFrame([{
                "city":    "TOTAL",
                "Orders":  bk_city["Orders"].sum(),
                "Units":   bk_city["Units"].sum(),
                "Revenue": fmt_inr(blinkit_geo["gross_value"].sum()),
            }])
            bk_display = pd.concat([bk_city, total_row], ignore_index=True).rename(columns={"city": "City"})

            def _bold_blinkit_total(styler):
                n = len(styler.data)
                return styler.set_properties(
                    subset=pd.IndexSlice[n - 1, :],
                    **{"font-weight": "bold", "border-top": "1px solid #ccc"},
                )

            st.dataframe(
                _bold_blinkit_total(bk_display.style),
                use_container_width=True,
                hide_index=True,
            )

        with col_right:
            bk_trend = (
                blinkit_geo.groupby(["month_dt", "city"])["quantity"].sum()
                .reset_index()
            )
            bk_trend["Month"] = bk_trend["month_dt"].dt.strftime("%b %Y")
            bk_trend = _project_col(bk_trend, "quantity", _cur_mult())
            month_order = bk_trend.sort_values("month_dt")["Month"].unique().tolist()

            # Merge NCR cities into one "NCR" line for THIS CHART ONLY — the
            # left-hand table keeps them separate, showing actual per-city
            # data. Collapsing 4 cities into 1 also frees up more top-10
            # slots for other cities to appear individually.
            _NCR_CITIES_BK = {"New Delhi", "Gurgaon", "Faridabad", "Ghaziabad"}
            bk_trend_ncr = bk_trend.copy()
            bk_trend_ncr["city"] = bk_trend_ncr["city"].where(
                ~bk_trend_ncr["city"].isin(_NCR_CITIES_BK), "NCR"
            )
            bk_trend_ncr = (
                bk_trend_ncr.groupby(["month_dt", "Month", "city"])["quantity"].sum()
                .reset_index()
            )

            # Top 10 (NCR now counts as one entrant) by total units; everything
            # else clubbed into a single "Others" line so the chart stays
            # readable regardless of how many cities are active.
            top10_cities = (
                bk_trend_ncr.groupby("city")["quantity"].sum()
                .sort_values(ascending=False)
                .head(10)
                .index.tolist()
            )
            bk_plot = bk_trend_ncr.copy()
            bk_plot["city"] = bk_plot["city"].where(bk_plot["city"].isin(top10_cities), "Others")
            bk_plot = (
                # month_dt (an actual datetime) must be the FIRST groupby key —
                # grouping by the "Month" string label first sorts rows
                # alphabetically ("Apr" < "Dec" < "Feb"...), and px.line() draws
                # each line by connecting dataframe rows in order, not by axis
                # position — that's what caused the zigzagging lines.
                bk_plot.groupby(["month_dt", "Month", "city"])["quantity"].sum()
                .reset_index()
                .sort_values("month_dt")
            )
            city_order = top10_cities + (["Others"] if "Others" in bk_plot["city"].values else [])

            fig_bk = px.line(
                bk_plot,
                x="Month", y="quantity", color="city",
                markers=True,
                labels={"quantity": "Units", "city": "City"},
                category_orders={"Month": month_order, "city": city_order},
                # Plotly's default qualitative palette only has 10 colors and
                # wraps around — with top-10 + "Others" (11 categories), the
                # 11th line silently reused the 1st city's color, making two
                # different cities' lines look like one tangled line. Dark24
                # has enough distinct colors to cover 11 without repeating.
                color_discrete_sequence=px.colors.qualitative.Dark24,
            )
            fig_bk.update_layout(
                xaxis_title=None, yaxis_title="Units",
                margin=dict(t=10),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            )
            st.caption("Monthly units — current month projected to full month")
            st.plotly_chart(fig_bk, use_container_width=True)


# ── Forecast tab ──────────────────────────────────────────────────────────────

_FC_CHANNEL_NAMES = {
    2: "Amazon", 3: "Amazon FBM", 4: "Blinkit", 5: "FnP",
    6: "First Cry", 7: "Peeko", 8: "Ozi", 9: "Kiddo", 10: "D2C",
}


@st.cache_data(ttl=60)
def _fc_load_display():
    return _fc_get_display()


@st.cache_data(ttl=60)
def _fc_load_channels(sku_id: str):
    return _fc_get_channels(sku_id)


def _fc_build_excel(display_df: pd.DataFrame) -> bytes:
    """Build 3-sheet forecast Excel."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb = openpyxl.Workbook()

    skus = get_skus()
    sku_names = {s["sku_id"]: s.get("sku_name", s["sku_id"]) for s in skus}

    forecast_dates = _fc_months(_FC_MONTHS)
    month_labels = [fm.strftime("%b %Y") for fm in forecast_dates]
    date_to_label = {fm: fm.strftime("%b %Y") for fm in forecast_dates}

    if display_df.empty:
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "No forecast data"
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    pivot = (
        display_df.pivot(index="sku_id", columns="forecast_month", values="forecast_units")
        .fillna(0).astype(int)
        .rename(columns=date_to_label)
    )
    locked_pivot = (
        display_df.pivot(index="sku_id", columns="forecast_month", values="is_user_locked")
        .fillna(False)
        .rename(columns=date_to_label)
    )

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    hdr_font = Font(bold=True)

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    headers = ["SKU", "Name"] + month_labels + ["6M Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font

    for r_idx, (sku_id, row) in enumerate(pivot.iterrows(), 2):
        ws.cell(row=r_idx, column=1, value=sku_id)
        ws.cell(row=r_idx, column=2, value=sku_names.get(sku_id, sku_id))
        total = 0
        for c_idx, m_label in enumerate(month_labels, 3):
            val = int(row.get(m_label, 0))
            total += val
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            is_locked = locked_pivot.loc[sku_id, m_label] if (sku_id in locked_pivot.index and m_label in locked_pivot.columns) else False
            if is_locked:
                cell.fill = green_fill
        ws.cell(row=r_idx, column=len(headers), value=total)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    for c_idx in range(3, len(headers) + 1):
        col_letter = ws.cell(row=1, column=c_idx).column_letter
        ws.column_dimensions[col_letter].width = 12

    # ── Sheet 2: Channel Breakdown ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Channel Breakdown")
    ws2.append(["SKU", "Channel", "Month", "Base Units", "Forecast Units", "Locked"])
    for cell in ws2[1]:
        cell.font = hdr_font

    for sku_id in sorted(display_df["sku_id"].unique()):
        ch_df = _fc_get_channels(sku_id)
        if ch_df.empty:
            continue
        for _, row in ch_df.iterrows():
            ws2.append([
                sku_id,
                _FC_CHANNEL_NAMES.get(int(row["channel_id"]), str(row["channel_id"])),
                str(row["forecast_month"]),
                int(row["base_units"]),
                int(row["forecast_units"]),
                "Yes" if row["is_user_locked"] else "No",
            ])

    for col in ("A", "B", "C"):
        ws2.column_dimensions[col].width = 14
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 16

    # ── Sheet 3: Assumptions ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Assumptions")
    ws3["A1"] = "Blinkit model: P75 of mean ADS across non-zero active DSes × planned DS count × 30 days"
    ws3["A2"] = "Growth model: P75 of MoM growth rates (capped at 35%) applied forward from last actuals"
    ws3["A3"] = "D2C: Manual input only — no engine-generated forecast"
    ws3["A4"] = "Locked cells (green) use USER_FINAL; unlocked cells use VELOCITY_BASE"
    ws3["A6"] = f"Generated: {date.today().isoformat()}"
    ws3.column_dimensions["A"].width = 90

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def tab_forecast():
    """Forecast tab: 6-month demand projection engine."""
    import subprocess

    # ── Top controls ──────────────────────────────────────────────────────────
    col_btn, col_status = st.columns([2, 4])
    with col_btn:
        regen = st.button("▶ Regenerate Base Forecast", key="fc_regen")
    with col_status:
        last_run = _fc_last_run()
        st.caption(f"Last run: {last_run}" if last_run else "No forecast generated yet — click Regenerate to start.")

    if regen:
        with st.spinner("Running forecast engine (~20s)..."):
            _r = subprocess.run(
                [sys.executable, "-m", "tcb.forecasting"],
                capture_output=True, text=True, encoding="utf-8",
                cwd=str(Path(".").resolve()),
            )
        if _r.returncode == 0:
            st.success("Forecast regenerated successfully!")
            st.cache_data.clear()
        else:
            st.error(f"Forecast failed:\n{_r.stderr[-800:]}")
        st.rerun()
        return

    _FC_EXCLUDE = {"TCB007", "TCB009", "TCB009_2"}

    # ── Load data ─────────────────────────────────────────────────────────────
    display_df = _fc_load_display()
    if display_df.empty:
        st.info("No forecast data found. Click ▶ Regenerate to generate the base forecast.")
        return

    active_sku_ids = {s["sku_id"] for s in get_skus()}
    display_df = display_df[
        display_df["sku_id"].isin(active_sku_ids) &
        ~display_df["sku_id"].isin(_FC_EXCLUDE)
    ]
    if display_df.empty:
        st.info("No active SKUs in forecast.")
        return

    forecast_dates = _fc_months(_FC_MONTHS)
    date_to_label  = {fm: fm.strftime("%b '%y") for fm in forecast_dates}
    month_labels   = [date_to_label[fm] for fm in forecast_dates]
    label_to_date  = {v: k for k, v in date_to_label.items()}

    # ── Coherency check: Forecast Month 1 vs Replenishment 30-day demand ─────────
    with st.expander("🔍 Forecast vs Replenishment Coherency — Month 1 Blinkit"):
        plan_df = _get_replen_plan()
        month1_str = forecast_dates[0].isoformat()
        if plan_df.empty:
            st.info("No replenishment plan found. Run the replenishment engine first.")
        else:
            replen_agg = (
                plan_df.groupby('sku_id', as_index=False)['total_ads']
                .sum()
                .assign(replen_30d=lambda d: (d['total_ads'] * 30).round(0).astype(int))
                [['sku_id', 'replen_30d']]
            )
            fc_df = _load_blinkit_fc_month1(month1_str)

            coh = replen_agg.merge(fc_df, on='sku_id', how='outer').fillna(0)
            coh['replen_30d'] = coh['replen_30d'].astype(int)
            coh['fc_m1']      = coh['fc_m1'].astype(int)
            replen_nonzero    = coh['replen_30d'].astype(float).replace(0.0, float('nan'))
            coh['delta_pct']  = ((coh['fc_m1'] - coh['replen_30d']) / replen_nonzero * 100).round(1)
            coh['status']     = coh['delta_pct'].apply(
                lambda d: '🔴 Under-call' if pd.notna(d) and d < -5 else '✅ OK'
            )

            under_calls = coh[coh['status'] == '🔴 Under-call']
            if not under_calls.empty:
                st.warning(f"⚠️ {len(under_calls)} SKU(s) where Blinkit forecast is >5% below replenishment demand — review before next ship.")
            else:
                st.success("All Blinkit SKUs: Forecast Month 1 ≥ Replenishment 30-day demand ✓")

            coh_display = coh.rename(columns={
                'sku_id':     'SKU',
                'replen_30d': 'Replen 30d',
                'fc_m1':      f'Forecast {forecast_dates[0].strftime("%b %y")} (Blinkit)',
                'delta_pct':  'Delta (%)',
                'status':     'Status',
            })
            st.dataframe(coh_display, use_container_width=True, hide_index=True)
            st.caption(
                "Replen 30d = SUM(total_ads) × 30 days — raw monthly demand signal, no coverage buffer. "
                "🔴 fires only when Forecast < Replen × 0.95 (over-forecast is always OK)."
            )

    # ── Build pivot ────────────────────────────────────────────────────────────
    pivot = (
        display_df.pivot(index="sku_id", columns="forecast_month", values="forecast_units")
        .fillna(0).astype(int)
        .rename(columns=date_to_label)
    )
    pivot_base = (
        display_df.pivot(index="sku_id", columns="forecast_month", values="base_units")
        .fillna(0).astype(int)
        .rename(columns=date_to_label)
    )
    pivot_locked = (
        display_df.pivot(index="sku_id", columns="forecast_month", values="is_user_locked")
        .fillna(False)
        .rename(columns=date_to_label)
    )

    avail_months = [m for m in month_labels if m in pivot.columns]
    for pv in (pivot, pivot_base, pivot_locked):
        for m in avail_months:
            if m not in pv.columns:
                pv[m] = 0

    # ── Last 3 months actuals ─────────────────────────────────────────────────
    today_d = date.today()
    cur_m   = date(today_d.year, today_d.month, 1)

    def _prev_month(d, n):
        month, year = d.month - n, d.year
        while month <= 0:
            month += 12
            year  -= 1
        return date(year, month, 1)

    hist_starts = [_prev_month(cur_m, 2), _prev_month(cur_m, 1), cur_m]
    hist_labels = [m.strftime("%b '%y") for m in hist_starts]

    raw_df = load_data()
    if not raw_df.empty:
        act_dict = (
            raw_df[
                (raw_df["order_date"].dt.date >= hist_starts[0]) &
                (raw_df["status"].isin(GROSS_STATUSES))
            ]
            .assign(mth=lambda x: x["order_date"].dt.to_period("M").dt.start_time.dt.date)
            .groupby(["sku_id", "mth"])["quantity"].sum()
            .to_dict()
        )
    else:
        act_dict = {}

    # ── Editable forecast grid ─────────────────────────────────────────────────
    grid_df = pivot[avail_months].reset_index().rename(columns={"sku_id": "SKU"})

    # Current month: pro-rate actuals to full-month projection
    days_elapsed   = today_d.day
    days_in_cur_m  = calendar.monthrange(today_d.year, today_d.month)[1]
    proj_factor    = days_in_cur_m / max(days_elapsed, 1)

    for lbl, mstart in zip(hist_labels, hist_starts):
        is_cur_month = (mstart == cur_m)
        pos = grid_df.columns.get_loc(avail_months[0])
        grid_df.insert(pos, lbl, grid_df["SKU"].apply(
            lambda s, ms=mstart, cur=is_cur_month: (
                round(act_dict.get((s, ms), 0) * proj_factor)
                if cur else int(act_dict.get((s, ms), 0))
            )
        ))

    def _locked_summary(sku_id: str) -> str:
        if sku_id not in pivot_locked.index:
            return ""
        row = pivot_locked.loc[sku_id]
        return ", ".join(m[:3] for m in avail_months if row.get(m, False)) or ""

    grid_df["Locked"] = grid_df["SKU"].apply(_locked_summary)

    # Append TOTAL row inside the same table
    num_cols  = hist_labels + avail_months
    total_row = {"SKU": "── TOTAL", **{c: int(grid_df[c].sum()) for c in num_cols}, "Locked": ""}
    grid_df   = pd.concat([grid_df, pd.DataFrame([total_row])], ignore_index=True)

    st.markdown("**Forecast — SKU × Month**")
    st.caption(
        f"Edit forecast cells to override, then click **Lock Edited Cells** to save. "
        f"Green = USER_FINAL locked. "
        f"*{hist_labels[0]}–{hist_labels[2]} = actual sales (read-only)*"
    )

    col_config = {
        "SKU": st.column_config.TextColumn("SKU", disabled=True, width="small"),
        **{lbl: st.column_config.NumberColumn(lbl, format="%d", disabled=True, width="small")
           for lbl in hist_labels},
        **{m: st.column_config.NumberColumn(m, format="%d", step=1, min_value=0, width="small")
           for m in avail_months},
        "Locked": st.column_config.TextColumn("Locked", disabled=True, width="small"),
    }

    edited_df = st.data_editor(
        grid_df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config=col_config,
        key="fc_editor",
    )

    # Base reference (collapsible)
    with st.expander("Base forecast (VELOCITY_BASE) — engine output before any user override"):
        base_view = pivot_base[avail_months].reset_index().rename(columns={"sku_id": "SKU"})
        st.dataframe(base_view, use_container_width=True, hide_index=True)

    # ── Lock / Reset ───────────────────────────────────────────────────────────
    st.markdown("---")
    col_lock, col_reset = st.columns(2)

    with col_lock:
        st.markdown("**Lock Edited Cells**")
        st.caption("Saves any edited value as USER_FINAL — distributed proportionally across channels.")
        if st.button("🔒 Lock Edited Cells", key="fc_lock"):
            changes: list[tuple] = []
            for _, row in edited_df.iterrows():
                sku = row["SKU"]
                if sku not in active_sku_ids or sku in _FC_EXCLUDE:
                    continue
                orig_row = grid_df[grid_df["SKU"] == sku]
                if orig_row.empty:
                    continue
                for m_label in avail_months:
                    orig_val = int(orig_row[m_label].values[0])
                    new_val  = int(row.get(m_label, orig_val) or 0)
                    if new_val != orig_val:
                        changes.append((sku, label_to_date[m_label], new_val))

            if changes:
                for sku, fm, val in changes:
                    _fc_lock(sku, fm, val)
                st.success(f"Locked {len(changes)} cell(s).")
                st.cache_data.clear()
                st.rerun()
            else:
                st.info("No changes detected — edit values in the grid above first.")

    with col_reset:
        st.markdown("**Reset to Base**")
        st.caption("Removes USER_FINAL — engine base forecast shows through on next load.")
        reset_sku = st.selectbox(
            "SKU to reset",
            options=["— select —"] + sorted(display_df["sku_id"].unique().tolist()),
            key="fc_reset_sku",
        )
        reset_months = st.multiselect(
            "Months (blank = all locked months)",
            options=avail_months,
            key="fc_reset_months",
        )
        if st.button("↩ Reset", key="fc_reset_btn") and reset_sku != "— select —":
            months_to_reset = [label_to_date[m] for m in (reset_months or avail_months)]
            for fm in months_to_reset:
                _fc_reset(reset_sku, fm)
            st.success(f"Reset {reset_sku} for {len(months_to_reset)} month(s).")
            st.cache_data.clear()
            st.rerun()

    # ── Channel breakdown ──────────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("Channel Breakdown — per SKU"):
        bd_sku = st.selectbox(
            "Select SKU",
            options=sorted(display_df["sku_id"].unique().tolist()),
            key="fc_bd_sku",
        )
        if bd_sku:
            ch_df = _fc_load_channels(bd_sku)
            if not ch_df.empty:
                ch_df = ch_df.copy()
                ch_df["Channel"] = ch_df["channel_id"].apply(
                    lambda cid: _FC_CHANNEL_NAMES.get(int(cid), str(cid))
                )
                ch_df["Month"] = ch_df["forecast_month"].apply(
                    lambda d: date_to_label.get(d, str(d))
                )
                ch_pivot = ch_df.pivot_table(
                    index="Channel", columns="Month",
                    values="forecast_units", aggfunc="sum", fill_value=0,
                )
                ch_pivot = ch_pivot[[m for m in avail_months if m in ch_pivot.columns]]
                st.dataframe(ch_pivot, use_container_width=True)
            else:
                st.info("No channel breakdown available for this SKU.")

    # ── Excel download ─────────────────────────────────────────────────────────
    st.markdown("---")
    if st.button("⬇ Prepare Forecast Excel", key="fc_excel_prep"):
        with st.spinner("Building Excel workbook..."):
            xl_bytes = _fc_build_excel(display_df)
        st.session_state["_fc_xl_bytes"] = xl_bytes
        st.session_state["_fc_xl_fname"] = f"tcb_forecast_{date.today().isoformat()}.xlsx"

    xl_bytes = st.session_state.get("_fc_xl_bytes")
    xl_fname = st.session_state.get("_fc_xl_fname", "tcb_forecast.xlsx")
    if xl_bytes:
        st.download_button(
            "⬇ Download Forecast Excel",
            data=xl_bytes,
            file_name=xl_fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="fc_dl_excel",
        )


if __name__ == "__main__":
    main()
