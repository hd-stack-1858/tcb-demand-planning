"""
Blinkit Shipment Invoice Generator.

Parses Blinkit RO Excel/PDF and generates a GST-compliant tax invoice Excel.
All invoice data comes from RO files + manual UI inputs — no DB values in invoice content.
DB is used only for cross-check alerts (price deviations, address mismatches).
"""

import io
import re
from datetime import date
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

SUPPLIER_NAME    = "Goodsense Trading India Private Limited"
SUPPLIER_ADDR1   = "First Floor, No. 2731, HAL 3rd Stage"
SUPPLIER_ADDR2   = "Bengaluru, Karnataka - 560075"
SUPPLIER_GSTIN   = "29AALCG8970F1Z0"

DELIVERY_PARTNER_GSTIN = {
    "Delhivery": "06AAPCS9575E1ZR",
    "Self Ship":  "29AALCG8970F1Z0",
}

# Known Blinkit WH consignee data — keyed by short WH name shown in the UI dropdown.
# All are BLINK COMMERCE PRIVATE LIMITED; GSTIN state code determines IGST vs CGST+SGST.
# Addresses match partner_locations DB exactly (sourced from Blinkit RO PDFs at onboarding).
# WH addresses are stable — Blinkit issues a new WH code when an address changes.
# gstin = "" means not yet confirmed (first RO from that WH will reveal it — update dict then).
WH_CONSIGNEE = {
    # ── Karnataka (GSTIN 29) ────────────────────────────────────────────────────
    "Bengaluru B3": {
        "gstin":   "29AAFCG9846E1Z7",
        "address": (
            "Sy. No. 89P, 90, 91, 92P, 93P, 94P, 95P, 96P, 100P, 101, 102P, and 103P,\n"
            "Iggalur Village, Attibele Hobli, Anekal Taluk,\n"
            "Bengaluru, Bengaluru Urban, Karnataka, 562106"
        ),
    },
    "Bengaluru B5": {
        "gstin":   "29AAFCG9846E1Z7",
        "address": (
            "Block C, Sy No 42/1, 42/3, 42/4, 42/5, 52, 53/1, 54, 55/3, 55/4\n"
            "Doddenahalli Village, Nangagudi Hobli, Hoskote Taluk,\n"
            "Bangaluru, 562114"
        ),
    },
    # ── Maharashtra (GSTIN 27) ──────────────────────────────────────────────────
    "Mumbai M10": {
        "gstin":   "27AAFCG9846E1ZB",
        "address": (
            "Survey No 122/1, 122/2, 123, 124 /1, 128/1, 128/4, 128/5, 133/4,\n"
            "Building No B-08-A, ESR Taloja, Industrial and Logistics Park,\n"
            "Village Usatane, Khoni Taloja Road, Taluka Ambernath,\n"
            "District Thane, Maharashtra - 421306"
        ),
    },
    "Pune P3": {
        "gstin":   "27AAFCG9846E1ZB",
        "address": "Gate No - 415, Nighoje, Khed Pune Nighoje - 410501",
    },
    "Nagpur N1": {
        "gstin":   "27AAFCG9846E1ZB",
        "address": (
            "Khasra No. 126/3/1, Mouza - Ketapar, NH 547E,\n"
            "Kalmeshwar - Gondkhairi Road, Tah - Kalmeshwar,\n"
            "Dist. Nagpur, Pin: - 441501"
        ),
    },
    # ── Haryana (GSTIN 06) ──────────────────────────────────────────────────────
    "Faridabad": {
        "gstin":   "06AAFCG9846E1ZF",
        "address": "20/3, Mathura Road, Faridabad, Faridabad, Haryana - 121006",
    },
    "Kundli": {
        "gstin":   "06AAFCG9846E1ZF",
        "address": (
            "Shed A, B and C, Musthil 41, Khasra No. 26//23, 41//1/2 min,\n"
            "2/2, 3/1/1 3/1/2, 3/2, 4/1, 4/2, 7, 8, 9 and 42//5/2,\n"
            "NH-1, village Kumaspur, Adjoining to Jurasik park inn,\n"
            "Sonipat, Haryana, 131001"
        ),
    },
    # ── Telangana (GSTIN 36) ────────────────────────────────────────────────────
    "Hyderabad H3": {
        "gstin":   "36AAFCG9846E2ZB",
        "address": (
            "Survey Nos. 664 & 665, admeasuring Acres 13-05 Guntas situated at\n"
            "Lalgadi Malakpet Village, Shameerpet Mandal, Ranga Reddy District,\n"
            "Presently Medchal-Malkajgiri District, Telangana State - 500101"
        ),
    },
    # ── Tamil Nadu (GSTIN 33) ───────────────────────────────────────────────────
    "Chennai C5": {
        "gstin":   "33AAFCG9846E1ZI",
        "address": (
            "Survey No. 55/1A2A2, Popular Auto Parts, Poochi Athipattu Village,\n"
            "Uthukottai Taluk, Chennai, Tiruvallur, Tamil Nadu, 600052"
        ),
    },
    "Coimbatore C1": {
        "gstin":   "33AAFCG9846E1ZI",
        "address": (
            "186/2A, 192/1A, 2, 3, 4, 193/1, 2A, 3A, 194/2D, 2E and 197/2,\n"
            "Appanaickenpatti village, Sulur TK, Coimbatore 641402"
        ),
    },
    # ── Uttar Pradesh (GSTIN 09) ────────────────────────────────────────────────
    "Noida N1": {
        "gstin":   "09AAFCG9846E1Z9",
        "address": (
            "Industrial Plot No. 2, Udyog Kendra, Ecotech -03,\n"
            "Greater Noida Dist - GB Nagar UP, 201306"
        ),
    },
    "Lucknow L4": {
        "gstin":   "09AAFCG9846E1Z9",
        "address": (
            "Khasra no 1479 Gulab khera, Kurauni Bijnaur Road,\n"
            "Near Banthra Thana, Kanpur Road Lucknow - 227101"
        ),
    },
    # ── Punjab (GSTIN unknown) ──────────────────────────────────────────────────
    "Rajpura R2": {
        "gstin":   "",
        "address": (
            "K No.174/222, Swift Logistics LLP, Village Alimajra Tehsil,\n"
            "Rajpura, 547/2MINA(0-8-1), Rajgarh, Patiala, Punjab, 140417"
        ),
    },
    # ── Rajasthan (GSTIN unknown) ───────────────────────────────────────────────
    "Jaipur J3": {
        "gstin":   "",
        "address": (
            "Jaipur DTA-01-12, Mahindra World City, Sanganer,\n"
            "DTA Phase-II, SEZ, Jaipur 302037"
        ),
    },
    # ── Bihar (GSTIN unknown) ───────────────────────────────────────────────────
    "Patna P1": {
        "gstin":   "",
        "address": (
            "VS LOGISTICS PARK located at Sabalpur, Deedarganj,\n"
            "Near Patna Samastipur Overbridge, Patna - 800009"
        ),
    },
    # ── Gujarat (GSTIN unknown) ─────────────────────────────────────────────────
    "Ahmedabad A2": {
        "gstin":   "",
        "address": (
            "Ashwika-4, Survey No 95, Behind Bharat Benz,\n"
            "Off Changodar-Bavla Highway (NH 8A),\n"
            "Village Vasna, Chacharvadi, Sanand, Ahmedabad, Gujarat, 382213"
        ),
    },
    # ── West Bengal (GSTIN unknown) ─────────────────────────────────────────────
    "Kolkata K4": {
        "gstin":   "",
        "address": (
            "Kalyani Expressway, Mohispota Ps near Muragachha,\n"
            "Crossing Ghola, Deshpriya Nagar, Bandipur,\n"
            "North Twenty Four Parganas, West Bengal, 700113"
        ),
    },
    "Kolkata K6": {
        "gstin":   "",
        "address": (
            "775, 777, 778, 783, 784, 1362, 1363, 785, 786, 787, 813,\n"
            "1361/1808, 600, 603, 1365, 1371, 1369, 1370 & Khatian Nos 8102 & 5177,\n"
            "Mouza Dankuni Bill & Gobra, JL Nos 83 & 92, Dankuni, 712311"
        ),
    },
    # ── Andhra Pradesh (GSTIN unknown) ──────────────────────────────────────────
    "Visakhapatnam V1": {
        "gstin":   "",
        "address": (
            "Land-8 admeasuring 16.47 Acres at IP-Extension-Gurrampalem,\n"
            "Gurrampalem Village, Pendurthi Mandal,\n"
            "Visakhapatnam - 531173"
        ),
    },
    # ── Assam (GSTIN unknown) ───────────────────────────────────────────────────
    "Guwahati G1": {
        "gstin":   "",
        "address": (
            "Khasra No. Dag No.1066, Patta No.34,\n"
            "Revenue Village Mikirpara Chakardoi, Mouza Ramcharani\n"
            "Azara Revenue Circle, Kamrup (Metro), Guwahati - 781017"
        ),
    },
}

# Map partial WH names from PDF "Delivered To" line → WH_CONSIGNEE key.
# Used to auto-select the WH dropdown when a PDF is uploaded.
_PDF_WH_NAME_TO_KEY: dict[str, str] = {
    "bengaluru b3":      "Bengaluru B3",
    "bengaluru b5":      "Bengaluru B5",
    "mumbai m10":        "Mumbai M10",
    "pune p3":           "Pune P3",
    "nagpur n1":         "Nagpur N1",
    "faridabad":         "Faridabad",
    "farukhnagar":       "Faridabad",   # Farukhnagar SR feeds Faridabad zone
    "kundli":            "Kundli",
    "hyderabad h3":      "Hyderabad H3",
    "chennai c5":        "Chennai C5",
    "coimbatore c1":     "Coimbatore C1",
    "noida n1":          "Noida N1",
    "lucknow l4":        "Lucknow L4",
    "rajpura r2":        "Rajpura R2",
    "jaipur j3":         "Jaipur J3",
    "patna p1":          "Patna P1",
    "ahmedabad a2":      "Ahmedabad A2",
    "kolkata k4":        "Kolkata K4",
    "kolkata k6":        "Kolkata K6",
    "visakhapatnam v1":  "Visakhapatnam V1",
    "guwahati g1":       "Guwahati G1",
}

CONSIGNEE_NAME = "BLINK COMMERCE PRIVATE LIMITED"


def wh_key_from_pdf_name(pdf_wh_name: str) -> str:
    """Map a raw 'BCPL - ...' WH name from the PDF to a WH_CONSIGNEE key. Returns '' if unknown."""
    lower = pdf_wh_name.lower()
    for fragment, key in _PDF_WH_NAME_TO_KEY.items():
        if fragment in lower:
            return key
    return ""

_FOOTER_DECLARATION = (
    "Declaration: We declare that this invoice shows the actual price of the goods "
    "described and that all particulars are true and correct."
)
_FOOTER_TC1 = "1. Once goods are handed over to the transporter/courier, the risk and title transfer to the buyer."
_FOOTER_TC2 = "2. Products must remain unused, unwashed, and in original packaging for any approved returns or replacements."

# ── Number to Words (Indian) ──────────────────────────────────────────────────

_ONES = [
    '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
    'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
    'Seventeen', 'Eighteen', 'Nineteen',
]
_TENS = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty',
         'Sixty', 'Seventy', 'Eighty', 'Ninety']


def _below_hundred(n: int) -> str:
    if n < 20:
        return _ONES[n]
    return _TENS[n // 10] + (' ' + _ONES[n % 10] if n % 10 else '')


def _below_thousand(n: int) -> str:
    if n < 100:
        return _below_hundred(n)
    return _ONES[n // 100] + ' Hundred' + (' ' + _below_hundred(n % 100) if n % 100 else '')


def amount_in_words(amount: float) -> str:
    """Convert rupee amount to Indian words: 'INR X Lakh Y Thousand Z only'."""
    rupees = int(amount)
    paise  = round((amount - rupees) * 100)
    if rupees == 0 and paise == 0:
        return 'INR Zero only'
    parts = []
    n = rupees
    crores = n // 10_000_000; n %= 10_000_000
    lakhs  = n // 100_000;    n %= 100_000
    thous  = n // 1_000;      n %= 1_000
    if crores: parts.append(_below_hundred(crores) + ' Crore')
    if lakhs:  parts.append(_below_hundred(lakhs)  + ' Lakh')
    if thous:  parts.append(_below_hundred(thous)  + ' Thousand')
    if n:      parts.append(_below_thousand(n))
    result = 'INR ' + ' '.join(parts)
    if paise:
        result += f' and {_below_hundred(paise)} Paise'
    return result + ' only'


# ── RO Excel Parser ───────────────────────────────────────────────────────────

def parse_ro_excel(file_bytes: bytes) -> dict:
    """
    Parse Blinkit RO .xlsx.
    Returns {line_items, net_amount, total_qty, item_count}.
    Line items contain all fields needed for invoice generation.
    """
    df = pd.read_excel(io.BytesIO(file_bytes), header=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    def _col(*names):
        return next((n for n in names if n in df.columns), None)

    c_item  = _col('Item Code')
    c_hsn   = _col('HSN Code', 'HSN')
    c_upc   = _col('Product UPC', 'UPC')
    c_desc  = _col('Product Description', 'Description')
    c_cgst  = _col('CGST %', 'CGST%')
    c_sgst  = _col('SGST %', 'SGST%')
    c_igst  = _col('IGST %', 'IGST%')
    c_land  = _col('Landing Rate', 'Landing Price')
    c_qty   = _col('Quantity', 'Qty')
    c_mrp   = _col('MRP')
    c_total = _col('Total Amount', 'Total')

    def _f(row, col, default=0.0):
        if col is None:
            return default
        try:
            return float(str(row[col]).replace(',', '').strip() or 0)
        except (ValueError, TypeError):
            return default

    def _s(row, col, default=''):
        if col is None:
            return default
        v = str(row[col] if row[col] is not None else '').strip()
        return '' if v in ('nan', 'None', '') else v

    def _pct(row, col):
        v = _f(row, col)
        return v / 100 if v > 1 else v  # normalise 2.5 → 0.025

    line_items = []
    for _, row in df.iterrows():
        item_raw = _s(row, c_item)
        try:
            int(float(item_raw))    # data rows have numeric item code
        except (ValueError, TypeError):
            continue
        line_items.append({
            'item_code':    item_raw,
            'hsn':          _s(row, c_hsn),
            'upc':          _s(row, c_upc),
            'description':  _s(row, c_desc),
            'cgst_pct':     _pct(row, c_cgst),
            'sgst_pct':     _pct(row, c_sgst),
            'igst_pct':     _pct(row, c_igst),
            'landing_rate': _f(row, c_land),
            'quantity':     int(_f(row, c_qty)),
            'mrp':          _f(row, c_mrp),
            'total_amount': _f(row, c_total),
        })

    return {
        'line_items': line_items,
        'net_amount': sum(li['total_amount'] for li in line_items),
        'total_qty':  sum(li['quantity'] for li in line_items),
        'item_count': len(line_items),
    }


# ── RO PDF Parser (Bill To / Ship To) ────────────────────────────────────────

def parse_ro_pdf(file_bytes: bytes) -> dict:
    """
    Extract consignee details from Blinkit RO PDF.
    Returns {name, address, gstin, ro_number} — empty strings on parse failure.

    PDF format (Blinkit RO):
      R.O. Number :<number>
      ...
      Delivered To :<NAME> GST No. :<GSTIN>
      <address line 1>
      <address line 2>
      ...
      # Item HSN ... (table header — marks end of address)
    """
    result = {'name': '', 'address': '', 'gstin': '', 'ro_number': '', 'wh_name': '', 'tax_type': ''}
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception:
        return result

    # RO number from "R.O. Number :<digits>"
    m = re.search(r'R\.O\.\s+Number\s*:\s*(\d+)', text, re.IGNORECASE)
    if m:
        result['ro_number'] = m.group(1)
    else:
        m = re.search(r'\b(\d{10,16})\b', text)
        if m:
            result['ro_number'] = m.group(1)

    # WH name — always "BCPL - <name>" on line 5 (index 5) of first page
    lines = text.split('\n')
    for line in lines[:10]:
        m = re.match(r'BCPL\s*[-–]\s*(.+)', line.strip(), re.IGNORECASE)
        if m:
            result['wh_name'] = m.group(1).strip()
            break

    # Two PDF formats seen in the wild:
    # Format A (BLR):       "Delivered To :<NAME> GST No. :<GSTIN>\n<addr line 1>\n..."
    # Format B (Faridabad): "Delivered :<NAME> GST No. :<GSTIN>\nTo <addr line 1>\n..."
    #
    # pdfplumber occasionally splits long lines across two rows even when the visual
    # layout is identical (font metrics, PDF stream order). To handle this, we first
    # look for "Delivered" and "BLINK" on the same line, then fall back to finding
    # the "Delivered :" anchor and scanning the next 3 lines for "BLINK" + GSTIN.
    lines = text.split('\n')

    delivered_line_idx = None
    blink_line_idx = None

    # Pass 1: both on same line (common case)
    for i, line in enumerate(lines):
        if re.search(r'Delivered\s*(?:To\s*)?:', line, re.IGNORECASE) \
                and 'BLINK' in line.upper():
            delivered_line_idx = i
            blink_line_idx = i
            break

    # Pass 2: "Delivered :" found but "BLINK" is on the next 1-3 lines (split-line case)
    if delivered_line_idx is None:
        for i, line in enumerate(lines):
            if re.search(r'Delivered\s*(?:To\s*)?:', line, re.IGNORECASE):
                for j in range(i, min(i + 4, len(lines))):
                    if 'BLINK' in lines[j].upper() and \
                            re.search(r'GST\s+No\.', lines[j], re.IGNORECASE):
                        delivered_line_idx = i
                        blink_line_idx = j
                        break
                if delivered_line_idx is not None:
                    break

    if delivered_line_idx is not None:
        blink_line = lines[blink_line_idx]

        # Extract GSTIN from the BLINK line
        gstin_m = re.search(r'GST\s+No\.\s*:\s*([A-Z0-9]{15})', blink_line, re.IGNORECASE)
        if gstin_m:
            result['gstin'] = gstin_m.group(1).strip()

        # Extract name from the BLINK line (between "Delivered [To] :" or start of line and "GST No.")
        name_m = re.search(
            r'(?:Delivered\s*(?:To\s*)?:\s*)?(.+?)\s+GST\s+No\.', blink_line, re.IGNORECASE
        )
        if name_m:
            candidate = name_m.group(1).strip()
            if 'BLINK' in candidate.upper():
                result['name'] = candidate

        # Collect address lines after the BLINK line
        addr_lines = []
        for line in lines[blink_line_idx + 1:]:
            line = line.strip()
            if not line:
                continue
            if re.match(r'^#\s+Item', line) or 'Total Quantity' in line:
                break
            # Format B: first address line starts with "To " — strip that prefix
            if line.startswith('To '):
                line = line[3:].strip()
            addr_lines.append(line)
            if len(addr_lines) >= 5:
                break

        # Strip trailing pdfplumber duplicate tokens (e.g. "Bengaluru 562114" at end)
        if addr_lines:
            last = re.sub(r'\s+\w[\w\s]*\d{6}\s*$', '', addr_lines[-1]).strip()
            addr_lines[-1] = last if last else addr_lines.pop() or ''

        # Clean double-comma artifacts and trailing "City PIN" duplicate lines
        addr_lines = [re.sub(r',\s*,', ',', l).strip().strip(',') for l in addr_lines if l.strip()]
        # Remove last line if it's just "CityName NNNNNN" (pdfplumber duplicate)
        if addr_lines and re.fullmatch(r'[\w\s]+ \d{6}', addr_lines[-1]):
            addr_lines.pop()
        result['address'] = '\n'.join(addr_lines)

    # Derive tax type from GSTIN state code (first 2 digits)
    # Karnataka = 29 → CGST+SGST (intra-state); all others → IGST
    if result['gstin']:
        state_code = result['gstin'][:2]
        result['tax_type'] = 'CGST_SGST' if state_code == '29' else 'IGST'

    return result


# ── DB Cross-check Helpers ────────────────────────────────────────────────────

def check_deviations(db, line_items: list[dict]) -> list[dict]:
    """
    Compare RO MRP and Landing Rate against DB values.
    Returns list of deviation dicts {item_code, description, ro_mrp, db_mrp,
    ro_landing, db_landing} — display only, no writes.
    """
    if not line_items:
        return []

    item_codes = [li['item_code'] for li in line_items]

    # Map Blinkit item code → sku_id
    rows = (db.table("sku_channel_ids")
              .select("sku_id, platform_pid_additional")
              .eq("channel_code", "BLK")
              .in_("platform_pid_additional", item_codes)
              .execute().data)
    code_to_sku = {r["platform_pid_additional"]: r["sku_id"] for r in rows}

    sku_ids = list(code_to_sku.values())
    if not sku_ids:
        return []

    # Latest pricing per SKU
    pricing_rows = (db.table("sku_pricing")
                      .select("sku_id, sp, mrp, effective_date")
                      .in_("sku_id", sku_ids)
                      .order("effective_date", desc=True)
                      .execute().data)
    latest_price = {}
    for p in pricing_rows:
        if p["sku_id"] not in latest_price:
            latest_price[p["sku_id"]] = p

    deviations = []
    for li in line_items:
        sku_id = code_to_sku.get(li['item_code'])
        if not sku_id or sku_id not in latest_price:
            continue
        p = latest_price[sku_id]
        db_mrp     = float(p.get('mrp') or 0)
        db_landing = float(p.get('sp')  or 0)
        ro_mrp     = li['mrp']
        ro_landing = li['landing_rate']

        if (db_mrp and abs(ro_mrp - db_mrp) > 0.01) or \
           (db_landing and abs(ro_landing - db_landing) > 0.01):
            deviations.append({
                'item_code':  li['item_code'],
                'description': li['description'],
                'ro_mrp':      ro_mrp,
                'db_mrp':      db_mrp,
                'ro_landing':  ro_landing,
                'db_landing':  db_landing,
            })
    return deviations


def check_wh_address(db, wh_name: str, dict_address: str) -> Optional[str]:
    """
    Cross-check the WH_CONSIGNEE dict address against the partner_locations DB record.
    Returns a warning string if a mismatch is detected, None if OK or no data to compare.
    Checks PIN code first (most reliable), then falls back to city/state.
    """
    if not dict_address or wh_name == "Manual":
        return None
    search_key = wh_name.split(" - ")[0]
    rows = (db.table("partner_locations")
              .select("address, city, state")
              .ilike("name", f"%{search_key}%")
              .execute().data)
    if not rows:
        return None
    db_row  = rows[0]
    db_addr = (db_row.get("address") or "").lower().strip()
    dict_addr_lower = dict_address.lower().strip()

    # Primary check: PIN code match
    db_pins   = set(re.findall(r"\b\d{6}\b", db_addr))
    dict_pins = set(re.findall(r"\b\d{6}\b", dict_addr_lower))
    if db_pins and dict_pins:
        if not db_pins.intersection(dict_pins):
            return (
                f"WH address mismatch — dict PIN {', '.join(dict_pins)} "
                f"≠ DB PIN {', '.join(db_pins)} for {wh_name}. "
                "Update WH_CONSIGNEE in tcb/blinkit_invoice.py if the WH address has changed."
            )
        return None  # PINs match — all good

    # Fallback: city/state check when addresses lack PINs
    db_state   = (db_row.get("state") or "").lower()
    dict_state_tokens = set(re.findall(r"[a-z]+", dict_addr_lower))
    db_state_tokens   = set(db_state.split())
    if db_state_tokens and not db_state_tokens.intersection(dict_state_tokens):
        return (
            f"WH state mismatch — DB shows state '{db_row.get('state')}' for {wh_name} "
            "but dict address doesn't mention it. "
            "Update WH_CONSIGNEE in tcb/blinkit_invoice.py if the WH address has changed."
        )
    return None


# ── Invoice Excel Generator ───────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _med_border():
    s = Side(style='medium')
    return Border(left=s, right=s, top=s, bottom=s)


def _set_outer_border(ws, r1, c1, r2, c2, style='medium'):
    """Set medium outer border on every perimeter cell, preserving interior borders."""
    med = Side(style=style)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            is_top    = (r == r1)
            is_bottom = (r == r2)
            is_left   = (c == c1)
            is_right  = (c == c2)
            if not (is_top or is_bottom or is_left or is_right):
                continue  # interior cell — skip
            cell = ws.cell(row=r, column=c)
            cur  = cell.border
            cell.border = Border(
                top    = med if is_top    else cur.top,
                bottom = med if is_bottom else cur.bottom,
                left   = med if is_left   else cur.left,
                right  = med if is_right  else cur.right,
            )


def _cell(ws, row, col, value='', bold=False, size=10, align='left',
          valign='center', wrap=False, border=True, bg=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Calibri', bold=bold, size=size)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        c.border = _thin_border()
    if bg:
        c.fill = PatternFill(fill_type='solid', fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    return c


def _merge(ws, r1, c1, r2, c2, value='', bold=False, size=10, align='left',
           valign='center', wrap=False, border=True, bg=None):
    ws.merge_cells(
        start_row=r1, start_column=c1, end_row=r2, end_column=c2
    )
    c = ws.cell(row=r1, column=c1, value=value)
    c.font      = Font(name='Calibri', bold=bold, size=size)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        c.border = _thin_border()
    if bg:
        c.fill = PatternFill(fill_type='solid', fgColor=bg)
    else:
        c.fill = PatternFill(fill_type='solid', fgColor='FFFFFF')   # explicit white
    return c


def generate_invoice_excel(
    line_items: list[dict],
    ro_number: str,
    invoice_no: str,
    invoice_date: date,
    delivery_date: date,
    delivery_partner: str,
    consignee_name: str,
    consignee_address: str,
    consignee_gstin: str,
    total_qty: int,
    item_count: int,
    tax_type: str,           # 'IGST' or 'CGST_SGST'
) -> bytes:
    """
    Generate Blinkit tax invoice Excel.
    tax_type: 'IGST' for inter-state, 'CGST_SGST' for intra-state Karnataka.
    Returns bytes of .xlsx file.
    """
    is_cgst = (tax_type == "CGST_SGST")

    # Column layout (matches template exactly):
    # A Sr.No | B Item ID | C Description | D UPC | E HSN | F MRP | G Box | H Units/Box
    # I Total Units | J CGST%/IGST% | [K SGST% CGST+SGST only] | ... | last = Total Amount
    N_COLS      = 14 if is_cgst else 13
    COL_TOTAL   = N_COLS
    COL_LANDING = N_COLS - 1
    COL_BASIC   = N_COLS - 2
    COL_UNITS   = 9    # I
    HALF        = 4    # left side ends at col D (matches template A-D / E-N split)

    # Summary section constants (from template)
    INWORDS_C2  = 7            # in-words left merge ends at col G
    SUMM_LBL_C1 = 8            # H — summary labels start
    SUMM_LBL_C2 = N_COLS - 1  # one before last col
    SUMM_VAL_C  = N_COLS       # last col = value

    # Exact colours from template
    HDR_BG   = "113B97"   # dark navy blue for column headers
    BSEC_BG  = "F2F3F4"   # very light grey for Bill To / Ship To header

    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"

    # Column widths — exact from template
    col_widths = [5.5, 10.0, 33.1, 14.5, 11.0, 6.5, 5.3, 8.1, 7.4, 8.5, 9.9, 9.9, 14.7, 12.1]
    for i, w in enumerate(col_widths[:N_COLS], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 65    # sized for Calibri 48pt
    _merge(ws, row, 1, row, N_COLS, 'TAX INVOICE', bold=True, size=48, align='center', valign='center', border=False)
    _set_outer_border(ws, row, 1, row, N_COLS)
    row += 1

    # ── Row 2: One big merged cell each side — no internal half-borders ──────────
    # Left: supplier + consignee summary | Right: all invoice meta
    import re as _re
    addr_lines_raw = [l.strip() for l in consignee_address.split('\n') if l.strip()]
    if addr_lines_raw:
        last = _re.sub(r'\s+[\w]+\s+\d{6}\s*$', '', addr_lines_raw[-1]).strip()
        addr_lines_raw[-1] = last if last else addr_lines_raw[-1]
    addr_lines = [l for l in addr_lines_raw if l][:6]

    left_text = '\n'.join([
        SUPPLIER_NAME,
        SUPPLIER_ADDR1,
        SUPPLIER_ADDR2,
        f'GST: {SUPPLIER_GSTIN}',
        '',
        'Consignee:',
        consignee_name,
        f'GST: {consignee_gstin}',
    ])

    right_text = '\n'.join([
        f'RO Number: {ro_number}',
        f'Invoice Number: {invoice_no}',
        f'Invoice Date: {invoice_date.strftime("%d-%m-%Y")}',
        '',
        f'Delivery Date: {delivery_date.strftime("%d-%m-%Y")}',
        f'Total Quantity: {total_qty}',
        f'Total Items: {item_count}',
        '',
        'Shipment Details:',
        f'Delivery Partner: {delivery_partner}',
        f'Transporter GST: {DELIVERY_PARTNER_GSTIN.get(delivery_partner, "")}',
    ])

    # Row 2 — one big block per side, medium outer border
    # Height: count lines × 14pt each side, take the larger, add 10pt padding
    _h2 = max(
        (left_text.count('\n') + 1) * 14,
        (right_text.count('\n') + 1) * 14,
        120
    ) + 10
    ws.row_dimensions[row].height = _h2
    _merge(ws, row, 1, row, HALF, left_text, size=10, wrap=True, valign='top', align='left', border=False)
    _merge(ws, row, HALF + 1, row, N_COLS, right_text, size=10, wrap=True, valign='top', align='left', border=False)
    _set_outer_border(ws, row, 1, row, HALF)
    _set_outer_border(ws, row, HALF + 1, row, N_COLS)
    row += 1

    # ── Bill To / Ship To — one big block per side ────────────────────────────
    # Build content for left and right blocks including all address lines
    billed_text = '\n'.join(filter(None, [
        'Billed To',
        '',
        consignee_name,
    ] + addr_lines + [
        f'GST: {consignee_gstin}',
    ]))
    ship_text = '\n'.join(filter(None, [
        'Ship To',
        '',
        consignee_name,
    ] + addr_lines + [
        f'GST: {consignee_gstin}',
    ]))

    _h3 = max((billed_text.count('\n') + 1) * 14, (ship_text.count('\n') + 1) * 14, 80) + 10
    ws.row_dimensions[row].height = _h3
    lc = _merge(ws, row, 1, row, HALF, billed_text, size=10, wrap=True, valign='top', align='left', border=False)
    lc.fill = PatternFill(fill_type='solid', fgColor=BSEC_BG)
    rc = _merge(ws, row, HALF + 1, row, N_COLS, ship_text, size=10, wrap=True, valign='top', align='left', border=False)
    rc.fill = PatternFill(fill_type='solid', fgColor=BSEC_BG)
    _set_outer_border(ws, row, 1, row, HALF)
    _set_outer_border(ws, row, HALF + 1, row, N_COLS)
    row += 1

    # ── Column Headers ────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    headers_cgst = [
        'Sr.\nNo', 'Item ID', 'Description of\nGoods', 'UPC', 'HSN\nNo',
        'MRP', 'Box', 'Units\nin Box', 'Total\nUnits',
        'CGST%', 'SGST%', 'Unit Basic\nPrice', 'Unit Landing\nPrice', 'Total\nAmount ₹'
    ]
    headers_igst = [
        'Sr.\nNo', 'Item ID', 'Description of\nGoods', 'UPC', 'HSN\nNo',
        'MRP', 'Box', 'Units\nin Box', 'Total\nUnits',
        'IGST%', 'Unit Basic\nPrice', 'Unit Landing\nPrice', 'Total\nAmount ₹'
    ]
    headers = headers_cgst if is_cgst else headers_igst
    for col_idx, hdr in enumerate(headers, start=1):
        c = _cell(ws, row, col_idx, hdr, bold=True, size=9,
                  align='center', valign='center', wrap=True, bg=HDR_BG)
        c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')  # white text on dark blue
    header_row = row
    row += 1

    # ── Line Items ────────────────────────────────────────────────────────────
    first_item_row = row
    for sr, li in enumerate(line_items, start=1):
        ws.row_dimensions[row].height = 14
        gst_rate = li['cgst_pct'] + li['sgst_pct'] + li['igst_pct']  # total GST as decimal
        unit_basic  = round(li['landing_rate'] / (1 + gst_rate), 2) if gst_rate else li['landing_rate']
        total_amt   = round(li['landing_rate'] * li['quantity'], 2)

        if is_cgst:
            row_vals = [
                sr, li['item_code'], li['description'], li['upc'], li['hsn'],
                li['mrp'], '', '',                              # Box, Units in Box blank
                li['quantity'],
                li['cgst_pct'],                                # CGST%
                li['sgst_pct'],                                # SGST%
                unit_basic, li['landing_rate'], total_amt,
            ]
            tax_num_cols = [10, 11]   # CGST%, SGST%
        else:
            row_vals = [
                sr, li['item_code'], li['description'], li['upc'], li['hsn'],
                li['mrp'], '', '',
                li['quantity'],
                li['igst_pct'],                                # IGST%
                unit_basic, li['landing_rate'], total_amt,
            ]
            tax_num_cols = [10]   # IGST%

        ws.row_dimensions[row].height = None   # auto-fit row height
        for col_idx, val in enumerate(row_vals, start=1):
            # Description column: left-aligned, wrap text
            if col_idx == 3:
                c = _cell(ws, row, col_idx, val, size=9, align='left', wrap=True)
            else:
                c = _cell(ws, row, col_idx, val, size=9, align='center')
            if col_idx in tax_num_cols:
                c.number_format = '0.0%'
            if col_idx in (COL_BASIC, COL_LANDING, COL_TOTAL, 6):
                c.number_format = '#,##0.00'
        row += 1

    last_item_row = row - 1

    # ── Totals Row ────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    _merge(ws, row, 1, row, COL_UNITS - 1, 'TOTAL', bold=True, size=10, align='center')
    _cell(ws, row, COL_UNITS, total_qty, bold=True, size=10, align='center')
    # Blank the tax/price columns in totals row
    for col_idx in range(COL_UNITS + 1, COL_TOTAL):
        _cell(ws, row, col_idx, '')
    total_row = row

    # Total Amount sum
    total_chargeable = sum(
        round(li['landing_rate'] * li['quantity'], 2) for li in line_items
    )
    _cell(ws, row, COL_TOTAL, total_chargeable, bold=True, size=10,
          align='right', num_fmt='#,##0.00')

    # Outer border around entire item table (headers + items + TOTAL row)
    _set_outer_border(ws, header_row, 1, total_row, N_COLS)
    row += 1

    # ── Summary Section ───────────────────────────────────────────────────────
    # Compute values
    gross_basic = sum(
        round(li['landing_rate'] / (1 + (li['cgst_pct'] + li['sgst_pct'] + li['igst_pct'])), 2)
        * li['quantity']
        for li in line_items
    )
    gross_basic = round(gross_basic, 2)

    # Use first item's rate (all TCB items same GST rate)
    li0 = line_items[0] if line_items else {}
    cgst_rate = li0.get('cgst_pct', 0)
    sgst_rate = li0.get('sgst_pct', 0)
    igst_rate = li0.get('igst_pct', 0)

    cgst_tax  = round(gross_basic * cgst_rate, 2)
    sgst_tax  = round(gross_basic * sgst_rate, 2)
    igst_tax  = round(gross_basic * igst_rate, 2)
    total_tax = cgst_tax + sgst_tax + igst_tax
    round_off = round(total_chargeable) - total_chargeable
    total_chargeable_rounded = round(total_chargeable)

    words_total = amount_in_words(total_chargeable_rounded)
    words_tax   = amount_in_words(total_tax)

    # ── Summary section ────────────────────────────────────────────────────────
    ROW_H = 15.75
    summ_start_row = row

    def _summ_label(r, text, bold=False):
        return _merge(ws, r, SUMM_LBL_C1, r, SUMM_LBL_C2,
                      text, bold=bold, size=9, wrap=True, valign='center', border=False)

    def _summ_val(r, value, bold=False):
        return _cell(ws, r, SUMM_VAL_C, value, bold=bold, size=9,
                     align='right', num_fmt='#,##0.00', border=False)

    # Row pair 1: in-words total (2-row merge left) | Gross Total + CGST/IGST right
    ws.row_dimensions[row].height = ROW_H
    _merge(ws, row, 1, row + 1, INWORDS_C2,
           f'Total Amount Chargeable (in words): {words_total}',
           size=9, wrap=True, valign='top', border=False)
    _summ_label(row, 'Gross Total (Basic Price)')
    _summ_val(row, gross_basic)
    row += 1

    ws.row_dimensions[row].height = ROW_H
    if is_cgst:
        _summ_label(row, 'CGST-Output Tax')
        _summ_val(row, cgst_tax)
        row += 1
        ws.row_dimensions[row].height = ROW_H
        _merge(ws, row, 1, row, INWORDS_C2, '', border=False)
        _summ_label(row, 'SGST-Output Tax')
        _summ_val(row, sgst_tax)
        row += 1
    else:
        _summ_label(row, 'IGST-Output Tax')
        _summ_val(row, igst_tax)
        row += 1

    # Row pair 2: in-words tax (2-row merge left) | Round Off + Total right
    ws.row_dimensions[row].height = ROW_H
    _merge(ws, row, 1, row + 1, INWORDS_C2,
           f'Total Tax Amount (in words): {words_tax}',
           size=9, wrap=True, valign='top', border=False)
    _summ_label(row, 'Less: Round Off')
    _summ_val(row, round_off)
    row += 1

    ws.row_dimensions[row].height = 16.15
    _summ_label(row, 'Total Amount Chargeable', bold=True)
    _summ_val(row, total_chargeable_rounded, bold=True)
    row += 1

    # Outer boxes around entire summary section
    summ_end_row = row - 1
    _set_outer_border(ws, summ_start_row, 1,          summ_end_row, INWORDS_C2)
    _set_outer_border(ws, summ_start_row, SUMM_LBL_C1, summ_end_row, N_COLS)

    # ── Footer — entire section wrapped in one outer border ──────────────────
    footer_start = row
    for text, bold in [
        ('Thanks for shopping with us.', False),
        (_FOOTER_DECLARATION, False),
        ('Terms & Conditions', True),
        (_FOOTER_TC1, False),
        (_FOOTER_TC2, False),
    ]:
        ws.row_dimensions[row].height = 14
        _merge(ws, row, 1, row, N_COLS, text, bold=bold, size=9, wrap=True, border=False)
        row += 1

    # Authorised Signatory label
    ws.row_dimensions[row].height = 13
    _merge(ws, row, 1, row, N_COLS, 'Authorised Signatory', bold=True, size=10, border=False)
    row += 1

    # Signature space (4 blank rows merged)
    sig_start = row
    for _ in range(4):
        ws.row_dimensions[row].height = 18
        row += 1
    _merge(ws, sig_start, 1, row - 1, N_COLS, '', border=False)

    # Single outer border around entire footer + signature area
    _set_outer_border(ws, footer_start, 1, row - 1, N_COLS)

    # ── Print setup — fit to one A4 landscape page ────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left   = 0.25
    ws.page_margins.right  = 0.25
    ws.page_margins.top    = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
