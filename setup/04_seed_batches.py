"""
Seed script: populates suppliers, item_batches, and inventory (loose stock in OWN_WH).
Run after 02_seed_data.py.
"""
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from tcb.db import get_client
import openpyxl

EXCEL = os.path.join(os.path.dirname(__file__), '..', 'master files', 'Batch wise-cost-inventory.xlsx')
db = get_client()

# ── Helpers ───────────────────────────────────────────────────
def upsert(table, rows, conflict_col):
    for row in rows:
        db.table(table).upsert(row, on_conflict=conflict_col).execute()
    print(f"  {table}: {len(rows)} rows upserted")

# ── Load Excel ────────────────────────────────────────────────
wb   = openpyxl.load_workbook(EXCEL, data_only=True)
ws   = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))[1:]  # skip header

# ── 1. SUPPLIERS ──────────────────────────────────────────────
print("Seeding suppliers...")
raw_suppliers = {str(r[6]).strip() for r in rows if r[6] and str(r[6]).strip() not in ('', 'Find out ?')}
supplier_rows = [{"name": s} for s in sorted(raw_suppliers)]
upsert("suppliers", supplier_rows, "name")

supplier_map = {r["name"]: r["supplier_id"]
                for r in db.table("suppliers").select("supplier_id,name").execute().data}

# ── 2. Item lookup ────────────────────────────────────────────
item_map = {r["item_code"]: r["item_id"]
            for r in db.table("items").select("item_id,item_code").execute().data}

# ── 3. OWN_WH channel_id ─────────────────────────────────────
own_wh_id = db.table("channels").select("channel_id").eq("code", "OWN_WH").single().execute().data["channel_id"]

# ── 4. ITEM BATCHES + INVENTORY ───────────────────────────────
print("Seeding item_batches and inventory...")
batch_rows     = []
inventory_rows = []
warnings       = []

for r in rows:
    if not r[0]: continue
    item_code    = str(r[0]).strip()
    receipt_date = r[3]
    cost         = float(r[4]) if r[4] else None
    qty_received = int(r[5]) if r[5] else 0
    supplier_raw = str(r[6]).strip() if r[6] else None
    qty_remaining = int(r[7]) if r[7] is not None else 0

    if item_code not in item_map:
        warnings.append(f"  WARNING: {item_code} not found in items table — skipped")
        continue

    if qty_remaining > qty_received:
        warnings.append(f"  WARNING: {item_code} {receipt_date.date()} — qty_remaining ({qty_remaining}) > qty_received ({qty_received})")

    item_id    = item_map[item_code]
    batch_code = receipt_date.strftime("%Y%m%d")
    supplier_id = supplier_map.get(supplier_raw)  # None for 'Find out ?' or missing

    batch_rows.append({
        "item_id":       item_id,
        "batch_code":    batch_code,
        "received_date": receipt_date.strftime("%Y-%m-%d"),
        "supplier_id":   supplier_id,
        "cost_per_unit": cost,
        "qty_received":  qty_received,
        "qty_remaining": qty_remaining,
        "is_current":    qty_remaining > 0,
    })

    if qty_remaining > 0:
        inventory_rows.append({
            "item_id":            item_id,
            "batch_code":         batch_code,
            "channel_id":         own_wh_id,
            "quantity_on_hand":   qty_remaining,
            "quantity_reserved":  0,
            "quantity_intransit": 0,
        })

upsert("item_batches", batch_rows, "item_id,batch_code")

# Build batch_id map for inventory FK
batch_id_map = {(r["item_id"], r["batch_code"]): r["batch_id"]
                for r in db.table("item_batches").select("batch_id,item_id,batch_code").execute().data}

for row in inventory_rows:
    row["batch_id"] = batch_id_map[(row["item_id"], row["batch_code"])]
    del row["batch_code"]

upsert("inventory", inventory_rows, "item_id,batch_id,channel_id")

if warnings:
    print("\nWarnings:")
    for w in warnings:
        print(w)
else:
    print("  No warnings.")

print(f"\nBatch seed complete.")
print(f"  {len(batch_rows)} batches loaded, {len(inventory_rows)} inventory rows (non-zero stock)")
print(f"  Items with 'Find out ?' supplier — supplier_id left NULL, update when confirmed.")
