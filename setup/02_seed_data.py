"""
Seed script: reads Item-Packaging-SKU mapping.xlsx and populates
suppliers, items, skus, sku_channel_ids, bom, channels tables.
Run after 01_create_tables.sql has been executed in Supabase.
"""
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from tcb.db import get_client
import openpyxl

EXCEL = os.path.join(os.path.dirname(__file__), '..', 'master files', 'Item-Packaging-SKU mapping.xlsx')
db = get_client()

def upsert(table, rows, conflict_col):
    for row in rows:
        db.table(table).upsert(row, on_conflict=conflict_col).execute()
    print(f"  {table}: {len(rows)} rows upserted")

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ── 1. CHANNELS ───────────────────────────────────────────────
print("Seeding channels...")
CHANNELS = [
    {"name": "Own Warehouse",  "code": "OWN_WH",   "business_model": None,         "fulfillment_from": "OWN_WH",             "is_location": True,  "is_active": True,  "legal_name": "Goodsense Trading India Private Limited"},
    {"name": "Amazon FBA",     "code": "AZ_FBA",   "business_model": "FBA",        "fulfillment_from": "FBA",                "is_location": False, "is_active": True,  "legal_name": "Amazon Seller Services Private Limited"},
    {"name": "Amazon FBM",     "code": "AZ_FBM",  "business_model": "DROP_SHIP",  "fulfillment_from": "OWN_WH",             "is_location": False, "is_active": True,  "legal_name": "Amazon Seller Services Private Limited",
     "notes": "Az SKU for TCB001=TCB002 and TCB002=TCB001 in platform (reversed at listing, cannot change)"},
    {"name": "Blinkit",        "code": "BLK",      "business_model": "SOR",        "fulfillment_from": "BLK_DARKSTORE",      "is_location": False, "is_active": True,  "legal_name": "Blink Commerce Private Limited"},
    {"name": "Ferns & Petals", "code": "FNP",      "business_model": "DROP_SHIP",  "fulfillment_from": "OWN_WH",             "is_location": False, "is_active": True,  "legal_name": "FNP E Retail Private Limited"},
    {"name": "First Cry",      "code": "FC",       "business_model": "DROP_SHIP",  "fulfillment_from": "OWN_WH",             "is_location": False, "is_active": True,  "legal_name": "Digital Age Retail Private Limited"},
    {"name": "Peeko",          "code": "PEEKO",    "business_model": "OUTRIGHT",   "fulfillment_from": "PEEKO_DARKSTORE",    "is_location": False, "is_active": True,  "legal_name": "Zippycubs Private Limited"},
    {"name": "Ozi",            "code": "OZI",      "business_model": "SOR",        "fulfillment_from": "OZI_DARKSTORE",      "is_location": False, "is_active": True,  "legal_name": "Ozi Technologies Private Limited"},
    {"name": "Kiddo",          "code": "KIDDO",    "business_model": "OUTRIGHT",   "fulfillment_from": "KIDDO_DARKSTORE",    "is_location": False, "is_active": True,  "legal_name": "Babyswift Private Limited"},
    {"name": "Own Website",    "code": "D2C",      "business_model": "DIRECT",     "fulfillment_from": "OWN_WH",             "is_location": False, "is_active": True,  "legal_name": "Goodsense Trading India Private Limited"},
    {"name": "Zepto",          "code": "ZEPTO",    "business_model": None,         "fulfillment_from": "ZEPTO_DARKSTORE",    "is_location": False, "is_active": False, "legal_name": "Zepto Private Limited"},
    {"name": "Instamart",      "code": "INSTAMART","business_model": None,         "fulfillment_from": "INSTAMART_DARKSTORE","is_location": False, "is_active": False, "legal_name": "Swiggy Instamart Private Limited"},
]
upsert("channels", CHANNELS, "code")

# ── 2. SKUs (from SKU List tab) ────────────────────────────────
print("Seeding SKUs...")
ws = wb["SKU List"]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

DISCONTINUED = {"TCB007": "Good quality photo frame could not be procured", "TCB009": "Replaced by TCB009_1 (2026 mugs)"}
NAME_OVERRIDE = {"TCB004": "Cosy Cub (5pcs)"}
HSN = {
    "TCB009": "69120010", "TCB009_1": "69120010",  # ceramic mugs
}
HSN_DEFAULT = "61112000"
GST = {}  # all SKUs are 5%

GS1_NOT_REGISTERED = {"TCB007", "TCB010", "TCB011", "TCB012"}

# Build Amazon UPC lookup first (used as fallback UPC for other channels)
az_upc_map = {}
for _row in rows[1:]:
    if not _row[0]: continue
    _col = dict(zip(headers, _row))
    _sid = str(_row[0]).strip()
    _v = _col.get("Amazon UPC")
    if _v and str(_v).strip() not in ("Not Listed", "None", ""):
        az_upc_map[_sid] = str(_v).strip()

skus_data = []
sku_channel_ids_data = []

for row in rows[1:]:
    if not row[0]: continue
    sku_id = str(row[0]).strip()
    name   = NAME_OVERRIDE.get(sku_id, str(row[1]).strip() if row[1] else "")
    skus_data.append({
        "sku_id": sku_id, "name": name,
        "hsn_code": HSN.get(sku_id, HSN_DEFAULT),
        "gst_pct": float(GST.get(sku_id, "5.0")),
        "is_discontinued": sku_id in DISCONTINUED,
        "discontinued_note": DISCONTINUED.get(sku_id),
    })
    # Platform IDs -- columns: Az SKU, Az ASIN, Blinkit PID, Blinkit Item ID, Blinkit UPC, FnP PID, FC Product ID, Peeko, Kiddo, Ozi
    col = dict(zip(headers, row))
    def val(k): v = col.get(k); return str(v) if v and str(v).strip() not in ("Not Listed","None","") else None

    def get_upc(upc_col):
        """Platform UPC → Amazon UPC fallback → 'Not registered with GS1' for unregistered SKUs."""
        if upc_col:
            upc = val(upc_col)
            if upc:
                return upc
        if sku_id in GS1_NOT_REGISTERED:
            return "Not registered with GS1"
        return az_upc_map.get(sku_id)

    # platform_pid = primary product ID on each platform
    #   Amazon → ASIN, Blinkit → PID, FnP → PID, FC → Product ID
    #   Peeko/Kiddo/Ozi → no dedicated PID, use our own sku_id as fallback
    # blinkit_item_id → Blinkit's second distinct ID (unique to that platform)
    # TCB006 note: Amazon UPC (8904492390033) ≠ Blinkit UPC (8904492390064) — listing mix-up, cannot change
    # platform_sku: BLK/FNP/FC have no separate SKU code on platform, use our own sku_id
    for ch_code, sku_col, pid_col, pid_additional_col, upc_col, use_sku_as_pid in [
        ("AZ_FBM","Az SKU",   "Az ASIN",             "Amazon FNSKU",    "Amazon UPC",   False),
        ("BLK",   None,       "Blinkit PID",          "Blinkit Item ID", "Blinkit UPC",  False),
        ("FNP",   None,       "FnP PID",              None,              None,           False),
        ("FC",    None,       "First Cry Product ID", None,              None,           False),
        ("PEEKO", "Peeko ID", None,                   None,              None,           True),
        ("KIDDO", "Kiddo ID", None,                   None,              None,           True),
        ("OZI",   "Ozi ID",   None,                   None,              None,           True),
    ]:
        platform_sku = val(sku_col) if sku_col else None
        if platform_sku is None:
            platform_sku = sku_id  # always fall back to our own sku_id
        platform_pid = sku_id if use_sku_as_pid else (val(pid_col) if pid_col else None) or "Not listed"
        pid_additional_raw = val(pid_additional_col) if pid_additional_col else None
        if pid_additional_raw is None:
            pid_additional = "Not listed" if platform_pid == "Not listed" else "NA"
        else:
            pid_additional = pid_additional_raw
        entry = {"sku_id": sku_id, "channel_code": ch_code,
                 "platform_sku":            platform_sku,
                 "platform_pid":            platform_pid,
                 "platform_pid_additional": pid_additional,
                 "platform_upc":            get_upc(upc_col)}
        sku_channel_ids_data.append(entry)

upsert("skus", skus_data, "sku_id")
upsert("sku_channel_ids", sku_channel_ids_data, "sku_id,channel_code")

# ── 3. ITEMS (from Item List tab) ─────────────────────────────
print("Seeding items...")
ITEM_UNIT = {
    "TCBP00010": "set",    # T-Shirt, Pajama, Cap, Mitten Little Looker
    "TCBP00015": "set",    # Romper, Cap, Mitten Cosy Cub
    "TCBP00020": "set",    # T-Shirt Growing Joy 0-6 month (6-piece milestone set)
    "TCBP00021": "set",    # T-Shirt Growing Joy 7-12 month (6-piece milestone set)
}

ws2 = wb["Item List"]
items_data = []
for row in list(ws2.iter_rows(values_only=True))[1:]:
    if not row[0]: continue
    item_code = str(row[0]).strip()
    name      = str(row[1]).strip() if row[1] else ""
    itype     = str(row[2]).strip().upper() if row[2] else "PRODUCT"
    itype     = "PACKAGING" if itype == "PACKAGING" else "PRODUCT"
    items_data.append({
        "item_code": item_code, "name": name, "item_type": itype,
        "unit": ITEM_UNIT.get(item_code, "piece"), "reorder_point": 0, "safety_stock": 0,
        "lead_time_days": 7, "moq": 1, "is_active": True
    })
upsert("items", items_data, "item_code")

# Delete items in DB that no longer exist in Excel
excel_item_codes = {r["item_code"] for r in items_data}
db_item_codes    = {r["item_code"] for r in db.table("items").select("item_code").execute().data}
stale_codes      = db_item_codes - excel_item_codes
if stale_codes:
    stale_ids = [r["item_id"] for r in db.table("items").select("item_id,item_code")
                 .in_("item_code", list(stale_codes)).execute().data]
    db.table("bom").delete().in_("item_id", stale_ids).execute()
    db.table("items").delete().in_("item_code", list(stale_codes)).execute()
    print(f"  items: deleted {len(stale_codes)} removed items: {sorted(stale_codes)}")

# ── 4. BOM (from SKU-Item Mapping tab) ────────────────────────
print("Seeding BOM...")
# Build item_code -> item_id map
item_map = {r["item_code"]: r["item_id"]
            for r in db.table("items").select("item_id,item_code").execute().data}

ws3 = wb["SKU-Item Mapping"]
bom_data = []
for row in list(ws3.iter_rows(values_only=True))[1:]:
    if not row[0] or not row[1]: continue
    sku_id    = str(row[0]).strip()
    item_code = str(row[1]).strip()
    if item_code not in item_map:
        print(f"  WARNING: item_code {item_code} not found, skipping BOM row")
        continue
    bom_data.append({"sku_id": sku_id, "item_id": item_map[item_code], "quantity_per_sku": 1})

upsert("bom", bom_data, "sku_id,item_id")

# Delete BOM rows in DB that no longer exist in Excel
excel_bom_pairs = {(r["sku_id"], r["item_id"]) for r in bom_data}
db_bom_rows     = db.table("bom").select("bom_id,sku_id,item_id").execute().data
stale_bom_ids   = [r["bom_id"] for r in db_bom_rows if (r["sku_id"], r["item_id"]) not in excel_bom_pairs]
if stale_bom_ids:
    db.table("bom").delete().in_("bom_id", stale_bom_ids).execute()
    print(f"  bom: deleted {len(stale_bom_ids)} stale rows")

print("\nSeed complete.")
print("Next steps:")
print("  1. Add suppliers (name, lead time, MOQ) via app or SQL")
print("  2. Add batch/cost data for all items")
print("  3. Add SKU pricing (MRP, SP per channel, COGS)")
print("  4. Enter current opening stock quantities")
