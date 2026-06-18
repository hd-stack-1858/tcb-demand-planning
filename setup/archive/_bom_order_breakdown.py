"""
One-off: BOM breakdown for manual order planning → exports Excel.
"""
import os, sys
from datetime import date
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if "TCB_ENV" in os.environ:
    del os.environ["TCB_ENV"]

from tcb.db import get_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORDER = {
    "TCB001": 400,
    "TCB002": 250,
    "TCB003": 175,
    "TCB004": 500,
    "TCB005": 900,
    "TCB006": 600,
    "TCB007": 0,
    "TCB008": 1400,
    "TCB009": 0,
    "TCB010": 200,
    "TCB011": 760,
    "TCB012": 500,
}

db = get_client()

bom_rows = db.table("bom").select("sku_id,item_id,quantity_per_sku").in_("sku_id", list(ORDER.keys())).execute().data
item_ids = list({r["item_id"] for r in bom_rows})
items_rows = db.table("items").select("item_id,item_code,name,item_type,moq,lead_time_days,latest_supplier_id").in_("item_id", item_ids).execute().data
items = {r["item_id"]: r for r in items_rows}

sup_ids = list({r["latest_supplier_id"] for r in items_rows if r["latest_supplier_id"]})
sup_rows = db.table("suppliers").select("supplier_id,name").in_("supplier_id", sup_ids).execute().data if sup_ids else []
suppliers = {r["supplier_id"]: r["name"] for r in sup_rows}

batches_rows = db.table("item_batches").select("item_id,cost_per_unit,received_date").in_("item_id", item_ids).order("received_date", desc=True).execute().data
latest_batch = {}
for b in batches_rows:
    iid = b["item_id"]
    if iid not in latest_batch:
        latest_batch[iid] = (b["cost_per_unit"], str(b["received_date"])[:10])

skus_rows = db.table("skus").select("sku_id,name").in_("sku_id", list(ORDER.keys())).execute().data
sku_names = {r["sku_id"]: r["name"] for r in skus_rows}

# ── Aggregate by item ─────────────────────────────────────────────────────────
item_totals: dict[int, dict] = {}

for row in bom_rows:
    iid = row["item_id"]
    sid = row["sku_id"]
    qty_per = int(row["quantity_per_sku"])
    order_qty = ORDER.get(sid, 0)
    needed = qty_per * order_qty

    if iid not in item_totals:
        item = items[iid]
        sup_id = item["latest_supplier_id"]
        cogs_info = latest_batch.get(iid)
        item_totals[iid] = {
            "item_code":        item["item_code"],
            "item_name":        item["name"],
            "item_type":        item["item_type"],
            "supplier":         suppliers.get(sup_id, "TBD") if sup_id else "TBD",
            "moq":              item["moq"],
            "lead_time_days":   item["lead_time_days"],
            "latest_cogs":      cogs_info[0] if cogs_info else None,
            "latest_batch_date":cogs_info[1] if cogs_info else "—",
            "sku_parts":        [],
            "total_needed":     0,
        }

    item_totals[iid]["sku_parts"].append(f"{sid}({needed})" if needed > 0 else f"{sid}(0)")
    item_totals[iid]["total_needed"] += needed

# Build rows, separated by type; exclude items with 0 total_needed
def build_rows(type_filter):
    rows = []
    for iid, d in sorted(item_totals.items(), key=lambda x: x[1]["item_code"]):
        if d["item_type"] != type_filter:
            continue
        if d["total_needed"] == 0:
            continue
        total = d["total_needed"]
        moq = d["moq"]
        order_qty = max(total, moq)
        moq_flag = f"↑ {order_qty}" if order_qty > total else ""
        used_in = "  +  ".join(p for p in d["sku_parts"] if not p.endswith("(0)"))
        rows.append([
            d["item_code"],
            d["item_name"],
            total,
            order_qty,
            moq_flag,
            used_in,
            d["supplier"],
            d["latest_cogs"],
            d["latest_batch_date"],
            moq,
            d["lead_time_days"],
        ])
    return rows

product_rows   = build_rows("PRODUCT")
packaging_rows = build_rows("PACKAGING")

HEADERS = [
    "Item Code", "Item Name",
    "Total Qty Needed", "Order Qty (after MOQ)", "MOQ Adjustment",
    "Used In (SKU × units)",
    "Supplier", "Latest COGS (₹)", "Latest Batch Date",
    "MOQ", "Lead Time (days)",
]

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Order Breakdown"

# Palette
C_HEADER    = "1F3864"   # dark navy
C_SECTION_P = "D9E1F2"   # light blue — products
C_SECTION_K = "E2EFDA"   # light green — packaging
C_FLAG      = "FFF2CC"   # amber — MOQ bump
C_FLAG_TEXT = "C55A11"   # dark orange
C_ALT       = "F7F7F7"   # alternating row

thin = Side(border_style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfont(bold=False, color="FFFFFF", sz=10):
    return Font(name="Calibri", bold=bold, color=color, size=sz)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_header_row(ws, row_num, labels, bg_color):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row_num, column=col, value=label)
        c.font = hfont(bold=True, color="FFFFFF" if bg_color == C_HEADER else "1F3864")
        c.fill = fill(bg_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

def write_section_header(ws, row_num, title, ncols, bg_color):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=ncols)
    c = ws.cell(row=row_num, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, color="1F3864", size=11)
    c.fill = fill(bg_color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, ncols + 1):
        ws.cell(row=row_num, column=col).border = border

current_row = 1

# Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
title_cell = ws.cell(row=1, column=1, value=f"TCB Manual Order — Item BOM Breakdown  ({date.today().strftime('%d %b %Y')})")
title_cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
title_cell.fill = fill(C_HEADER)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24
current_row = 2

# SKU order summary row (small)
sku_summary = "  |  ".join(f"{k}: {v}" for k, v in ORDER.items() if v > 0)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
sc = ws.cell(row=2, column=1, value=f"SKU Order Quantities:  {sku_summary}")
sc.font = Font(name="Calibri", size=9, color="595959", italic=True)
sc.fill = fill("F2F2F2")
sc.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 14
current_row = 3

def write_data_rows(ws, data_rows, start_row, alt_color):
    r = start_row
    for i, row in enumerate(data_rows):
        bg = alt_color if i % 2 == 0 else "FFFFFF"
        is_moq_bump = bool(row[4])  # MOQ Adjustment column not empty
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Calibri", size=10,
                          color=C_FLAG_TEXT if is_moq_bump else "000000",
                          bold=(col in (3, 4) and is_moq_bump))
            c.fill = fill(C_FLAG if is_moq_bump else bg)
            c.alignment = Alignment(
                horizontal="right" if col in (3, 4, 8, 10, 11) else "left",
                vertical="center", wrap_text=(col == 6)
            )
            c.border = border
        ws.row_dimensions[r].height = 15
        r += 1
    return r

# ── PRODUCT section ───────────────────────────────────────────────────────────
write_section_header(ws, current_row, "▶  PRODUCT Items", len(HEADERS), C_SECTION_P)
ws.row_dimensions[current_row].height = 18
current_row += 1

write_header_row(ws, current_row, HEADERS, C_HEADER)
ws.row_dimensions[current_row].height = 30
current_row += 1

current_row = write_data_rows(ws, product_rows, current_row, "EDF2F8")

# blank row
current_row += 1

# ── PACKAGING section ─────────────────────────────────────────────────────────
write_section_header(ws, current_row, "▶  PACKAGING Items", len(HEADERS), C_SECTION_K)
ws.row_dimensions[current_row].height = 18
current_row += 1

write_header_row(ws, current_row, HEADERS, C_HEADER)
ws.row_dimensions[current_row].height = 30
current_row += 1

current_row = write_data_rows(ws, packaging_rows, current_row, "EDF7EA")

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [14, 38, 16, 20, 16, 55, 20, 16, 16, 8, 16]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze pane below title+subtitle+first section header+header row
ws.freeze_panes = "A5"

out_path = f"data/blinkit/auto/replenishment/order_bom_breakdown_{date.today().strftime('%Y%m%d')}.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"Saved: {out_path}")
