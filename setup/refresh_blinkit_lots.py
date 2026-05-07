"""
Reconcile sku_cogs_lots (channel_id=4, Blinkit) against a fresh SOH snapshot.

Usage:
  python setup/refresh_blinkit_lots.py --file <path/to/InventoryData_DDMMMYYYY.xlsx>
                                        [--env dev|prod] [--dry-run]

What it does:
  1. Parses the Blinkit SOH Excel (same format as InventoryData_6May2026.xlsx).
     Column used: "Total sellable" (col index 10, 0-based).
  2. Resolves item_id → sku_id and facility_id → partner_location_id via DB.
  3. Reads the file generation timestamp (row 1, col 1).
  4. Adds any TRANSFER_OUT to Blinkit that occurred AFTER the file timestamp —
     those units are in transit / arrived after the snapshot.
  5. Computes expected_qty = file_total_sellable + post_snapshot_transfers.
  6. Compares against sum(qty_remaining) across all lots per (sku_id, location_id).
  7. Adjusts lots to reconcile:
       - Overstatement (DB > expected): reduce FIFO from oldest lots.
       - Understatement (DB < expected): add delta to the oldest lot.
  8. Prints a summary diff. With --dry-run, shows changes without writing.
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger(__name__)

IST = timezone(timedelta(hours=5, minutes=30))
BLK_CHANNEL_ID = 4
_TOTAL_SELLABLE_COL = 10   # 0-based index in the data row


def _parse_file_timestamp(ws) -> datetime:
    raw = ws.cell(1, 1).value  # "This sheet was generated at 2026-05-06 15:09:26"
    try:
        ts_str = str(raw).replace("This sheet was generated at ", "").strip()
        dt = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
        return dt.replace(tzinfo=IST)
    except Exception:
        logger.warning("Could not parse file timestamp from row 1: %r — using epoch", raw)
        return datetime.min.replace(tzinfo=timezone.utc)


def _load_soh(filepath: Path, fac_to_loc: dict[int, int],
              item_to_sku: dict[int, str]) -> dict[tuple[str, int], int]:
    """Parse SOH file. Returns {(sku_id, location_id): total_sellable}."""
    wb = openpyxl.load_workbook(str(filepath), read_only=False, data_only=True)
    ws = wb.active
    soh: dict[tuple[str, int], int] = {}
    skipped = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        item_id = row[0]
        fac_id  = row[5]
        tot_sell = row[_TOTAL_SELLABLE_COL]

        sku_id = item_to_sku.get(int(item_id))
        loc_id = fac_to_loc.get(int(fac_id))

        if sku_id is None or loc_id is None:
            skipped += 1
            continue

        qty = int(tot_sell) if tot_sell is not None else 0
        soh[(sku_id, loc_id)] = qty

    wb.close()
    if skipped:
        logger.warning("%d rows skipped (unmapped item_id or facility_id)", skipped)
    return soh


def _post_snapshot_transfers(db, file_ts: datetime) -> dict[tuple[str, int], int]:
    """
    Sum TRANSFER_OUT quantities to Blinkit that were created AFTER file_ts.
    Returns {(sku_id, partner_location_id): qty}.
    """
    ts_str = file_ts.isoformat()
    txns = (db.table("sku_inventory_transactions")
              .select("sku_id, quantity, partner_location_id")
              .eq("type", "TRANSFER_OUT")
              .eq("to_channel_id", BLK_CHANNEL_ID)
              .gt("created_at", ts_str)
              .execute().data)

    adds: dict[tuple[str, int], int] = {}
    for t in txns:
        loc_id = t.get("partner_location_id")
        if loc_id is None:
            continue
        key = (t["sku_id"], loc_id)
        adds[key] = adds.get(key, 0) + int(t["quantity"])

    if adds:
        logger.info("Found %d post-snapshot TRANSFER_OUT groups to add back", len(adds))
    return adds


def _get_current_lots(db) -> dict[tuple[str, int], list[dict]]:
    """
    Fetch all Blinkit lots from sku_cogs_lots.
    Returns {(sku_id, partner_location_id): [lot rows ordered by assembled_at]}.
    """
    rows = (db.table("sku_cogs_lots")
              .select("lot_id, sku_id, partner_location_id, assembled_at, unit_cogs, qty_remaining")
              .eq("channel_id", BLK_CHANNEL_ID)
              .order("assembled_at")
              .order("lot_id")
              .execute().data)

    lots: dict[tuple[str, int], list[dict]] = {}
    for r in rows:
        loc = r["partner_location_id"]
        if loc is None:
            continue
        key = (r["sku_id"], loc)
        lots.setdefault(key, []).append(r)
    return lots


def main(filepath: Path, env: str, dry_run: bool) -> None:
    os.environ["TCB_ENV"] = env
    from tcb.db import get_client
    db = get_client()

    # ── 1. Build lookup maps from DB ──────────────────────────────────────────
    pid_rows = (db.table("sku_channel_ids")
                  .select("sku_id, platform_pid_additional")
                  .eq("channel_code", "BLK")
                  .execute().data)
    item_to_sku: dict[int, str] = {}
    for r in pid_rows:
        pid = r.get("platform_pid_additional") or ""
        try:
            iid = int(pid)
        except (ValueError, TypeError):
            continue
        if iid not in item_to_sku or r["sku_id"] == "TCB009_1":
            item_to_sku[iid] = r["sku_id"]

    loc_rows = (db.table("partner_locations")
                  .select("location_id, external_id, name")
                  .eq("channel_id", BLK_CHANNEL_ID)
                  .execute().data)
    fac_to_loc: dict[int, int] = {}
    loc_name: dict[int, str] = {}
    for r in loc_rows:
        ext = r.get("external_id")
        if ext:
            try:
                fac_to_loc[int(ext)] = r["location_id"]
                loc_name[r["location_id"]] = r["name"]
            except (ValueError, TypeError):
                pass

    # ── 2. Parse SOH file ─────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(str(filepath), read_only=False, data_only=True)
    ws = wb.active
    file_ts = _parse_file_timestamp(ws)
    wb.close()

    logger.info("SOH file timestamp: %s", file_ts)
    soh = _load_soh(filepath, fac_to_loc, item_to_sku)

    # ── 3. Add post-snapshot TRANSFER_OUTs ───────────────────────────────────
    post_transfers = _post_snapshot_transfers(db, file_ts)
    expected: dict[tuple[str, int], int] = dict(soh)
    for key, add_qty in post_transfers.items():
        expected[key] = expected.get(key, 0) + add_qty
        logger.info("  Post-snapshot transfer: %s loc=%s +%s", key[0], key[1], add_qty)

    # ── 4. Compare against current DB lots ───────────────────────────────────
    current_lots = _get_current_lots(db)

    # Union of all keys
    all_keys = set(expected.keys()) | set(current_lots.keys())

    print(f"\n{'SKU':<12} {'Loc':<6} {'Location':<30} {'Expected':>9} {'DB_Total':>9} {'Delta':>7}")
    print("-" * 80)

    updates: list[dict] = []   # {lot_id, new_qty_remaining}
    no_lot_keys: list[tuple] = []

    for key in sorted(all_keys):
        sku_id, loc_id = key
        exp_qty = expected.get(key, 0)
        lots = current_lots.get(key, [])
        db_total = sum(l["qty_remaining"] for l in lots)
        delta = exp_qty - db_total
        loc_label = loc_name.get(loc_id, str(loc_id))
        marker = "" if delta == 0 else " <<<"

        print(f"{sku_id:<12} {loc_id:<6} {loc_label:<30} {exp_qty:>9} {db_total:>9} {delta:>7}{marker}")

        if delta == 0:
            continue

        if not lots:
            no_lot_keys.append(key)
            logger.warning("  No lots in DB for %s loc=%s (expected %s) — cannot adjust", sku_id, loc_id, exp_qty)
            continue

        if delta < 0:
            # Overstatement: reduce oldest lots first (FIFO)
            to_remove = abs(delta)
            for lot in lots:
                if to_remove <= 0:
                    break
                cut = min(int(lot["qty_remaining"]), to_remove)
                updates.append({"lot_id": lot["lot_id"],
                                 "new_qty": int(lot["qty_remaining"]) - cut})
                to_remove -= cut
        else:
            # Understatement: add to oldest lot
            oldest = lots[0]
            updates.append({"lot_id": oldest["lot_id"],
                             "new_qty": int(oldest["qty_remaining"]) + delta})

    print(f"\nTotal lots to update: {len(updates)}")
    if no_lot_keys:
        print(f"Locations with no lots (manual action needed): {no_lot_keys}")

    if dry_run or not updates:
        tag = " [DRY RUN]" if dry_run else ""
        print(f"\nNo changes written{tag}.")
        return

    for u in updates:
        db.table("sku_cogs_lots").update(
            {"qty_remaining": u["new_qty"]}
        ).eq("lot_id", u["lot_id"]).execute()

    print(f"\n{len(updates)} lot(s) updated in {env} DB.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Reconcile Blinkit sku_cogs_lots against a fresh SOH snapshot"
    )
    parser.add_argument("--file", required=True, help="Path to Blinkit SOH Excel file")
    parser.add_argument("--env",  choices=["dev", "prod"], default="prod")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    main(Path(args.file), args.env, args.dry_run)
